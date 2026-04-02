"""BOQ upload parsing helpers shared by SMU import flows."""

from __future__ import annotations

import csv
import io
import re
import zipfile
from typing import Any

from fastapi import HTTPException

from services.api.domain.boq.runtime.utxo import BoqItem
from services.api.domain.smu.templates import HEADER_ALIASES

ITEM_NO_PATTERN = re.compile(r"^\d{3}(?:-[0-9A-Za-z]+)*$")


def _to_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    text = _to_text(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        m = re.search(r"[-+]?\d+(?:\.\d+)?", text)
        if not m:
            return None
        try:
            return float(m.group(0))
        except Exception:
            return None


def _round4(value: float | None) -> float | None:
    if value is None:
        return None
    try:
        return round(float(value), 4)
    except Exception:
        return None


def _normalize_item_no(value: Any) -> str:
    text = _to_text(value).strip()
    if not text:
        return ""
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[./]", "-", text)
    text = re.sub(r"[()]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text


def _normalize_unit(value: Any) -> str:
    text = _to_text(value).strip()
    if not text:
        return ""
    norm = text.replace(" ", "").replace("　", "")
    unit_map = {
        "m3": "m3",
        "m³": "m3",
        "立方米": "m3",
        "方": "m3",
        "m2": "m2",
        "m²": "m2",
        "平方米": "m2",
        "平米": "m2",
        "延米": "m",
        "米": "m",
        "kg": "kg",
        "千克": "kg",
        "吨": "t",
        "t": "t",
    }
    return unit_map.get(norm.lower(), unit_map.get(norm, text)).strip()


def _fallback_approved_quantity(
    *,
    item_no: str,
    unit: str,
    approved_quantity: float | None,
    approved_amount: float | None,
    design_quantity: float | None,
    unit_price: float | None,
) -> tuple[float | None, bool]:
    if approved_quantity is not None:
        return approved_quantity, False
    unit_norm = _normalize_unit(unit)
    if unit_norm and unit_norm != "总额":
        return approved_quantity, False
    if approved_amount is None:
        return approved_quantity, False

    if unit_price and unit_price > 0:
        return _round4(approved_amount / unit_price), True
    if design_quantity and design_quantity > 0 and approved_amount > 0:
        return approved_amount, True
    if item_no.startswith("1") or item_no.startswith("2"):
        return approved_amount, True
    return approved_quantity, False


def _normalize_header(value: Any) -> str:
    text = _to_text(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = text.replace("_", "")
    text = text.replace("-", "")
    text = text.replace("/", "")
    return text


def _detect_header_map(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    normalized_aliases = {
        field: {_normalize_header(alias) for alias in aliases}
        for field, aliases in HEADER_ALIASES.items()
    }
    best_idx = -1
    best_score = -1
    best_map: dict[str, int] = {}
    best_found_fields: set[str] = set()

    def _header_match(key: str, alias: str) -> bool:
        if not alias:
            return False
        if key == alias:
            return True
        if key.startswith(alias) or key.endswith(alias):
            return True
        if alias in key:
            return True
        return False

    for idx, row in enumerate(rows):
        mapping: dict[str, int] = {}
        score = 0
        for col_idx, cell in enumerate(row):
            key = _normalize_header(cell)
            if not key:
                continue
            for field, aliases in normalized_aliases.items():
                if field in mapping:
                    continue
                if any(_header_match(key, alias) for alias in aliases):
                    mapping[field] = col_idx
                    score += 1
        if score > best_score and "item_no" in mapping and "name" in mapping:
            best_idx = idx
            best_score = score
            best_map = mapping
            best_found_fields = set(mapping.keys())
    if best_idx < 0:
        found = ",".join(sorted(best_found_fields)) if best_found_fields else "none"
        raise HTTPException(
            400,
            f"Failed to detect BOQ header row from upload file. required=item_no,name found={found}",
        )
    return best_idx, best_map


def _rows_from_csv_bytes(content: bytes) -> list[list[Any]]:
    if not content:
        return []
    text = ""
    for enc in ("utf-8-sig", "gb18030", "gbk"):
        try:
            text = content.decode(enc)
            break
        except Exception:
            continue
    if not text:
        text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def _rows_from_xlsx_bytes(content: bytes) -> tuple[list[list[Any]], str]:
    try:
        import openpyxl
    except Exception as exc:
        raise HTTPException(500, f"openpyxl not available: {exc}") from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except zipfile.BadZipFile:
        if content[:4] == b"\xD0\xCF\x11\xE0":
            # Fallback to legacy .xls parser when file is mis-labeled as .xlsx
            return _rows_from_xls_bytes(content)
        raise HTTPException(400, "Invalid .xlsx file. Please re-save as .xlsx or .csv.")
    except Exception as exc:
        raise HTTPException(400, f"Failed to read .xlsx file: {exc}")
    best_rows: list[list[Any]] = []
    best_sheet = wb.sheetnames[0]
    best_score = -1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) >= 300000:
                break
        if not rows:
            continue
        score = -1
        try:
            _, colmap = _detect_header_map(rows[:120])
            score = len(colmap)
        except HTTPException:
            score = -1
        if score > best_score:
            best_rows = rows
            best_sheet = sheet_name
            best_score = score
    if not best_rows:
        return [], best_sheet
    return best_rows, best_sheet


def _rows_from_xls_bytes(content: bytes) -> tuple[list[list[Any]], str]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise HTTPException(
            400,
            f".xls parser missing (xlrd). Please convert to CSV/XLSX or install xlrd in services/api env. detail={exc}",
        ) from exc
    wb = xlrd.open_workbook(file_contents=content)
    best_rows: list[list[Any]] = []
    best_sheet = wb.sheet_by_index(0).name if wb.nsheets > 0 else "sheet0"
    best_score = -1

    for i in range(wb.nsheets):
        sheet = wb.sheet_by_index(i)
        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
            if len(rows) >= 300000:
                break
        if not rows:
            continue
        score = -1
        try:
            _, colmap = _detect_header_map(rows[:120])
            score = len(colmap)
        except HTTPException:
            score = -1
        if score > best_score:
            best_rows = rows
            best_sheet = _to_text(getattr(sheet, "name", "")).strip() or best_sheet
            best_score = score
    if not best_rows:
        return [], best_sheet
    return best_rows, best_sheet


def parse_boq_upload(file_name: str, content: bytes) -> list[BoqItem]:
    name = _to_text(file_name).strip()
    lower = name.lower()
    header_sig = content[:4] if content else b""
    is_ole2 = header_sig == b"\xD0\xCF\x11\xE0"
    is_zip = header_sig == b"\x50\x4B\x03\x04"

    if lower.endswith(".csv"):
        rows = _rows_from_csv_bytes(content)
        source_sheet = "csv"
    elif lower.endswith(".xlsx"):
        if is_ole2:
            rows, source_sheet = _rows_from_xls_bytes(content)
        else:
            rows, source_sheet = _rows_from_xlsx_bytes(content)
        source_sheet = _to_text(source_sheet).strip() or "xlsx"
    elif lower.endswith(".xls"):
        rows, source_sheet = _rows_from_xls_bytes(content)
        source_sheet = _to_text(source_sheet).strip() or "xls"
    elif is_ole2:
        rows, source_sheet = _rows_from_xls_bytes(content)
        source_sheet = _to_text(source_sheet).strip() or "xls"
    elif is_zip:
        rows, source_sheet = _rows_from_xlsx_bytes(content)
        source_sheet = _to_text(source_sheet).strip() or "xlsx"
    else:
        rows = _rows_from_csv_bytes(content)
        source_sheet = "unknown"

    if not rows:
        raise HTTPException(400, "Upload file is empty.")

    probe = rows[:120]
    header_row, colmap = _detect_header_map(probe)

    out: list[BoqItem] = []
    for row_idx in range(header_row + 1, len(rows)):
        row = rows[row_idx]
        item_no_raw = _to_text(row[colmap["item_no"]] if colmap["item_no"] < len(row) else "").strip()
        item_no = _normalize_item_no(item_no_raw)
        if not item_no or not ITEM_NO_PATTERN.match(item_no):
            continue
        name = _to_text(row[colmap["name"]] if colmap["name"] < len(row) else "").strip()
        unit = _normalize_unit(row[colmap.get("unit", -1)] if colmap.get("unit", -1) < len(row) and colmap.get("unit", -1) >= 0 else "")
        division = _to_text(row[colmap.get("division", -1)] if colmap.get("division", -1) < len(row) and colmap.get("division", -1) >= 0 else "").strip()
        subdivision = _to_text(row[colmap.get("subdivision", -1)] if colmap.get("subdivision", -1) < len(row) and colmap.get("subdivision", -1) >= 0 else "").strip()
        hierarchy_raw = _to_text(row[colmap.get("hierarchy", -1)] if colmap.get("hierarchy", -1) < len(row) and colmap.get("hierarchy", -1) >= 0 else "").strip()

        dq_raw = _to_text(row[colmap.get("design_quantity", -1)] if colmap.get("design_quantity", -1) < len(row) and colmap.get("design_quantity", -1) >= 0 else "").strip()
        up_raw = _to_text(row[colmap.get("unit_price", -1)] if colmap.get("unit_price", -1) < len(row) and colmap.get("unit_price", -1) >= 0 else "").strip()
        aq_raw = _to_text(row[colmap.get("approved_quantity", -1)] if colmap.get("approved_quantity", -1) < len(row) and colmap.get("approved_quantity", -1) >= 0 else "").strip()
        aa_raw = _to_text(row[colmap.get("approved_amount", -1)] if colmap.get("approved_amount", -1) < len(row) and colmap.get("approved_amount", -1) >= 0 else "").strip()
        remark = _to_text(row[colmap.get("remark", -1)] if colmap.get("remark", -1) < len(row) and colmap.get("remark", -1) >= 0 else "").strip()

        design_quantity = _round4(_to_float(dq_raw))
        unit_price = _round4(_to_float(up_raw))
        approved_quantity = _round4(_to_float(aq_raw))
        approved_amount = _round4(_to_float(aa_raw))
        approved_quantity, used_amount = _fallback_approved_quantity(
            item_no=item_no,
            unit=unit,
            approved_quantity=approved_quantity,
            approved_amount=approved_amount,
            design_quantity=design_quantity,
            unit_price=unit_price,
        )
        if used_amount:
            aq_raw = aa_raw
            if not unit:
                unit = "总额"

        out.append(
            BoqItem(
                item_no=item_no,
                name=name,
                unit=unit,
                division=division,
                subdivision=subdivision,
                hierarchy_raw=hierarchy_raw,
                design_quantity=design_quantity,
                design_quantity_raw=dq_raw,
                unit_price=unit_price,
                unit_price_raw=up_raw,
                approved_quantity=approved_quantity,
                approved_quantity_raw=aq_raw,
                remark=remark,
                row_index=row_idx + 1,
                sheet_name=source_sheet,
            )
        )
    if not out:
        raise HTTPException(400, "No BOQ item rows parsed from upload file.")
    return out


__all__ = ["parse_boq_upload"]
