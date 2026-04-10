"""NormRef logic scaffold and Tab-to-Peg runtime utilities.

This module provides BOQPeg-facing helpers for:
- Bootstrapping NormRef L0/L1/L2 scaffold.
- Converting quality tables into executable protocol blocks.
"""

from __future__ import annotations

from datetime import UTC, datetime
import csv
import hashlib
import io
import json
from pathlib import Path
import re
from typing import Any

from fastapi import HTTPException

from services.api.domain.projects.gitpeg_sdk import register_uri
from services.api.domain.specir.runtime.registry import ensure_specir_object
from services.api.domain.utxo.integrations import ProofUTXOEngine

try:  # optional dependency
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:  # optional dependency
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None

NORMREF_ROOT_URI = "v://normref.com"
NORMREF_CORE_URI = "v://normref.com/core@v1"
NORMREF_HIGHWAY_URI = "v://normref.com/construction/highway@v1"
NORMREF_SCHEMA_QC_V1_URI = "v://normref.com/schema/qc-v1"
NORMREF_QC_GENERAL_TEMPLATE_URI = "v://normref.com/qc/template/general-quality-inspection@v1"
NORMREF_SEED_REBAR_PROTOCOL_URI = "v://normref.com/qc/rebar-processing@v1"
NORMREF_CONCRETE_QC_PROTOCOL_URI = "v://normref.com/qc/concrete-compressive-test@v1"
NORMREF_PILE_QC_PROTOCOL_URI = "v://normref.com/qc/pile-foundation@v1"
NORMREF_RAFT_SPU_URI = "v://normref.com/spu/raft-foundation@v1"
NORMREF_RAFT_QC_PROTOCOL_URI = "v://normref.com/qc/raft-foundation@v1"
NORMREF_PROMPT_T2P_URI = "v://normref.com/prompt/tab-to-peg-engine@v1"


def _to_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _to_bool(value: Any, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    text = _to_text(value).strip().lower()
    if not text:
        return bool(default)
    return text in {"1", "true", "yes", "on"}


def _as_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    return {}


def _clean_number(value: float) -> float | int:
    if float(value).is_integer():
        return int(value)
    return float(value)


def _to_number(value: Any) -> float | int | None:
    text = _to_text(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return _clean_number(float(text))
    except Exception:
        return None


def _slug(value: Any) -> str:
    text = re.sub(r"\s+", "-", _to_text(value).strip().lower())
    text = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff._@/-]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "protocol"


def _stable_hash(payload: Any) -> str:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _hash16(payload: Any) -> str:
    return _stable_hash(payload)[:16]


def _normref_docs_root() -> Path:
    root = Path(__file__).resolve().parents[5]
    return (root / "docs" / "normref").resolve()


def _normalize_header(value: Any) -> str:
    text = _to_text(value).strip().lower()
    text = re.sub(r"[\s_\-()\[\]{}:：]+", "", text)
    return text


def _normalize_severity(value: Any) -> str:
    text = _to_text(value).strip().lower()
    if text in {"warning", "warn", "soft", "提示", "预警", "警告"}:
        return "warning"
    return "mandatory"


def _extract_unit(text: str) -> str:
    match = re.search(r"(mm|cm|m|km|kg|t|%|‰|mpa|kn|m3|m²|m2|个|项)$", text.lower())
    if match:
        return match.group(1)
    cn_unit = re.search(r"(毫米|厘米|米|千米|千克|吨|百分比)$", text)
    if cn_unit:
        return cn_unit.group(1)
    return ""


def _parse_threshold(raw_value: Any, fallback_unit: str = "") -> dict[str, Any]:
    raw = _to_text(raw_value).strip()
    text = raw.replace(" ", "")
    unit = fallback_unit.strip() or _extract_unit(text)

    range_match = re.search(r"([-+]?\d+(?:\.\d+)?)\s*(?:~|～|至|-)\s*([-+]?\d+(?:\.\d+)?)", text)
    if range_match:
        left = float(range_match.group(1))
        right = float(range_match.group(2))
        if left > right:
            left, right = right, left
        return {
            "value": [_clean_number(left), _clean_number(right)],
            "operator": "range",
            "unit": unit,
            "raw": raw,
        }

    plus_minus_match = re.search(r"^(?:±|\+/-)(\d+(?:\.\d+)?)", text)
    if plus_minus_match:
        value = _clean_number(float(plus_minus_match.group(1)))
        return {"value": value, "operator": "lte", "unit": unit, "symmetric": True, "raw": raw}

    lte_match = re.search(r"^(?:<=|≤|不大于|小于等于)([-+]?\d+(?:\.\d+)?)", text)
    if lte_match:
        return {
            "value": _clean_number(float(lte_match.group(1))),
            "operator": "lte",
            "unit": unit,
            "raw": raw,
        }

    gte_match = re.search(r"^(?:>=|≥|不小于|大于等于)([-+]?\d+(?:\.\d+)?)", text)
    if gte_match:
        return {
            "value": _clean_number(float(gte_match.group(1))),
            "operator": "gte",
            "unit": unit,
            "raw": raw,
        }

    eq_match = re.search(r"^(?:=|等于)([-+]?\d+(?:\.\d+)?)", text)
    if eq_match:
        return {
            "value": _clean_number(float(eq_match.group(1))),
            "operator": "eq",
            "unit": unit,
            "raw": raw,
        }

    numeric = _to_number(text.replace("%", ""))
    if numeric is not None:
        if "%" in text and not unit:
            unit = "%"
        return {"value": numeric, "operator": "eq", "unit": unit, "raw": raw}

    return {"value": raw, "operator": "eq", "unit": unit, "raw": raw}


def _read_csv_rows(content: bytes) -> list[dict[str, Any]]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = content.decode(encoding)
            reader = csv.DictReader(io.StringIO(text))
            return [dict(row) for row in reader if isinstance(row, dict)]
        except Exception as exc:  # pragma: no cover
            last_error = exc
    raise HTTPException(400, f"failed to decode csv: {last_error}")


def _read_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise HTTPException(400, "openpyxl is required for .xlsx parsing")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_to_text(v).strip() for v in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        record: dict[str, Any] = {}
        has_value = False
        for idx, cell in enumerate(raw):
            key = headers[idx] if idx < len(headers) else f"col_{idx + 1}"
            record[key or f"col_{idx + 1}"] = cell
            if _to_text(cell).strip():
                has_value = True
        if has_value:
            out.append(record)
    return out


def _read_xls_rows(content: bytes) -> list[dict[str, Any]]:
    if xlrd is None:
        raise HTTPException(400, "xlrd is required for .xls parsing")
    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    if sheet.nrows <= 0:
        return []
    headers = [_to_text(sheet.cell_value(0, col)).strip() for col in range(sheet.ncols)]
    out: list[dict[str, Any]] = []
    for row in range(1, sheet.nrows):
        record: dict[str, Any] = {}
        has_value = False
        for col in range(sheet.ncols):
            key = headers[col] if col < len(headers) else f"col_{col + 1}"
            value = sheet.cell_value(row, col)
            record[key or f"col_{col + 1}"] = value
            if _to_text(value).strip():
                has_value = True
        if has_value:
            out.append(record)
    return out


def _read_table_rows(file_name: str, content: bytes) -> list[dict[str, Any]]:
    suffix = Path(file_name or "table.csv").suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(content)
    if suffix == ".xlsx":
        return _read_xlsx_rows(content)
    if suffix == ".xls":
        return _read_xls_rows(content)
    raise HTTPException(400, f"unsupported quality table format: {suffix}")


def _row_value(row: dict[str, Any], aliases: list[str]) -> Any:
    if not row:
        return None
    normalized_map = {_normalize_header(k): v for k, v in row.items()}
    for alias in aliases:
        val = normalized_map.get(_normalize_header(alias))
        if _to_text(val).strip():
            return val
    return None


def _extract_gates(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    gates: list[dict[str, Any]] = []
    for row in rows:
        label = _to_text(
            _row_value(
                row,
                [
                    "label",
                    "check_item",
                    "检查项",
                    "检查项目",
                    "检测项",
                    "项目",
                    "指标",
                    "name",
                ],
            )
        ).strip()
        threshold_value = _row_value(
            row,
            [
                "threshold",
                "allowance",
                "limit",
                "允许偏差",
                "阈值",
                "判定值",
                "标准值",
            ],
        )
        if not label and not _to_text(threshold_value).strip():
            continue

        check_id = _to_text(
            _row_value(row, ["check_id", "check_code", "检查编号", "检查项编码", "id", "code"])
        ).strip() or _slug(label)
        norm_ref = _to_text(
            _row_value(row, ["norm_ref", "norm", "规范", "规范条文", "条文", "标准引用"])
        ).strip()
        unit = _to_text(_row_value(row, ["unit", "单位"])).strip()
        severity = _normalize_severity(_row_value(row, ["severity", "严重级别", "级别", "严重性"]))

        threshold = _parse_threshold(threshold_value, fallback_unit=unit)

        gates.append(
            {
                "check_id": check_id,
                "label": label or check_id,
                "norm_ref": norm_ref,
                "threshold": threshold,
                "severity": severity,
            }
        )

    if not gates:
        raise HTTPException(400, "no gate rows found in table")
    return gates


def _logic_inputs_from_gates(gates: list[dict[str, Any]], description: str) -> list[dict[str, Any]]:
    text = " ".join([_to_text(g.get("label")).lower() + " " + _to_text(g.get("check_id")).lower() for g in gates])
    text = f"{text} {_to_text(description).lower()}"
    out: list[dict[str, Any]] = []

    def _append(name: str, hint: str, unit: str = "") -> None:
        if any(_to_text(item.get("name")).strip() == name for item in out):
            return
        out.append({"name": name, "hint": hint, "unit": unit})

    if "diameter" in text or "直径" in text:
        _append("design_diameter", "Design diameter from drawing", "mm")
        _append("measured_diameter", "Measured diameter from field", "mm")
    if "spacing" in text or "间距" in text:
        _append("measured_spacing", "Measured spacing from field", "mm")
    if "保护层" in text or "protection" in text:
        _append("measured_protection_layer", "Measured protection layer thickness", "mm")
    if "weld" in text or "焊" in text:
        _append("weld_quality_level", "Weld quality level (I/II/III)")
    if "raft" in text or "筏" in text:
        _append("design_thickness", "Design raft thickness from drawing", "mm")
        _append("measured_thickness", "Measured raft thickness from field", "mm")
        _append("measured_concrete_strength", "Measured concrete strength", "MPa")
        _append("measured_rebar_spacing", "Measured rebar spacing", "mm")
    if not out:
        _append("measured_value", "Measured value from field")
    return out


def _state_matrix(
    *,
    topology_component_count: int,
    forms_per_component: int,
    generated_qc_table_count: int,
    signed_pass_table_count: int,
) -> dict[str, int]:
    comp = max(int(topology_component_count), 0)
    per_comp = max(int(forms_per_component), 1)
    expected = comp * per_comp
    generated = max(int(generated_qc_table_count), 0)
    signed = max(int(signed_pass_table_count), 0)
    pending = expected - generated
    if pending < 0:
        pending = 0
    return {
        "component_count": comp,
        "forms_per_component": per_comp,
        "expected_qc_table_count": expected,
        "generated_qc_table_count": generated,
        "signed_pass_table_count": signed,
        "pending_qc_table_count": pending,
        "total_qc_tables": expected,
        "total": expected,
        "generated": generated,
        "signed": signed,
        "pending": pending,
    }


def _empty_qc_state_matrix(
    *,
    total: int = 0,
    generated: int = 0,
    signed: int = 0,
    pending: int | None = None,
    component_count: int = 0,
    forms_per_component: int = 0,
) -> dict[str, int]:
    total_val = max(int(total), 0)
    generated_val = max(int(generated), 0)
    signed_val = max(int(signed), 0)
    pending_val = max(int(total_val - generated_val), 0) if pending is None else max(int(pending), 0)
    return {
        "component_count": max(int(component_count), 0),
        "forms_per_component": max(int(forms_per_component), 0),
        "expected_qc_table_count": total_val,
        "generated_qc_table_count": generated_val,
        "signed_pass_table_count": signed_val,
        "pending_qc_table_count": pending_val,
        "total_qc_tables": total_val,
        "total": total_val,
        "generated": generated_val,
        "signed": signed_val,
        "pending": pending_val,
    }


def _schema_qc_v1() -> dict[str, Any]:
    return {
        "uri": NORMREF_SCHEMA_QC_V1_URI,
        "version": "v1",
        "protocol_type": "QualityCheckProtocol",
        "domain": "construction/highway",
        "anchor": {
            "boq_item_uri": "v://.../boq/{boq_item_id}",
            "component_uri": "v://.../component/{component_id}",
        },
        "structure": {
            "metadata": {
                "norm_code": "string",
                "boq_item_id": "string",
                "bridge_uri": "string?",
                "component_type": "string?",
            },
            "gates": [
                {
                    "check_id": "string",
                    "label": "string",
                    "norm_ref": "string",
                    "threshold": {
                        "value": "number|[number,number]|string",
                        "operator": "gte|lte|range|eq",
                        "unit": "string?",
                    },
                    "severity": "mandatory|warning",
                    "explain": "string?",
                }
            ],
            "verdict_logic": "input -> boolean + explain + proof",
            "output_schema": {
                "result": "PASS|FAIL|WARNING",
                "failed_gates": "array<string>",
                "explain": "string",
                "proof_hash": "string",
                "sealed_at": "ISO-8601",
            },
            "layers": {
                "header": {
                    "doc_type": "string",
                    "doc_id": "string",
                    "v_uri": "string",
                    "project_ref": "string",
                    "version": "string",
                    "created_at": "ISO-8601",
                    "jurisdiction": "string",
                },
                "gate": {
                    "pre_conditions": "array<string>",
                    "entry_rules": "array<object>",
                    "required_trip_roles": "array<string>",
                },
                "body": {
                    "basic": "object",
                    "test_data": "array<object>",
                    "relations": "array<string>",
                },
                "proof": {
                    "data_hash": "string",
                    "proof_hash": "string",
                    "signatures": "array<string>",
                    "witness_logs": "array<string>",
                    "timestamps": "object",
                },
                "state": {
                    "lifecycle_stage": "draft|inspecting|approved|rejected|archived",
                    "state_matrix": {
                        "component_count": "number",
                        "forms_per_component": "number",
                        "expected_qc_table_count": "number",
                        "generated_qc_table_count": "number",
                        "signed_pass_table_count": "number",
                        "pending_qc_table_count": "number",
                        "total_qc_tables": "number",
                        "total": "number",
                        "generated": "number",
                        "signed": "number",
                        "pending": "number",
                    },
                    "next_action": "string",
                    "valid_until": "ISO-8601|string",
                },
            },
        },
    }


def _compose_five_layers(
    *,
    protocol_uri: str,
    jurisdiction: str,
    doc_type_uri: str = "v://normref.com/doc-type/quality-inspection@v1",
    trip_role: str = "quality.check",
    industry: str = "highway",
    required_trip_roles: list[str] | None = None,
    pre_conditions: list[str] | None = None,
    entry_rules: list[dict[str, Any]] | None = None,
    basic_body: dict[str, Any] | None = None,
    body_test_data_schema: list[dict[str, Any]] | None = None,
    state_matrix_schema: dict[str, Any] | None = None,
    next_action: str = "",
    project_ref: str = "",
    doc_id: str = "",
    created_at: str = "",
    valid_until: str = "",
    lifecycle_stage: str = "draft",
) -> dict[str, Any]:
    created = created_at or datetime.now(UTC).isoformat()
    doc_token = _slug(doc_id or protocol_uri.split("/")[-1]).replace("-", "").upper()[:10] or "DOC0000001"
    resolved_doc_id = doc_id or f"NINST-{doc_token}"
    return {
        "header": {
            "doc_type": doc_type_uri,
            "doc_id": resolved_doc_id,
            "v_uri": protocol_uri,
            "project_ref": project_ref,
            "jurisdiction": jurisdiction,
            "version": "v1",
            "created_at": created,
            "industry": industry,
            "trip_role": trip_role,
        },
        "gate": {
            "required_trip_roles": required_trip_roles
            or ["inspector.quality.check", "supervisor.approve"],
            "pre_conditions": pre_conditions or ["原材料合格", "设备校准有效"],
            "entry_rules": entry_rules or [],
        },
        "body": {
            "basic": basic_body or {},
            "test_data": body_test_data_schema or [],
            "relations": ["v://.../boq/{boq_item_id}", "v://.../component/{component_id}"],
        },
        "proof": {
            "signatures": [],
            "data_hash": "sha256:...",
            "witness_logs": [],
            "proof_hash": "",
            "timestamps": {"created_at": created, "sealed_at": ""},
        },
        "state": {
            "lifecycle_stage": lifecycle_stage,
            "state_matrix": state_matrix_schema or {},
            "next_action": next_action or "等待质检数据输入后执行 Gate 判定并生成 Proof",
            "valid_until": valid_until,
        },
    }


def _general_quality_template_v1() -> dict[str, Any]:
    return {
        "uri": NORMREF_QC_GENERAL_TEMPLATE_URI,
        "schema_uri": NORMREF_SCHEMA_QC_V1_URI,
        "version": "v1",
        "metadata": {
            "norm_code": "GB50204 / JTG F80 / SL223",
            "description": "通用质检表格协议模板",
            "doc_type": "v://normref.com/doc-type/quality-inspection@v1",
            "domain": "construction/highway",
        },
        "gates": [],
        "verdict_logic": {
            "rule": "Load concrete gates from concrete protocol and evaluate all mandatory checks.",
            "input": {"actual_values": {}, "design_values": {}},
            "output": {"result": "PASS|FAIL|WARNING", "failed_gates": [], "explain": "", "proof_hash": ""},
        },
        "output_schema": {
            "result": "PASS|FAIL|WARNING",
            "failed_gates": [],
            "explain": "",
            "proof_hash": "",
            "sealed_at": "ISO-8601",
        },
        "layers": _compose_five_layers(
            protocol_uri=NORMREF_QC_GENERAL_TEMPLATE_URI,
            jurisdiction="GB50204 / JTG F80 / SL223",
            doc_type_uri="v://normref.com/doc-type/quality-inspection@v1",
            entry_rules=[],
            basic_body={"location": "string", "component_type": "string", "quantity": 0},
            body_test_data_schema=[
                {
                    "item": "string",
                    "standard": "string",
                    "measured": "any",
                    "unit": "string",
                    "result": "合格|不合格|警告",
                }
            ],
            state_matrix_schema=_empty_qc_state_matrix(),
            next_action="按具体协议块补齐 entry_rules 和 gates",
        ),
    }


def _seed_rebar_protocol_v1() -> dict[str, Any]:
    protocol = {
        "uri": NORMREF_SEED_REBAR_PROTOCOL_URI,
        "schema_uri": NORMREF_SCHEMA_QC_V1_URI,
        "version": "v1",
        "metadata": {
            "norm_code": "GB50204-2015 5.3.2 + JTG F80/1-2017",
            "boq_item_id": "403-1-2",
            "description": "钢筋加工及安装（制作、运输、安装、焊接）",
            "doc_type": "v://normref.com/doc-type/rebar-processing@v1",
            "applicable_component": ["pile", "pier", "cap", "beam"],
            "domain": "construction/highway",
        },
        "gates": [
            {
                "check_id": "diameter",
                "label": "直径允许偏差",
                "norm_ref": "GB50204-2015 5.3.2",
                "threshold": {"value": 0.02, "operator": "lte", "unit": "%"},
                "severity": "mandatory",
                "explain": "实际直径与设计值偏差不超过2%",
            },
            {
                "check_id": "spacing",
                "label": "钢筋间距偏差",
                "norm_ref": "GB50204-2015 5.3.2",
                "threshold": {"value": 10, "operator": "lte", "unit": "mm"},
                "severity": "mandatory",
            },
            {
                "check_id": "protection_layer",
                "label": "保护层厚度",
                "norm_ref": "JTG F80/1-2017",
                "threshold": {"value": 5, "operator": "lte", "unit": "mm"},
                "severity": "mandatory",
            },
        ],
        "verdict_logic": {
            "input": {
                "actual_values": {"diameter": 19.8, "spacing": 152},
                "design_values": {"diameter": 20.0, "spacing": 150},
            },
            "output": {
                "result": "PASS|FAIL|WARNING",
                "failed_gates": [],
                "explain": "所有检查项均满足规范要求",
                "proof_hash": "PF-XXXX...",
            },
            "rule": "For each gate evaluate actual vs threshold. All mandatory pass => PASS.",
        },
        "output_schema": {
            "result": "PASS",
            "failed_gates": [],
            "explain": "所有检查项均满足规范要求",
            "proof_hash": "PF-XXXX...",
            "sealed_at": "2026-04-03T13:08:00Z",
        },
    }
    protocol["layers"] = _compose_five_layers(
        protocol_uri=NORMREF_SEED_REBAR_PROTOCOL_URI,
        jurisdiction="GB50204-2015 5.3.2 + JTG F80/1-2017",
        doc_type_uri="v://normref.com/doc-type/rebar-processing@v1",
        entry_rules=[
            {"name": "diameter_deviation", "operator": "lte", "value": 0.02, "unit": "%"},
            {"name": "spacing_deviation", "operator": "lte", "value": 10, "unit": "mm"},
            {"name": "protection_layer", "operator": "lte", "value": 5, "unit": "mm"},
        ],
        basic_body={"rebar_spec": "HRB400-Φ20", "total_weight": 0},
        body_test_data_schema=[
            {"check_item": "diameter", "measured": 0, "result": "合格|不合格"},
            {"check_item": "spacing", "measured": 0, "result": "合格|不合格"},
            {"check_item": "protection", "measured": 0, "result": "合格|不合格"},
        ],
        state_matrix_schema=_empty_qc_state_matrix(),
    )
    return protocol


def _concrete_compressive_protocol_v1() -> dict[str, Any]:
    protocol = {
        "uri": NORMREF_CONCRETE_QC_PROTOCOL_URI,
        "schema_uri": NORMREF_SCHEMA_QC_V1_URI,
        "version": "v1",
        "metadata": {
            "norm_code": "GB/T 50081 + JTG F80/1-2017",
            "description": "混凝土抗压强度检测协议块",
            "doc_type": "v://normref.com/doc-type/concrete-compressive-test@v1",
            "domain": "construction/highway",
        },
        "gates": [
            {
                "check_id": "sample_size",
                "label": "试件组最小数量",
                "norm_ref": "GB/T 50081",
                "threshold": {"value": 3, "operator": "gte", "unit": "个"},
                "severity": "mandatory",
            },
            {
                "check_id": "test_age",
                "label": "试验龄期",
                "norm_ref": "GB/T 50081",
                "threshold": {"value": 28, "operator": "eq", "unit": "d"},
                "severity": "mandatory",
            },
            {
                "check_id": "compressive_strength",
                "label": "抗压强度",
                "norm_ref": "JTG F80/1-2017",
                "threshold": {"value": 30, "operator": "gte", "unit": "MPa"},
                "severity": "mandatory",
            },
        ],
        "verdict_logic": {
            "rule": "Sample size / test age / compressive strength all pass => PASS.",
            "input": {"actual_values": {}, "design_values": {}},
            "output": {"result": "PASS|FAIL|WARNING", "failed_gates": [], "explain": "", "proof_hash": ""},
        },
        "output_schema": {
            "result": "PASS|FAIL|WARNING",
            "failed_gates": [],
            "explain": "",
            "proof_hash": "",
            "sealed_at": "ISO-8601",
        },
    }
    protocol["layers"] = _compose_five_layers(
        protocol_uri=NORMREF_CONCRETE_QC_PROTOCOL_URI,
        jurisdiction="GB/T 50081 + JTG F80/1-2017",
        doc_type_uri="v://normref.com/doc-type/concrete-compressive-test@v1",
        entry_rules=[
            {"name": "min_sample_size", "value": 3},
            {"name": "curing_method", "value": "标准养护28d"},
            {"name": "test_age", "value": "28d"},
        ],
        basic_body={"concrete_grade": "C30", "pouring_volume": 0},
        body_test_data_schema=[
            {
                "group_id": "string",
                "measured": [0, 0, 0],
                "average": 0,
                "min_value": 0,
                "result": "合格|不合格",
            }
        ],
        state_matrix_schema={**_empty_qc_state_matrix(), "total_groups": 0, "qualified_groups": 0},
        next_action="完成组强度检测后自动生成混凝土强度 Proof",
    )
    return protocol


def _pile_foundation_protocol_v1() -> dict[str, Any]:
    protocol = {
        "uri": NORMREF_PILE_QC_PROTOCOL_URI,
        "schema_uri": NORMREF_SCHEMA_QC_V1_URI,
        "version": "v1",
        "metadata": {
            "norm_code": "JTG F80/1-2017",
            "description": "桩基检测协议块",
            "doc_type": "v://normref.com/doc-type/pile-foundation@v1",
            "domain": "construction/highway",
        },
        "gates": [
            {
                "check_id": "verticality",
                "label": "垂直度偏差",
                "norm_ref": "JTG F80/1-2017",
                "threshold": {"value": 1, "operator": "lte", "unit": "%"},
                "severity": "mandatory",
            },
            {
                "check_id": "concrete_strength",
                "label": "桩身混凝土强度",
                "norm_ref": "JTG F80/1-2017",
                "threshold": {"value": 30, "operator": "gte", "unit": "MPa"},
                "severity": "mandatory",
            },
        ],
        "verdict_logic": {
            "rule": "Verticality and concrete strength both pass => PASS.",
            "input": {"actual_values": {}, "design_values": {}},
            "output": {"result": "PASS|FAIL|WARNING", "failed_gates": [], "explain": "", "proof_hash": ""},
        },
        "output_schema": {
            "result": "PASS|FAIL|WARNING",
            "failed_gates": [],
            "explain": "",
            "proof_hash": "",
            "sealed_at": "ISO-8601",
        },
    }
    protocol["layers"] = _compose_five_layers(
        protocol_uri=NORMREF_PILE_QC_PROTOCOL_URI,
        jurisdiction="JTG F80/1-2017",
        doc_type_uri="v://normref.com/doc-type/pile-foundation@v1",
        entry_rules=[
            {"name": "verticality", "operator": "lte", "value": 1, "unit": "%"},
            {"name": "concrete_strength", "operator": "gte", "value": 30, "unit": "MPa"},
        ],
        basic_body={"pile_id": "P3", "length": 0, "diameter": 0},
        body_test_data_schema=[
            {"item": "verticality", "measured": 0, "result": "合格|不合格"},
            {"item": "integrity", "measured": 0, "result": "合格|不合格"},
            {"item": "bearing_capacity", "measured": 0, "result": "合格|不合格"},
        ],
        state_matrix_schema=_empty_qc_state_matrix(),
        next_action="完成桩基检测后自动关联桥级状态矩阵",
    )
    return protocol



def _raft_spu_v1() -> dict[str, Any]:
    return {
        "uri": NORMREF_RAFT_SPU_URI,
        "version": "v1",
        "kind": "spu",
        "metadata": {
            "name": "Raft foundation",
            "description": "Standard SPU for raft foundation works.",
            "unit": "m3",
            "domain": "construction/highway",
            "norm_refs": ["GB50204-2015", "JTG F80/1-2017"],
        },
        "measure_rule": {
            "statement": "Meter by effective concrete volume of approved raft sections.",
            "operator": "design-volume",
            "expression": "raft_length * raft_width * raft_thickness",
        },
        "consumption": {
            "materials": [
                {"name": "C30 concrete", "unit": "m3", "quantity_per_unit": 1.0},
                {"name": "HRB400 rebar", "unit": "kg", "quantity_per_unit": 110.0},
            ],
            "labor": [{"name": "raft-casting-crew", "unit": "shift", "quantity_per_unit": 0.08}],
        },
        "qc_gate": {
            "gate_refs": [
                "v://normref.com/gate/raft-thickness@v1",
                "v://normref.com/gate/raft-concrete-strength@v1",
                "v://normref.com/gate/raft-rebar-spacing@v1",
            ]
        },
        "thresholds": {
            "thickness_tolerance_mm": 10,
            "concrete_strength_mpa_min": 30,
            "rebar_spacing_tolerance_mm": 10,
        },
    }


def _raft_qc_protocol_v1() -> dict[str, Any]:
    protocol = {
        "uri": NORMREF_RAFT_QC_PROTOCOL_URI,
        "schema_uri": NORMREF_SCHEMA_QC_V1_URI,
        "version": "v1",
        "metadata": {
            "norm_code": "GB50204-2015 + JTG F80/1-2017",
            "boq_item_id": "403-raft-foundation",
            "description": "筏基础施工质检协议块",
            "doc_type": "v://normref.com/doc-type/raft-foundation@v1",
            "applicable_component": ["raft_foundation"],
            "ref_spu_uri": NORMREF_RAFT_SPU_URI,
            "domain": "construction/highway",
        },
        "gates": [
            {
                "check_id": "raft_thickness",
                "label": "厚度偏差",
                "norm_ref": "GB50204-2015",
                "threshold": {"value": 10, "operator": "lte", "unit": "mm"},
                "severity": "mandatory",
                "explain": "实测厚度与设计厚度偏差不超过 10mm",
            },
            {
                "check_id": "concrete_strength",
                "label": "混凝土强度",
                "norm_ref": "GB50204-2015 7.4",
                "threshold": {"value": 30, "operator": "gte", "unit": "MPa"},
                "severity": "mandatory",
                "explain": "强度等级不低于 C30",
            },
            {
                "check_id": "rebar_spacing",
                "label": "钢筋间距偏差",
                "norm_ref": "JTG F80/1-2017",
                "threshold": {"value": 10, "operator": "lte", "unit": "mm"},
                "severity": "mandatory",
                "explain": "钢筋间距偏差不超过 10mm",
            },
        ],
        "verdict_logic": {
            "rule": "All mandatory gates pass => PASS; any mandatory fail => FAIL.",
            "input": {"actual_values": {}, "design_values": {}},
            "output": {
                "result": "PASS|FAIL|WARNING",
                "failed_gates": [],
                "explain": "",
                "proof_hash": "",
            },
        },
        "output_schema": {
            "result": "PASS|FAIL|WARNING",
            "failed_gates": [],
            "explain": "",
            "proof_hash": "",
            "sealed_at": "ISO-8601",
        },
    }
    protocol["layers"] = _compose_five_layers(
        protocol_uri=NORMREF_RAFT_QC_PROTOCOL_URI,
        jurisdiction="GB50204-2015 + JTG F80/1-2017",
        doc_type_uri="v://normref.com/doc-type/raft-foundation@v1",
        entry_rules=[
            {"name": "thickness_deviation", "operator": "lte", "value": 10, "unit": "mm"},
            {"name": "concrete_strength", "operator": "gte", "value": 30, "unit": "MPa"},
            {"name": "rebar_spacing", "operator": "lte", "value": 10, "unit": "mm"},
        ],
        basic_body={"component_type": "raft_foundation", "quantity": 0},
        body_test_data_schema=[
            {"item": "thickness", "standard": "<=10mm", "measured": 0, "unit": "mm", "result": "合格|不合格"},
            {"item": "concrete_strength", "standard": ">=30MPa", "measured": 0, "unit": "MPa", "result": "合格|不合格"},
        ],
        state_matrix_schema=_empty_qc_state_matrix(),
    )
    return protocol


def _prompt_t2p_template() -> str:
    return (
        "你是一个严格的工程质检结构化引擎（Tab-to-Peg Engine v1.0）。\n\n"
        "任务：把任意质检表格（Excel行、文字描述、扫描件内容）自动转换为 NormRef 协议块，并生成完整的五层结构文档。\n\n"
        "输入格式：\n"
        "- BOQItem 信息（编码、描述、工程量、单位）\n"
        "- 图纸拓扑（构件数量、节段、部位）\n"
        "- 质检表格内容（检测项目、标准值、实测值、规范依据）\n\n"
        "严格执行以下步骤：\n"
        "Step 1: 识别文档类型和关联 BOQItem\n"
        "Step 2: 计算检测频率和总质检表数量（根据规范 + 工程量 + 构件数）\n"
        "Step 3: 提取并参数化 Gate（把“允许偏差 ≤ 2%”转为 operator + value）\n"
        "Step 4: 生成五层结构（Header + Gate + Body + Proof + State）\n"
        "Step 5: 计算 State Matrix（兼容双命名）\n"
        "  - 规范命名：expected_qc_table_count / generated_qc_table_count / signed_pass_table_count / pending_qc_table_count\n"
        "  - 展示命名：total_qc_tables / generated / signed / pending\n"
        "Step 6: 输出完整 Markdown（带 v:// URI）\n\n"
        "输出必须严格使用以下五层结构：\n"
        "Layer 1: Header -> doc_type, v_uri, jurisdiction, trip_role\n"
        "Layer 2: Gate -> required_trip_roles, entry_rules\n"
        "Layer 3: Body -> basic, test_data, relations\n"
        "Layer 4: Proof -> proof_hash, signatures\n"
        "Layer 5: State -> lifecycle_stage, state_matrix(expected/generated_qc/signed_pass/pending_qc + total/generated/signed/pending), next_action\n\n"
        "现在处理以下质检表格输入：\n"
        "BOQItem: {{boq_item}}\n"
        "工程量: {{quantity}} {{unit}}\n"
        "图纸拓扑: {{drawing_topology}}\n"
        "质检内容: {{table_content}}\n\n"
        "请生成完整的五层结构文档。\n"
    )


def _uri_to_docs_paths(uri: str, root: Path) -> tuple[Path, Path]:
    normalized = _to_text(uri).strip().rstrip("/")
    if not normalized.startswith("v://normref.com"):
        token = _slug(normalized.replace("v://", "").replace("/", "__"))
        md_path = root / "misc" / f"{token}.md"
        return md_path, md_path.with_suffix(".json")

    rel = normalized[len("v://normref.com") :].lstrip("/")
    if not rel:
        rel = "index"
    md_path = root / f"{rel}.md"
    return md_path, md_path.with_suffix(".json")


def _render_core_md() -> str:
    return """# v://normref.com/core@v1

**协议类型**：NormRefCore  
**版本**：v1  
**描述**：所有工程逻辑的统一入口和版本控制机制  
**规则**：
- 所有协议必须有 `v://normref.com/...@版本` 格式
- 必须包含 `metadata`、`gates`、`verdict_logic`、`output_schema`
- 变更必须生成 Proof 并记录在对应 `.md` 中
- 支持行业域划分（highway、bridge、building 等）
"""


def _render_highway_md() -> str:
    return """# v://normref.com/construction/highway@v1

**描述**：公路工程质量检验评定标准逻辑域  
**子域**：
- qc/rebar-processing
- qc/raft-foundation
- spu/raft-foundation
- qc/concrete-pouring
- qc/pile-foundation
- boq/400-chapter
"""


def _render_protocol_md(protocol: dict[str, Any]) -> str:
    logic_inputs = [x for x in (protocol.get("logic_inputs") or []) if isinstance(x, dict)]
    state_matrix = _as_dict(protocol.get("state_matrix"))
    layers = _as_dict(protocol.get("layers"))
    return "\n".join(
        [
            f"# {_to_text(protocol.get('uri')).strip()}",
            "",
            "**元数据 (Anchor)**",
            f"- norm_code: {_to_text(_as_dict(protocol.get('metadata')).get('norm_code') or '-').strip()}",
            f"- boq_item_id: {_to_text(_as_dict(protocol.get('metadata')).get('boq_item_id') or '-').strip()}",
            f"- description: {_to_text(_as_dict(protocol.get('metadata')).get('description') or '-').strip()}",
            f"- applicable_component: {json.dumps(_as_dict(protocol.get('metadata')).get('applicable_component') or [], ensure_ascii=False)}",
            "",
            "**阈值阵列 (Gates)**",
            "```json",
            json.dumps(protocol.get("gates") or [], ensure_ascii=False, indent=2),
            "```",
            "",
            "**判定逻辑 (verdict_logic)**",
            "```json",
            json.dumps(protocol.get("verdict_logic") or {}, ensure_ascii=False, indent=2),
            "```",
            "",
            "**Logic Inputs（必填输入项 / 参数化 Gate）**",
            "```json",
            json.dumps(logic_inputs, ensure_ascii=False, indent=2),
            "```",
            "",
            "**State Matrix**",
            "```json",
            json.dumps(state_matrix, ensure_ascii=False, indent=2),
            "```",
            "",
            "**DocPeg 五层结构 (Layers)**",
            "```json",
            json.dumps(layers, ensure_ascii=False, indent=2),
            "```",
            "",
            "**输出 Schema (output_schema)**",
            "```json",
            json.dumps(protocol.get("output_schema") or {}, ensure_ascii=False, indent=2),
            "```",
            "",
        ]
    )


def _render_prompt_md() -> str:
    return f"# {NORMREF_PROMPT_T2P_URI}\n\n```text\n{_prompt_t2p_template()}```\n"


def _write_markdown_and_json(*, uri: str, markdown: str, payload: dict[str, Any] | None, output_root: Path | None) -> dict[str, str]:
    root = output_root.resolve() if isinstance(output_root, Path) else _normref_docs_root()
    md_path, json_path = _uri_to_docs_paths(uri, root)
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(markdown.rstrip() + "\n", encoding="utf-8")
    if payload is not None:
        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    return {"md_path": str(md_path), "json_path": str(json_path)}


def _ensure_specir_nodes(*, sb: Any, owner_uri: str, nodes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if sb is None:
        raise HTTPException(400, "supabase client is required when commit=true")
    upserts: list[dict[str, Any]] = []
    for node in nodes:
        row = ensure_specir_object(
            sb=sb,
            uri=_to_text(node.get("uri")).strip(),
            kind=_to_text(node.get("kind") or "schema").strip() or "schema",
            title=_to_text(node.get("title")).strip(),
            owner_uri=owner_uri,
            content=_as_dict(node.get("content")),
        )
        upserts.append({"uri": node.get("uri"), "row": row})
    return upserts


def _register_gitpeg_uris(*, sb: Any, owner_uri: str, commit: bool, uris: list[dict[str, str]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for item in uris:
        out.append(
            register_uri(
                sb=sb,
                uri=item["uri"],
                uri_type=item.get("uri_type") or "protocol",
                metadata={"owner_uri": owner_uri, "source": "normref-scaffold"},
                source_system="qcspec-normref",
                commit=bool(commit),
            )
        )
    return out


def _create_proof(
    *,
    sb: Any,
    owner_uri: str,
    commit: bool,
    proof_kind: str,
    segment_uri: str,
    payload: dict[str, Any],
    norm_uri: str,
) -> dict[str, Any]:
    preview_hash = _stable_hash(payload)
    result: dict[str, Any] = {
        "proof_id": f"GP-{proof_kind.upper()}-{_hash16(payload).upper()}",
        "proof_hash": preview_hash,
        "segment_uri": segment_uri,
        "committed": False,
    }
    if not commit:
        return result
    if sb is None:
        raise HTTPException(400, "supabase client is required when commit=true")

    row = ProofUTXOEngine(sb).create(
        proof_id=result["proof_id"],
        owner_uri=owner_uri,
        project_uri=NORMREF_ROOT_URI,
        proof_type="protocol",
        result="PASS",
        state_data={"proof_kind": proof_kind, "payload": payload},
        norm_uri=norm_uri,
        segment_uri=segment_uri,
        signer_uri=owner_uri,
        signer_role="SYSTEM",
    )
    result["committed"] = True
    result["row"] = row
    result["proof_hash"] = _to_text(row.get("proof_hash")).strip() or preview_hash
    return result


def bootstrap_normref_logic_scaffold(
    *,
    sb: Any,
    commit: bool = False,
    owner_uri: str = "v://normref.com/executor/system/",
    write_files: bool = False,
    output_root: Path | None = None,
) -> dict[str, Any]:
    schema = _schema_qc_v1()
    general_template = _general_quality_template_v1()
    seed_protocol = _seed_rebar_protocol_v1()
    concrete_protocol = _concrete_compressive_protocol_v1()
    pile_protocol = _pile_foundation_protocol_v1()
    raft_spu = _raft_spu_v1()
    raft_protocol = _raft_qc_protocol_v1()

    nodes = [
        {
            "uri": NORMREF_CORE_URI,
            "kind": "namespace",
            "title": "NormRef Core v1",
            "content": {
                "protocol_type": "NormRefCore",
                "version": "v1",
                "requires": ["metadata", "gates", "verdict_logic", "output_schema"],
            },
        },
        {
            "uri": NORMREF_HIGHWAY_URI,
            "kind": "namespace",
            "title": "Construction Highway namespace v1",
            "content": {"parent_uri": NORMREF_CORE_URI, "domain": "construction/highway"},
        },
        {
            "uri": NORMREF_SCHEMA_QC_V1_URI,
            "kind": "schema",
            "title": "Quality Check Protocol Schema v1",
            "content": schema,
        },
        {
            "uri": NORMREF_QC_GENERAL_TEMPLATE_URI,
            "kind": "protocol_template",
            "title": "General Quality Inspection Template v1",
            "content": general_template,
        },
        {
            "uri": NORMREF_SEED_REBAR_PROTOCOL_URI,
            "kind": "protocol",
            "title": "Rebar Processing QC Protocol v1",
            "content": seed_protocol,
        },
        {
            "uri": NORMREF_CONCRETE_QC_PROTOCOL_URI,
            "kind": "protocol",
            "title": "Concrete Compressive Test QC Protocol v1",
            "content": concrete_protocol,
        },
        {
            "uri": NORMREF_PILE_QC_PROTOCOL_URI,
            "kind": "protocol",
            "title": "Pile Foundation QC Protocol v1",
            "content": pile_protocol,
        },
        {
            "uri": NORMREF_RAFT_SPU_URI,
            "kind": "spu",
            "title": "Raft Foundation SPU v1",
            "content": raft_spu,
        },
        {
            "uri": NORMREF_RAFT_QC_PROTOCOL_URI,
            "kind": "protocol",
            "title": "Raft Foundation QC Protocol v1",
            "content": raft_protocol,
        },
        {
            "uri": NORMREF_PROMPT_T2P_URI,
            "kind": "prompt_spu",
            "title": "Tab-to-Peg PromptSPU v1",
            "content": {"prompt": _prompt_t2p_template()},
        },
    ]

    specir_upserts: list[dict[str, Any]] = []
    gitpeg_registrations: list[dict[str, Any]] = []
    if _to_bool(commit):
        specir_upserts = _ensure_specir_nodes(sb=sb, owner_uri=owner_uri, nodes=nodes)
        gitpeg_registrations = _register_gitpeg_uris(
            sb=sb,
            owner_uri=owner_uri,
            commit=True,
            uris=[
                {"uri": NORMREF_CORE_URI, "uri_type": "normref_core"},
                {"uri": NORMREF_HIGHWAY_URI, "uri_type": "normref_domain"},
                {"uri": NORMREF_SCHEMA_QC_V1_URI, "uri_type": "normref_schema"},
                {"uri": NORMREF_QC_GENERAL_TEMPLATE_URI, "uri_type": "normref_protocol_template"},
                {"uri": NORMREF_SEED_REBAR_PROTOCOL_URI, "uri_type": "normref_protocol"},
                {"uri": NORMREF_CONCRETE_QC_PROTOCOL_URI, "uri_type": "normref_protocol"},
                {"uri": NORMREF_PILE_QC_PROTOCOL_URI, "uri_type": "normref_protocol"},
                {"uri": NORMREF_RAFT_SPU_URI, "uri_type": "normref_spu"},
                {"uri": NORMREF_RAFT_QC_PROTOCOL_URI, "uri_type": "normref_protocol"},
                {"uri": NORMREF_PROMPT_T2P_URI, "uri_type": "prompt_spu"},
            ],
        )

    files: dict[str, Any] = {}
    if _to_bool(write_files):
        files = {
            "core": _write_markdown_and_json(
                uri=NORMREF_CORE_URI,
                markdown=_render_core_md(),
                payload={"uri": NORMREF_CORE_URI, "protocol_type": "NormRefCore", "version": "v1"},
                output_root=output_root,
            ),
            "highway": _write_markdown_and_json(
                uri=NORMREF_HIGHWAY_URI,
                markdown=_render_highway_md(),
                payload={"uri": NORMREF_HIGHWAY_URI, "domain": "construction/highway", "version": "v1"},
                output_root=output_root,
            ),
            "schema": _write_markdown_and_json(
                uri=NORMREF_SCHEMA_QC_V1_URI,
                markdown=(
                    f"# {NORMREF_SCHEMA_QC_V1_URI}\n\n"
                    "该文件定义全行业通用质检协议模板（metadata / gates / verdict_logic / output_schema）。\n"
                ),
                payload=schema,
                output_root=output_root,
            ),
            "general_template": _write_markdown_and_json(
                uri=NORMREF_QC_GENERAL_TEMPLATE_URI,
                markdown=_render_protocol_md(general_template),
                payload=general_template,
                output_root=output_root,
            ),
            "rebar_protocol": _write_markdown_and_json(
                uri=NORMREF_SEED_REBAR_PROTOCOL_URI,
                markdown=_render_protocol_md(seed_protocol),
                payload=seed_protocol,
                output_root=output_root,
            ),
            "concrete_protocol": _write_markdown_and_json(
                uri=NORMREF_CONCRETE_QC_PROTOCOL_URI,
                markdown=_render_protocol_md(concrete_protocol),
                payload=concrete_protocol,
                output_root=output_root,
            ),
            "pile_protocol": _write_markdown_and_json(
                uri=NORMREF_PILE_QC_PROTOCOL_URI,
                markdown=_render_protocol_md(pile_protocol),
                payload=pile_protocol,
                output_root=output_root,
            ),
            "raft_spu": _write_markdown_and_json(
                uri=NORMREF_RAFT_SPU_URI,
                markdown=(
                    f"# {NORMREF_RAFT_SPU_URI}\n\n"
                    "标准计量单元（SPU）：筏基础。包含厚度、混凝土强度、钢筋配置、允许偏差等规范阈值。\n"
                ),
                payload=raft_spu,
                output_root=output_root,
            ),
            "raft_protocol": _write_markdown_and_json(
                uri=NORMREF_RAFT_QC_PROTOCOL_URI,
                markdown=_render_protocol_md(raft_protocol),
                payload=raft_protocol,
                output_root=output_root,
            ),
            "prompt_spu": _write_markdown_and_json(
                uri=NORMREF_PROMPT_T2P_URI,
                markdown=_render_prompt_md(),
                payload={"uri": NORMREF_PROMPT_T2P_URI, "prompt": _prompt_t2p_template()},
                output_root=output_root,
            ),
        }

    payload = {
        "event": "normref_logic_scaffold_bootstrap",
        "nodes": [{"uri": n["uri"], "kind": n["kind"]} for n in nodes],
        "committed_specir": _to_bool(commit),
        "wrote_files": _to_bool(write_files),
    }
    proof = _create_proof(
        sb=sb,
        owner_uri=owner_uri,
        commit=_to_bool(commit),
        proof_kind="normref-scaffold",
        segment_uri=NORMREF_CORE_URI,
        payload=payload,
        norm_uri=NORMREF_SCHEMA_QC_V1_URI,
    )

    return {
        "ok": True,
        "root_uri": NORMREF_CORE_URI,
        "levels": {
            "l0": NORMREF_CORE_URI,
            "l1": NORMREF_HIGHWAY_URI,
            "l2": NORMREF_RAFT_QC_PROTOCOL_URI,
        },
        "schema": schema,
        "general_template": general_template,
        "seed_protocol": seed_protocol,
        "concrete_protocol": concrete_protocol,
        "pile_protocol": pile_protocol,
        "raft_spu": raft_spu,
        "raft_protocol": raft_protocol,
        "protocol_catalog": [
            NORMREF_QC_GENERAL_TEMPLATE_URI,
            NORMREF_SEED_REBAR_PROTOCOL_URI,
            NORMREF_CONCRETE_QC_PROTOCOL_URI,
            NORMREF_PILE_QC_PROTOCOL_URI,
            NORMREF_RAFT_QC_PROTOCOL_URI,
        ],
        "prompt_spu": {"uri": NORMREF_PROMPT_T2P_URI, "name": "Tab-to-Peg Engine"},
        "specir_upserts": specir_upserts,
        "gitpeg_registrations": gitpeg_registrations,
        "files": files,
        "proof": proof,
    }


def table_to_protocol_block(
    *,
    sb: Any,
    upload_file_name: str,
    upload_content: bytes,
    protocol_uri: str = "",
    norm_code: str = "",
    boq_item_id: str = "",
    description: str = "",
    bridge_uri: str = "",
    component_type: str = "",
    topology_component_count: int = 0,
    forms_per_component: int = 2,
    generated_qc_table_count: int = 0,
    signed_pass_table_count: int = 0,
    owner_uri: str = "v://normref.com/executor/system/",
    commit: bool = False,
    write_files: bool = False,
    output_root: Path | None = None,
) -> dict[str, Any]:
    rows = _read_table_rows(upload_file_name, upload_content)
    gates = _extract_gates(rows)
    created_at = datetime.now(UTC).isoformat()

    inferred_description = description.strip() or (_to_text(rows[0].get("description")).strip() if rows else "")
    inferred_boq_item_id = boq_item_id.strip() or (_to_text(rows[0].get("boq_item_id")).strip() if rows else "")
    inferred_norm = norm_code.strip() or (_to_text(rows[0].get("norm_code")).strip() if rows else "")

    token = _slug(inferred_description or inferred_boq_item_id or Path(upload_file_name).stem or "qc-protocol")
    maybe_raft = ("raft" in token) or ("筏" in inferred_description)
    resolved_uri = protocol_uri.strip() or (
        NORMREF_RAFT_QC_PROTOCOL_URI if maybe_raft else f"v://normref.com/qc/{token}@v1"
    )
    logic_inputs = _logic_inputs_from_gates(gates, inferred_description)
    state_matrix = _state_matrix(
        topology_component_count=int(topology_component_count or 0),
        forms_per_component=int(forms_per_component or 2),
        generated_qc_table_count=int(generated_qc_table_count or 0),
        signed_pass_table_count=int(signed_pass_table_count or 0),
    )
    bridge_ref = bridge_uri.strip()
    project_ref = ""
    if "/bridge/" in bridge_ref:
        project_ref = bridge_ref.split("/bridge/", 1)[0].rstrip("/")
    if not project_ref and rows:
        project_ref = _to_text(rows[0].get("project_ref")).strip()
    doc_id = f"NINST-{_hash16({'uri': resolved_uri, 'boq_item_id': inferred_boq_item_id, 'ts': created_at}).upper()}"
    lifecycle_stage = "draft"
    expected_tables = int(state_matrix.get("expected_qc_table_count") or 0)
    signed_tables = int(state_matrix.get("signed_pass_table_count") or 0)
    generated_tables = int(state_matrix.get("generated_qc_table_count") or 0)
    if expected_tables > 0 and signed_tables >= expected_tables:
        lifecycle_stage = "approved"
    elif generated_tables > 0:
        lifecycle_stage = "inspecting"

    protocol = {
        "uri": resolved_uri,
        "schema_uri": NORMREF_SCHEMA_QC_V1_URI,
        "version": "v1",
        "metadata": {
            "norm_code": inferred_norm,
            "boq_item_id": inferred_boq_item_id,
            "description": inferred_description,
            "doc_id": doc_id,
            "project_ref": project_ref,
            "bridge_uri": bridge_uri.strip(),
            "component_type": component_type.strip(),
            "ref_spu_uri": NORMREF_RAFT_SPU_URI if maybe_raft else "",
            "domain": "construction/highway",
        },
        "gates": gates,
        "logic_inputs": logic_inputs,
        "state_matrix": state_matrix,
        "verdict_logic": {
            "rule": "For each gate: evaluate actual vs threshold; all mandatory PASS => PASS.",
            "input": {"actual_values": {}, "design_values": {}},
            "output": {
                "result": "PASS|FAIL|WARNING",
                "failed_gates": [],
                "explain": "",
                "proof_hash": "",
            },
        },
        "output_schema": {
            "result": "PASS|FAIL|WARNING",
            "failed_gates": [],
            "explain": "",
            "proof_hash": "",
            "sealed_at": datetime.now(UTC).isoformat(),
        },
        "source": {
            "file_name": upload_file_name,
            "row_count": len(rows),
            "generated_at": created_at,
        },
    }
    protocol["layers"] = _compose_five_layers(
        protocol_uri=resolved_uri,
        jurisdiction=inferred_norm or "GB50204 / JTG F80",
        entry_rules=gates,
        basic_body={
            "location": bridge_uri.strip() or "unspecified",
            "component_type": component_type.strip() or "component",
            "quantity": int(topology_component_count or 0),
        },
        body_test_data_schema=[
            {
                "item": "string",
                "standard": "string",
                "measured": "any",
                "unit": "string",
                "result": "合格|不合格|警告",
            }
        ],
        state_matrix_schema={
            "total_qc_tables": int(state_matrix.get("total_qc_tables") or 0),
            "expected_qc_table_count": int(state_matrix.get("expected_qc_table_count") or 0),
            "generated_qc_table_count": int(state_matrix.get("generated_qc_table_count") or 0),
            "signed_pass_table_count": int(state_matrix.get("signed_pass_table_count") or 0),
            "pending_qc_table_count": int(state_matrix.get("pending_qc_table_count") or 0),
            "total": int(state_matrix.get("total") or 0),
            "generated": int(state_matrix.get("generated") or 0),
            "signed": int(state_matrix.get("signed") or 0),
            "pending": int(state_matrix.get("pending") or 0),
        },
        project_ref=project_ref,
        doc_id=doc_id,
        created_at=created_at,
        lifecycle_stage=lifecycle_stage,
    )

    specir_upserts: list[dict[str, Any]] = []
    gitpeg_registrations: list[dict[str, Any]] = []
    if _to_bool(commit):
        nodes = [
            {
                "uri": NORMREF_SCHEMA_QC_V1_URI,
                "kind": "schema",
                "title": "Quality Check Protocol Schema v1",
                "content": _schema_qc_v1(),
            },
            {
                "uri": resolved_uri,
                "kind": "protocol",
                "title": inferred_description or token,
                "content": protocol,
            },
        ]
        specir_upserts = _ensure_specir_nodes(sb=sb, owner_uri=owner_uri, nodes=nodes)
        gitpeg_registrations = _register_gitpeg_uris(
            sb=sb,
            owner_uri=owner_uri,
            commit=True,
            uris=[
                {"uri": NORMREF_SCHEMA_QC_V1_URI, "uri_type": "normref_schema"},
                {"uri": resolved_uri, "uri_type": "normref_protocol"},
            ],
        )

    files: dict[str, str] = {}
    if _to_bool(write_files):
        files = _write_markdown_and_json(
            uri=resolved_uri,
            markdown=_render_protocol_md(protocol),
            payload=protocol,
            output_root=output_root,
        )

    payload = {
        "event": "tab_to_protocol",
        "protocol_uri": resolved_uri,
        "gate_count": len(gates),
        "source": protocol["source"],
    }
    proof = _create_proof(
        sb=sb,
        owner_uri=owner_uri,
        commit=_to_bool(commit),
        proof_kind="tab-to-peg",
        segment_uri=resolved_uri,
        payload=payload,
        norm_uri=NORMREF_SCHEMA_QC_V1_URI,
    )

    return {
        "ok": True,
        "protocol": protocol,
        "summary": {
            "gate_count": len(gates),
            "source_row_count": len(rows),
            "protocol_uri": resolved_uri,
            "expected_qc_table_count": state_matrix.get("expected_qc_table_count", 0),
            "pending_qc_table_count": state_matrix.get("pending_qc_table_count", 0),
        },
        "specir_upserts": specir_upserts,
        "gitpeg_registrations": gitpeg_registrations,
        "files": files,
        "proof": proof,
    }


__all__ = [
    "bootstrap_normref_logic_scaffold",
    "table_to_protocol_block",
]
