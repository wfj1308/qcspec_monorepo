"""BOQPeg upload parser shared by BOQ/SMU genesis workflows."""

from __future__ import annotations

import csv
import io
import re
import zipfile
from typing import Any

from fastapi import HTTPException

from services.api.domain.boq.runtime.utxo import BoqItem

ITEM_NO_PATTERN = re.compile(r"^\d{3}(?:-[0-9A-Za-z]+)*$")

HEADER_ALIASES: dict[str, set[str]] = {
    "item_no": {"item_no", "itemno", "item", "itemcode", "code", "子目号", "清单编码", "细目号"},
    "name": {"name", "itemname", "title", "名称", "项目名称", "清单名称"},
    "unit": {"unit", "uom", "单位", "计量单位"},
    "division": {"division", "分部", "分部工程"},
    "subdivision": {"subdivision", "分项", "分项工程"},
    "hierarchy": {"hierarchy", "wbs", "层级", "路径"},
    "design_quantity": {"designquantity", "design_qty", "designqty", "设计数量", "设计工程量"},
    "unit_price": {"unitprice", "price", "unit_price", "单价", "综合单价"},
    "approved_quantity": {"approvedquantity", "approved_qty", "approvedqty", "批复数量", "审核数量"},
    "approved_amount": {"approvedamount", "amount", "total", "批复金额", "合价", "金额"},
    "remark": {"remark", "note", "备注"},
}


def _to_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _to_float(value: Any) -> float | None:
    text = _to_text(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            return float(match.group(0))
        except Exception:
            return None


def _round4(value: float | None) -> float | None:
    if value is None:
        return None
    try:
        return round(float(value), 4)
    except Exception:
        return None


def _normalize_item_no(value: Any) -> str:
    text = _to_text(value).strip()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[./]", "-", text)
    text = re.sub(r"[()]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text


def _normalize_header(value: Any) -> str:
    text = _to_text(value).strip().lower()
    text = re.sub(r"[\s_/\-]+", "", text)
    text = text.replace("（", "(").replace("）", ")")
    return text


def _normalize_unit(value: Any) -> str:
    text = _to_text(value).strip()
    if not text:
        return ""
    key = text.replace(" ", "").replace("\u3000", "").lower()
    mapping = {
        "m3": "m3",
        "m^3": "m3",
        "立方米": "m3",
        "方": "m3",
        "m2": "m2",
        "m^2": "m2",
        "平方米": "m2",
        "平米": "m2",
        "m": "m",
        "米": "m",
        "kg": "kg",
        "千克": "kg",
        "t": "t",
        "吨": "t",
        "cny": "CNY",
        "元": "CNY",
    }
    return mapping.get(key, text)


def _fallback_approved_quantity(
    *,
    item_no: str,
    unit: str,
    approved_quantity: float | None,
    approved_amount: float | None,
    design_quantity: float | None,
    unit_price: float | None,
) -> tuple[float | None, bool]:
    if approved_quantity is not None:
        return approved_quantity, False
    unit_norm = _normalize_unit(unit)
    if unit_norm and unit_norm not in {"CNY", "元", "金额"}:
        return approved_quantity, False
    if approved_amount is None:
        return approved_quantity, False
    if unit_price and unit_price > 0:
        return _round4(approved_amount / unit_price), True
    if design_quantity and design_quantity > 0 and approved_amount > 0:
        return approved_amount, True
    if item_no.startswith("1") or item_no.startswith("2"):
        return approved_amount, True
    return approved_quantity, False


def _header_match(key: str, alias: str) -> bool:
    if not alias:
        return False
    if key == alias:
        return True
    # Avoid short-token overmatching, e.g. "unit" falsely matching "unitprice".
    if len(alias) >= 4 and alias in key:
        return True
    return False


def _detect_header_map(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    normalized_aliases = {
        field: {_normalize_header(alias) for alias in aliases}
        for field, aliases in HEADER_ALIASES.items()
    }
    best_idx = -1
    best_score = -1
    best_map: dict[str, int] = {}
    for idx, row in enumerate(rows):
        mapping: dict[str, int] = {}
        score = 0
        for col_idx, cell in enumerate(row):
            key = _normalize_header(cell)
            if not key:
                continue
            for field, aliases in normalized_aliases.items():
                if field in mapping:
                    continue
                if any(_header_match(key, alias) for alias in aliases):
                    mapping[field] = col_idx
                    score += 1
        if score > best_score and "item_no" in mapping and "name" in mapping:
            best_idx = idx
            best_score = score
            best_map = mapping
    if best_idx < 0:
        raise HTTPException(400, "Failed to detect BOQ header row from upload file.")
    return best_idx, best_map


def _rows_from_csv_bytes(content: bytes) -> list[list[Any]]:
    text = ""
    for enc in ("utf-8-sig", "gb18030", "gbk"):
        try:
            text = content.decode(enc)
            break
        except Exception:
            continue
    if not text:
        text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def _rows_from_xlsx_bytes(content: bytes) -> tuple[list[list[Any]], str]:
    try:
        import openpyxl
    except Exception as exc:
        raise HTTPException(500, f"openpyxl not available: {exc}") from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except zipfile.BadZipFile:
        if content[:4] == b"\xD0\xCF\x11\xE0":
            return _rows_from_xls_bytes(content)
        raise HTTPException(400, "Invalid .xlsx file.")
    except Exception as exc:
        raise HTTPException(400, f"Failed to read .xlsx file: {exc}")
    best_rows: list[list[Any]] = []
    best_sheet = wb.sheetnames[0] if wb.sheetnames else "sheet1"
    best_score = -1
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) >= 300000:
                break
        if not rows:
            continue
        score = -1
        try:
            _, colmap = _detect_header_map(rows[:120])
            score = len(colmap)
        except HTTPException:
            pass
        if score > best_score:
            best_rows = rows
            best_sheet = sheet_name
            best_score = score
    return best_rows, best_sheet


def _rows_from_xls_bytes(content: bytes) -> tuple[list[list[Any]], str]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise HTTPException(
            400,
            f".xls parser missing (xlrd). Please convert to CSV/XLSX or install xlrd. detail={exc}",
        ) from exc
    wb = xlrd.open_workbook(file_contents=content)
    best_rows: list[list[Any]] = []
    best_sheet = wb.sheet_by_index(0).name if wb.nsheets > 0 else "sheet0"
    best_score = -1
    for i in range(wb.nsheets):
        sheet = wb.sheet_by_index(i)
        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
            if len(rows) >= 300000:
                break
        if not rows:
            continue
        score = -1
        try:
            _, colmap = _detect_header_map(rows[:120])
            score = len(colmap)
        except HTTPException:
            pass
        if score > best_score:
            best_rows = rows
            best_sheet = _to_text(getattr(sheet, "name", "")).strip() or best_sheet
            best_score = score
    return best_rows, best_sheet


def parse_boq_upload(file_name: str, content: bytes) -> list[BoqItem]:
    if not content:
        raise HTTPException(400, "Upload file is empty.")

    name = _to_text(file_name).strip()
    lower = name.lower()
    header_sig = content[:4] if content else b""
    is_ole2 = header_sig == b"\xD0\xCF\x11\xE0"
    is_zip = header_sig == b"\x50\x4B\x03\x04"

    if lower.endswith(".csv"):
        rows = _rows_from_csv_bytes(content)
        source_sheet = "csv"
    elif lower.endswith(".xlsx") or is_zip:
        rows, source_sheet = _rows_from_xlsx_bytes(content)
    elif lower.endswith(".xls") or is_ole2:
        rows, source_sheet = _rows_from_xls_bytes(content)
    else:
        rows = _rows_from_csv_bytes(content)
        source_sheet = "unknown"

    if not rows:
        raise HTTPException(400, "Upload file is empty.")

    header_row, colmap = _detect_header_map(rows[:120])
    out: list[BoqItem] = []
    for row_idx in range(header_row + 1, len(rows)):
        row = rows[row_idx]
        item_no_raw = _to_text(row[colmap["item_no"]] if colmap["item_no"] < len(row) else "").strip()
        item_no = _normalize_item_no(item_no_raw)
        if not item_no or not ITEM_NO_PATTERN.match(item_no):
            continue
        name = _to_text(row[colmap["name"]] if colmap["name"] < len(row) else "").strip()
        unit = _normalize_unit(row[colmap.get("unit", -1)] if 0 <= colmap.get("unit", -1) < len(row) else "")
        division = _to_text(row[colmap.get("division", -1)] if 0 <= colmap.get("division", -1) < len(row) else "").strip()
        subdivision = _to_text(row[colmap.get("subdivision", -1)] if 0 <= colmap.get("subdivision", -1) < len(row) else "").strip()
        hierarchy_raw = _to_text(row[colmap.get("hierarchy", -1)] if 0 <= colmap.get("hierarchy", -1) < len(row) else "").strip()

        dq_raw = _to_text(row[colmap.get("design_quantity", -1)] if 0 <= colmap.get("design_quantity", -1) < len(row) else "").strip()
        up_raw = _to_text(row[colmap.get("unit_price", -1)] if 0 <= colmap.get("unit_price", -1) < len(row) else "").strip()
        aq_raw = _to_text(row[colmap.get("approved_quantity", -1)] if 0 <= colmap.get("approved_quantity", -1) < len(row) else "").strip()
        aa_raw = _to_text(row[colmap.get("approved_amount", -1)] if 0 <= colmap.get("approved_amount", -1) < len(row) else "").strip()
        remark = _to_text(row[colmap.get("remark", -1)] if 0 <= colmap.get("remark", -1) < len(row) else "").strip()

        design_quantity = _round4(_to_float(dq_raw))
        unit_price = _round4(_to_float(up_raw))
        approved_quantity = _round4(_to_float(aq_raw))
        approved_amount = _round4(_to_float(aa_raw))
        approved_quantity, used_amount = _fallback_approved_quantity(
            item_no=item_no,
            unit=unit,
            approved_quantity=approved_quantity,
            approved_amount=approved_amount,
            design_quantity=design_quantity,
            unit_price=unit_price,
        )
        if used_amount:
            aq_raw = aa_raw
            if not unit:
                unit = "CNY"

        out.append(
            BoqItem(
                item_no=item_no,
                name=name,
                unit=unit,
                division=division,
                subdivision=subdivision,
                hierarchy_raw=hierarchy_raw,
                design_quantity=design_quantity,
                design_quantity_raw=dq_raw,
                unit_price=unit_price,
                unit_price_raw=up_raw,
                approved_quantity=approved_quantity,
                approved_quantity_raw=aq_raw,
                remark=remark,
                row_index=row_idx + 1,
                sheet_name=source_sheet,
            )
        )

    if not out:
        raise HTTPException(400, "No BOQ item rows parsed from upload file.")
    return out


def boq_items_to_dict(rows: list[BoqItem]) -> list[dict[str, Any]]:
    return [item.__dict__.copy() for item in rows]


__all__ = ["parse_boq_upload", "boq_items_to_dict"]
