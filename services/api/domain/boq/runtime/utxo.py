"""
BOQ parsing and UTXO initialization helpers.

Implements BOQ -> UTXO genesis workflow:
- parse Excel BOQ rows
- map each line item to sovereign v:// URI
- compute deterministic genesis hash
- optionally persist INITIAL UTXO into proof_utxo
"""

from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import UTC, datetime
import csv
import hashlib
import io
import json
from pathlib import Path
import re
from typing import Any

from services.api.domain.utxo.integrations import ProofUTXOEngine
from services.api.domain.boq.runtime.specdict_gate import (
    resolve_gate_binding,
    save_spec_dict,
    upsert_gate_binding,
)
from services.api.domain.specir.runtime.registry import (
    ensure_specir_object,
)
from services.api.domain.specir.runtime.refs import (
    resolve_spu_ref_pack,
)
from services.api.domain.specir.runtime.spu_schema import (
    build_spu_ultimate_content,
)
from services.api.domain.specir.runtime.layering import (
    build_project_boq_item_ref,
    resolve_standard_spu_snapshot,
)
from services.api.domain.projects.gitpeg_sdk import (
    register_boq_item,
)

ITEM_NO_PATTERN = re.compile(r"^\d{3}(?:-[0-9A-Za-z]+)*$")
CODE_HINT_PATTERN = re.compile(r"(?<!\d)([1-9]\d{2})(?!\d)")

NORMREF_BY_PREFIX: dict[str, list[str]] = {
    "401": ["JTG F80/1-2017"],
    "403": ["JTG F80/1-2017"],
    "405": ["JTG F80/1-2017"],
    "600": ["JTG F80/1-2017"],
    "702": ["JTG F80/1-2017"],
}


def _normalize_item_no(value: Any) -> str:
    text = _to_text(value).strip()
    if not text:
        return ""
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[./]", "-", text)
    text = re.sub(r"[()]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text


@dataclass(slots=True)
class BoqItem:
    item_no: str
    name: str
    unit: str
    division: str
    subdivision: str
    hierarchy_raw: str
    design_quantity: float | None
    design_quantity_raw: str
    unit_price: float | None
    unit_price_raw: str
    approved_quantity: float | None
    approved_quantity_raw: str
    remark: str
    row_index: int
    sheet_name: str


@dataclass(slots=True)
class BoqHierarchyNode:
    code: str
    uri: str
    norm_context_uri: str
    parent_code: str
    parent_uri: str
    depth: int
    node_type: str
    is_leaf: bool
    name: str
    unit: str
    design_quantity: float | None
    approved_quantity: float | None
    unit_price: float | None
    division: str
    subdivision: str
    hierarchy_raw: str
    source_file: str
    source_sheet: str
    source_row: int | None
    children_codes: list[str]
    children_uris: list[str]
    children_merkle_root: str
    node_hash: str
    subtree_hash: str
    linked_gate_id: str
    linked_gate_ids: list[str]
    linked_gate_rules: list[dict[str, Any]]
    linked_spec_uri: str
    spec_dict_key: str
    spec_item: str
    ref_gate_uri: str
    ref_gate_uris: list[str]
    ref_spec_uri: str
    ref_spec_dict_uri: str
    ref_spec_item_uri: str
    ref_spu_uri: str
    ref_quota_uri: str
    ref_meter_rule_uri: str
    gate_template_lock: bool
    hierarchy: dict[str, Any]
    bridge_uri: str
    genesis_hash: str


def _to_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    text = _to_text(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            return float(match.group(0))
        except Exception:
            return None


def _round4(value: float | None) -> float | None:
    if value is None:
        return None
    try:
        return round(float(value), 4)
    except Exception:
        return None


def _normalize_unit(value: Any) -> str:
    text = _to_text(value).strip()
    if not text:
        return ""
    norm = text.replace("平方米", "m2").replace("㎡", "m2").replace("m²", "m2").replace("M2", "m2").replace("m^2", "m2")
    norm = norm.replace("立方米", "m3").replace("m³", "m3").replace("M3", "m3").replace("m^3", "m3")
    norm = norm.replace("公里", "km").replace("千米", "km").replace("KM", "km").replace("Km", "km")
    norm = norm.replace("米", "m").replace("M", "m")
    if norm in {"金额", "费用", "价款"}:
        norm = "总额"
    return norm.strip()


def _unit_from_meter_rule_uri(uri: str) -> str:
    ref = _to_text(uri).strip().lower()
    if "/by-weight@" in ref:
        return "t"
    if "/by-volume@" in ref:
        return "m3"
    if "/by-area@" in ref:
        return "m2"
    if "/contract-payment@" in ref:
        return "CNY"
    if "/landscape-work@" in ref:
        return "m2"
    return ""


def _fallback_approved_quantity(
    *,
    item_no: str,
    unit: str,
    approved_quantity: float | None,
    approved_amount: float | None,
    design_quantity: float | None,
    unit_price: float | None,
) -> tuple[float | None, bool]:
    if approved_amount is None:
        return approved_quantity, False
    if approved_quantity is not None and approved_quantity > 0:
        return approved_quantity, False
    unit_norm = _normalize_unit(unit)
    if unit_norm in {"总额", "金额"} or not unit_norm:
        return approved_amount, True
    if _to_text(item_no).strip().startswith("1"):
        return approved_amount, True
    if unit_price is None and design_quantity is None:
        return approved_amount, True
    return approved_quantity, False


def _extract_code_hint(*values: Any) -> str:
    for raw in values:
        text = _to_text(raw).strip()
        if not text:
            continue
        match = CODE_HINT_PATTERN.search(text)
        if match:
            return match.group(1)
    return ""


def _default_norm_refs(*, item_no: str, item_name: str = "") -> list[str]:
    code = _to_text(item_no).strip()
    prefix = code.split("-")[0] if code else ""
    if prefix in NORMREF_BY_PREFIX:
        return list(NORMREF_BY_PREFIX[prefix])
    name = _to_text(item_name).strip()
    if "桥" in name:
        return list(NORMREF_BY_PREFIX.get("401") or [])
    if "绿化" in name:
        return list(NORMREF_BY_PREFIX.get("600") or [])
    if "种植" in name:
        return list(NORMREF_BY_PREFIX.get("702") or [])
    if "钢筋" in name:
        return list(NORMREF_BY_PREFIX.get("403") or [])
    return []


def _normalize_header(value: Any) -> str:
    text = _to_text(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("號", "号").replace("編", "编").replace("碼", "码")
    text = text.replace("稱", "称").replace("單", "单").replace("價", "价")
    text = re.sub(r"[\s\u3000]+", "", text)
    text = re.sub(r"[:：_/\\-]", "", text)
    return text


def _detect_header_map(rows: list[tuple[int, list[Any]]]) -> tuple[int, dict[str, int]]:
    aliases = {
        "item_no": {
            "子目号",
            "子目號",
            "子目编号",
            "子目編號",
            "细目号",
            "細目號",
            "细目",
            "細目",
            "itemno",
            "item",
            "itemcode",
        },
        "name": {"子目名称", "子目名稱", "细目名称", "細目名稱", "名称", "名稱", "项目名称", "項目名稱", "name"},
        "unit": {"单位", "單位", "unit"},
        "division": {"分部工程", "分部", "division"},
        "subdivision": {"分项工程", "分項工程", "子目", "subdivision"},
        "hierarchy": {"所属分部分项层级", "所屬分部分項層級", "分部分项层级", "分部分項層級", "层级", "層級", "hierarchy", "wbs"},
        "design_quantity": {
            "设计数量",
            "設計數量",
            "设计工程量",
            "設計工程量",
            "设计",
            "設計",
            "施工图数量",
            "施工圖數量",
            "施工图复核数量",
            "施工圖復核數量",
            "施工图复核工程量",
            "designqty",
            "designquantity",
        },
        "unit_price": {"单价", "單價", "综合单价", "綜合單價", "price", "unitprice"},
        "approved_quantity": {
            "批复数量",
            "批復數量",
            "批复工程量",
            "批復工程量",
            "审批数量",
            "審批數量",
            "审批工程量",
            "審批工程量",
            "批准数量",
            "批准數量",
            "批准工程量",
            "批准工程量",
            "合同数量",
            "合同數量",
            "合同工程量",
            "审定数量",
            "審定數量",
            "审定工程量",
            "審定工程量",
            "批复",
            "批復",
            "approvedqty",
            "approvedquantity",
        },
        "approved_amount": {
            "批复金额",
            "批復金額",
            "批复合价",
            "批復合價",
            "审定金额",
            "審定金額",
            "审定合价",
            "審定合價",
            "批准金额",
            "批准金額",
            "合同金额",
            "合同金額",
            "合同合价",
            "合同合價",
            "合价",
            "合價",
            "总价",
            "總價",
            "总额",
            "總額",
            "金额",
            "金額",
        },
        "remark": {"备注", "備註", "remark"},
    }
    normalized_aliases: dict[str, set[str]] = {
        field: {_normalize_header(alias) for alias in names}
        for field, names in aliases.items()
    }

    def _header_match(key: str, alias: str) -> bool:
        if not key or not alias:
            return False
        if key == alias:
            return True
        if len(alias) >= 3 and alias in key:
            return True
        if len(key) >= 4 and key in alias:
            return True
        return False

    best_row = -1
    best_score = -1
    best_map: dict[str, int] = {}

    for row_index, row_values in rows:
        mapping: dict[str, int] = {}
        score = 0
        for col_index, cell in enumerate(row_values):
            key = _normalize_header(cell)
            if not key:
                continue
            for field, names in normalized_aliases.items():
                if field in mapping:
                    continue
                if any(_header_match(key, alias) for alias in names):
                    mapping[field] = col_index
                    score += 1
        if score > best_score and "item_no" in mapping and "name" in mapping:
            best_row = row_index
            best_score = score
            best_map = mapping

    if best_row < 0:
        raise ValueError("Failed to detect BOQ header row (need at least item_no + name columns).")
    return best_row, best_map


def _merge_header_rows(primary: list[Any], secondary: list[Any]) -> list[Any]:
    merged: list[Any] = []
    max_len = max(len(primary), len(secondary))
    for i in range(max_len):
        a = _to_text(primary[i] if i < len(primary) else "").strip()
        b = _to_text(secondary[i] if i < len(secondary) else "").strip()
        if a and b:
            merged.append(f"{a}{b}")
        elif b:
            merged.append(b)
        else:
            merged.append(a)
    return merged


def _normalize_sheet_targets(sheet_name: Any) -> list[str]:
    if not sheet_name:
        return []
    if isinstance(sheet_name, (list, tuple, set)):
        return [str(x).strip() for x in sheet_name if str(x).strip()]
    text = _to_text(sheet_name).strip()
    if not text:
        return []
    if "," in text:
        return [seg.strip() for seg in text.split(",") if seg.strip()]
    return [text]


def _resolve_sheet_names(requested: list[str], available: list[str]) -> list[str]:
    if not requested:
        return []
    resolved: list[str] = []
    for req in requested:
        if req in available:
            resolved.append(req)
            continue
        hint = _extract_code_hint(req)
        if hint:
            matches = [name for name in available if hint in _to_text(name)]
            if matches:
                resolved.extend(matches)
                continue
        matches = [name for name in available if req in _to_text(name)]
        if matches:
            resolved.extend(matches)
            continue
    seen: set[str] = set()
    ordered: list[str] = []
    for name in resolved:
        if name in seen:
            continue
        seen.add(name)
        ordered.append(name)
    return ordered


def _is_quantified_item(item: BoqItem) -> bool:
    return item.design_quantity is not None or item.approved_quantity is not None or item.unit_price is not None


def _auto_split_items(
    items: list[BoqItem],
    *,
    min_depth: int,
    leaf_only: bool,
) -> list[BoqItem]:
    if min_depth <= 1 or not items:
        return items

    by_code: dict[str, BoqItem] = {item.item_no: item for item in items if item.item_no}
    expanded: list[BoqItem] = list(items)

    def _parts(code: str) -> list[str]:
        return [seg.strip() for seg in _to_text(code).split("-") if seg.strip()]

    # Detect existing children first.
    has_child: set[str] = set()
    for code in list(by_code.keys()):
        parts = _parts(code)
        for i in range(1, len(parts)):
            has_child.add("-".join(parts[:i]))

    for item in list(items):
        code = _to_text(item.item_no).strip()
        if not code:
            continue
        parts = _parts(code)
        if len(parts) >= min_depth:
            continue
        if code in has_child:
            continue
        if not _is_quantified_item(item):
            continue

        new_code = code
        while len(_parts(new_code)) < min_depth:
            new_code = f"{new_code}-1"
        if new_code in by_code:
            continue

        synthetic = BoqItem(
            item_no=new_code,
            name=item.name,
            unit=item.unit,
            division=item.division,
            subdivision=item.subdivision,
            hierarchy_raw=item.hierarchy_raw,
            design_quantity=item.design_quantity,
            design_quantity_raw=item.design_quantity_raw,
            unit_price=item.unit_price,
            unit_price_raw=item.unit_price_raw,
            approved_quantity=item.approved_quantity,
            approved_quantity_raw=item.approved_quantity_raw,
            remark=(item.remark or "AUTO_SPLIT"),
            row_index=item.row_index,
            sheet_name=item.sheet_name,
        )
        by_code[new_code] = synthetic
        expanded.append(synthetic)
        has_child.add(code)

    if not leaf_only:
        return expanded

    # Recompute children after expansion.
    has_child = set()
    for code in [item.item_no for item in expanded if item.item_no]:
        parts = _parts(code)
        for i in range(1, len(parts)):
            has_child.add("-".join(parts[:i]))

    return [item for item in expanded if item.item_no not in has_child]


def _parse_openpyxl_sheet(
    ws: Any,
    *,
    sheet_name: str,
    leaf_only: bool,
) -> list[BoqItem]:
    probe_rows = [
        (idx, list(row))
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=120, values_only=True), start=1)
    ]
    header_row, colmap = _detect_header_map(probe_rows)
    uses_double_header = False
    try:
        header_cells = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
        sub_header_cells = list(next(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True)))
        if header_cells and sub_header_cells and any(_to_text(x).strip() for x in sub_header_cells):
            merged = _merge_header_rows(header_cells, sub_header_cells)
            _, merged_map = _detect_header_map([(header_row, merged)])
            if merged_map:
                colmap = merged_map
                uses_double_header = True
    except Exception:
        uses_double_header = False

    out: list[BoqItem] = []
    start_row = header_row + (2 if uses_double_header else 1)
    for row_index, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        item = _parse_boq_row(
            cells=list(row),
            colmap=colmap,
            row_index=row_index,
            sheet_name=sheet_name,
            leaf_only=leaf_only,
        )
        if item is not None:
            out.append(item)
    return out


def _parse_xlrd_sheet(
    ws: Any,
    *,
    sheet_name: str,
    leaf_only: bool,
) -> list[BoqItem]:
    probe_rows = [(idx + 1, ws.row_values(idx)) for idx in range(min(120, ws.nrows))]
    header_row, colmap = _detect_header_map(probe_rows)
    uses_double_header = False
    if header_row < ws.nrows:
        header_cells = ws.row_values(header_row - 1)
        sub_cells = ws.row_values(header_row) if header_row < ws.nrows else []
        if header_cells and sub_cells and any(_to_text(x).strip() for x in sub_cells):
            merged = _merge_header_rows(header_cells, sub_cells)
            try:
                _, merged_map = _detect_header_map([(header_row, merged)])
                if merged_map:
                    colmap = merged_map
                    uses_double_header = True
            except Exception:
                uses_double_header = False

    out: list[BoqItem] = []
    start_row = header_row + (1 if uses_double_header else 0)
    for row_idx in range(start_row, ws.nrows):
        item = _parse_boq_row(
            cells=ws.row_values(row_idx),
            colmap=colmap,
            row_index=row_idx + 1,
            sheet_name=sheet_name,
            leaf_only=leaf_only,
        )
        if item is not None:
            out.append(item)
    return out


def _rows_from_csv_path(source_path: Path) -> list[list[Any]]:
    if not source_path.exists():
        return []
    raw = source_path.read_bytes()
    text = ""
    for enc in ("utf-8-sig", "gb18030", "gbk"):
        try:
            text = raw.decode(enc)
            break
        except Exception:
            continue
    if not text:
        text = raw.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def _parse_csv_rows(
    rows: list[list[Any]],
    *,
    sheet_name: str,
    leaf_only: bool,
) -> list[BoqItem]:
    if not rows:
        return []
    probe_rows = [(idx + 1, row) for idx, row in enumerate(rows[:120])]
    header_row, colmap = _detect_header_map(probe_rows)
    uses_double_header = False
    if header_row <= len(rows):
        header_cells = rows[header_row - 1]
        sub_cells = rows[header_row] if header_row < len(rows) else []
        if header_cells and sub_cells and any(_to_text(x).strip() for x in sub_cells):
            merged = _merge_header_rows(header_cells, sub_cells)
            try:
                _, merged_map = _detect_header_map([(header_row, merged)])
                if merged_map:
                    colmap = merged_map
                    uses_double_header = True
            except Exception:
                uses_double_header = False

    out: list[BoqItem] = []
    start_row = header_row + (1 if uses_double_header else 0)
    for row_idx in range(start_row, len(rows)):
        item = _parse_boq_row(
            cells=rows[row_idx],
            colmap=colmap,
            row_index=row_idx + 1,
            sheet_name=sheet_name,
            leaf_only=leaf_only,
        )
        if item is not None:
            out.append(item)
    return out


def _cell_text(cells: list[Any], colmap: dict[str, int], key: str) -> str:
    idx = colmap.get(key)
    if idx is None or idx >= len(cells):
        return ""
    return _to_text(cells[idx]).strip()


def _parse_boq_row(
    *,
    cells: list[Any],
    colmap: dict[str, int],
    row_index: int,
    sheet_name: str,
    leaf_only: bool,
) -> BoqItem | None:
    item_no_raw = _cell_text(cells, colmap, "item_no")
    item_no = _normalize_item_no(item_no_raw)
    if not item_no or not ITEM_NO_PATTERN.match(item_no):
        return None

    name = _cell_text(cells, colmap, "name")
    unit = _normalize_unit(_cell_text(cells, colmap, "unit"))
    division = _cell_text(cells, colmap, "division")
    subdivision = _cell_text(cells, colmap, "subdivision")
    hierarchy_raw = _cell_text(cells, colmap, "hierarchy")

    dq_raw = _cell_text(cells, colmap, "design_quantity")
    up_raw = _cell_text(cells, colmap, "unit_price")
    aq_raw = _cell_text(cells, colmap, "approved_quantity")
    aa_raw = _cell_text(cells, colmap, "approved_amount")

    design_quantity = _round4(_to_float(dq_raw))
    unit_price = _round4(_to_float(up_raw))
    approved_quantity = _round4(_to_float(aq_raw))
    approved_amount = _round4(_to_float(aa_raw))
    approved_quantity, used_amount = _fallback_approved_quantity(
        item_no=item_no,
        unit=unit,
        approved_quantity=approved_quantity,
        approved_amount=approved_amount,
        design_quantity=design_quantity,
        unit_price=unit_price,
    )
    if used_amount:
        aq_raw = aa_raw
        if not unit:
            unit = "总额"

    if leaf_only and design_quantity is None and approved_quantity is None:
        return None

    return BoqItem(
        item_no=item_no,
        name=name,
        unit=unit,
        division=division,
        subdivision=subdivision,
        hierarchy_raw=hierarchy_raw,
        design_quantity=design_quantity,
        design_quantity_raw=dq_raw,
        unit_price=unit_price,
        unit_price_raw=up_raw,
        approved_quantity=approved_quantity,
        approved_quantity_raw=aq_raw,
        remark=_cell_text(cells, colmap, "remark"),
        row_index=row_index,
        sheet_name=sheet_name,
    )

def parse_boq_excel(
    xlsx_path: str | Path,
    *,
    sheet_name: str | None = None,
    leaf_only: bool = True,
    auto_split_min_depth: int = 3,
) -> list[BoqItem]:
    source_path = Path(xlsx_path).expanduser().resolve()
    if not source_path.exists():
        raise FileNotFoundError(f"BOQ file not found: {source_path}")

    if source_path.suffix.lower() == ".csv":
        rows = _rows_from_csv_path(source_path)
        out = _parse_csv_rows(rows, sheet_name=sheet_name or source_path.stem or "CSV", leaf_only=leaf_only)
        if not out:
            raise ValueError("Failed to detect BOQ header row (need at least item_no + name columns).")
        return _auto_split_items(out, min_depth=auto_split_min_depth, leaf_only=leaf_only)

    use_xls = source_path.suffix.lower() == ".xls"
    if not use_xls:
        try:
            import openpyxl
        except Exception as exc:
            raise RuntimeError("openpyxl is required for BOQ parsing.") from exc
        try:
            wb = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
            available = list(wb.sheetnames)
            requested = _normalize_sheet_targets(sheet_name)
            targets = _resolve_sheet_names(requested, available) if requested else []
            if not targets:
                targets = list(available)

            out: list[BoqItem] = []
            for name in targets:
                if name not in wb.sheetnames:
                    continue
                ws = wb[name]
                try:
                    out.extend(_parse_openpyxl_sheet(ws, sheet_name=name, leaf_only=leaf_only))
                except Exception:
                    continue
            if out:
                return _auto_split_items(out, min_depth=auto_split_min_depth, leaf_only=leaf_only)
            raise ValueError("Failed to detect BOQ header row (need at least item_no + name columns).")
        except Exception:
            use_xls = True

    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise RuntimeError("xlrd is required for .xls BOQ parsing.") from exc

    wb = xlrd.open_workbook(source_path)
    available = list(wb.sheet_names())
    requested = _normalize_sheet_targets(sheet_name)
    targets = _resolve_sheet_names(requested, available) if requested else []
    if not targets:
        targets = list(available)
    if not targets:
        raise RuntimeError("No worksheet found in BOQ xls file.")

    out: list[BoqItem] = []
    for name in targets:
        try:
            ws = wb.sheet_by_name(name)
        except Exception:
            continue
        try:
            out.extend(_parse_xlrd_sheet(ws, sheet_name=name, leaf_only=leaf_only))
        except Exception:
            continue

    if not out:
        raise ValueError("Failed to detect BOQ header row (need at least item_no + name columns).")
    return _auto_split_items(out, min_depth=auto_split_min_depth, leaf_only=leaf_only)


def _derive_hierarchy(item: BoqItem) -> dict[str, Any]:
    codes = [seg.strip() for seg in _to_text(item.item_no).split("-") if seg.strip()]
    chapter_hint = _extract_code_hint(item.division, item.hierarchy_raw, item.subdivision, item.sheet_name)
    section_hint = _extract_code_hint(item.subdivision, item.hierarchy_raw)
    chapter = chapter_hint or (codes[0] if codes else "")
    section = section_hint or (codes[0] if codes else chapter)
    subgroup = "-".join(codes[:3]) if len(codes) >= 3 else (section or chapter)

    labels: list[str] = []
    if item.division:
        labels.append(item.division)
    if item.subdivision and item.subdivision not in labels:
        labels.append(item.subdivision)
    if item.hierarchy_raw and item.hierarchy_raw not in labels:
        labels.append(item.hierarchy_raw)
    if item.name and item.name not in labels:
        labels.append(item.name)

    return {
        "chapter_code": chapter,
        "section_code": section,
        "subgroup_code": subgroup,
        "item_code": item.item_no,
        "code_parts": codes,
        "level": len(codes),
        "division": item.division,
        "subdivision": item.subdivision,
        "hierarchy_raw": item.hierarchy_raw,
        "wbs_path": " / ".join(labels) if labels else item.item_no,
    }


def _item_code_parts(item_no: str) -> list[str]:
    return [seg.strip() for seg in _to_text(item_no).split("-") if seg.strip()]


def _item_code_sort_key(code: str) -> tuple[int, list[int], str]:
    parts = _item_code_parts(code)
    nums: list[int] = []
    for part in parts:
        try:
            nums.append(int(part))
        except Exception:
            nums.append(999999)
    return (len(parts), nums, _to_text(code))


def _direct_parent_code(code: str) -> str:
    parts = _item_code_parts(code)
    if len(parts) <= 1:
        return ""
    return "-".join(parts[:-1])


def _node_type_for_depth(depth: int, max_depth: int) -> str:
    if depth <= 1:
        return "chapter"
    if depth == 2:
        return "section"
    if depth == 3:
        return "item"
    if depth == max_depth:
        return "detail"
    return f"level_{depth}"


def _default_node_name(*, code: str, node_type: str) -> str:
    if node_type == "chapter":
        return f"第{code}章"
    if node_type == "section":
        return f"{code} 节"
    if node_type == "item":
        return f"{code} 目"
    if node_type == "detail":
        return f"{code} 细目"
    return f"{code} 分组"


DEFAULT_QCGATE_RULE_LIBRARY: dict[str, dict[str, Any]] = {
    "gate_403_rebar_hrb400_yield": {
        "gate_name": "HRB400 钢筋屈服强度门控",
        "spec_uri": "v://norm/GB50204@2015/5.3.2#diameter_tolerance",
        "metric": "yield_strength_or_diameter_deviation",
        "unit": "mm",
    },
    "gate_403_rebar_spacing": {
        "gate_name": "钢筋间距偏差门控",
        "spec_uri": "v://norm/GB50204@2015/5.3.3#spacing_tolerance",
        "metric": "spacing_deviation",
        "unit": "mm",
    },
    "gate_403_rebar_crack": {
        "gate_name": "裂缝宽度门控",
        "spec_uri": "v://norm/JTG_F80@2017/4.3#crack_width_max",
        "metric": "crack_width",
        "unit": "mm",
    },
    "gate_400_generic_quality": {
        "gate_name": "400章通用质量门控",
        "spec_uri": "",
        "metric": "generic_quality",
        "unit": "",
    },
}

DEFAULT_QCGATE_BINDINGS: dict[str, dict[str, list[str]]] = {
    "exact": {
        "403-1-2": ["gate_403_rebar_hrb400_yield", "gate_403_rebar_spacing"],
    },
    "prefix": {
        "403-1": ["gate_403_rebar_spacing"],
        "403": ["gate_403_rebar_crack"],
    },
    "chapter": {
        "400": ["gate_400_generic_quality"],
    },
}


def _item_code_from_boq_uri(boq_item_uri: str) -> str:
    uri = _to_text(boq_item_uri).strip().rstrip("/")
    if not uri:
        return ""
    return uri.split("/")[-1]


def _safe_ref_token(value: Any) -> str:
    text = _to_text(value).strip()
    if not text:
        return ""
    text = text.replace("::", "/").replace("#", "/").replace("@", "-")
    text = re.sub(r"[^0-9A-Za-z._/-]+", "-", text).strip("-/")
    text = re.sub(r"/{2,}", "/", text)
    return text[:180]


def _build_binding_refs(binding: dict[str, Any]) -> dict[str, Any]:
    gate_ids = [_to_text(x).strip() for x in list(binding.get("linked_gate_ids") or []) if _to_text(x).strip()]
    linked_gate_id = _to_text(binding.get("linked_gate_id") or "").strip()
    linked_spec_uri = _to_text(binding.get("linked_spec_uri") or "").strip()
    spec_dict_key = _to_text(binding.get("spec_dict_key") or "").strip()
    spec_item = _to_text(binding.get("spec_item") or "").strip()

    if linked_gate_id and linked_gate_id not in gate_ids:
        gate_ids.insert(0, linked_gate_id)
    ref_gate_uris = [f"v://norm/gate/{_safe_ref_token(gate_id)}@v1" for gate_id in gate_ids if _safe_ref_token(gate_id)]
    ref_gate_uri = ref_gate_uris[0] if ref_gate_uris else ""

    ref_spec_uri = linked_spec_uri if linked_spec_uri.startswith("v://") else ""
    ref_spec_dict_uri = f"v://norm/specdict/{_safe_ref_token(spec_dict_key)}@v1" if spec_dict_key else ""
    if spec_item and ref_spec_dict_uri:
        ref_spec_item_uri = f"{ref_spec_dict_uri}#{_safe_ref_token(spec_item)}"
    elif spec_item and ref_spec_uri:
        ref_spec_item_uri = f"{ref_spec_uri}#{_safe_ref_token(spec_item)}"
    else:
        ref_spec_item_uri = ""
    return {
        "ref_gate_uri": ref_gate_uri,
        "ref_gate_uris": ref_gate_uris,
        "ref_spec_uri": ref_spec_uri,
        "ref_spec_dict_uri": ref_spec_dict_uri,
        "ref_spec_item_uri": ref_spec_item_uri,
    }


def _register_specir_refs(
    *,
    sb: Any,
    item_code: str,
    binding: dict[str, Any],
) -> None:
    if sb is None:
        return
    refs = _build_binding_refs(binding)
    ref_spec_uri = _to_text(refs.get("ref_spec_uri") or "").strip()
    ref_spec_dict_uri = _to_text(refs.get("ref_spec_dict_uri") or "").strip()
    ref_spec_item_uri = _to_text(refs.get("ref_spec_item_uri") or "").strip()
    ref_gate_uris = [_to_text(x).strip() for x in list(refs.get("ref_gate_uris") or []) if _to_text(x).strip()]
    linked_gate_ids = [_to_text(x).strip() for x in list(binding.get("linked_gate_ids") or []) if _to_text(x).strip()]
    linked_gate_rules = list(binding.get("linked_gate_rules") or [])

    try:
        if ref_spec_uri:
            ensure_specir_object(
                sb=sb,
                uri=ref_spec_uri,
                kind="spec_rule",
                title=f"Spec Rule {item_code}",
                content={
                    "item_code": item_code,
                    "spec_dict_key": _to_text(binding.get("spec_dict_key") or "").strip(),
                    "spec_item": _to_text(binding.get("spec_item") or "").strip(),
                },
                metadata={"source": "boq.resolve_linked_gates"},
            )
        if ref_spec_dict_uri:
            ensure_specir_object(
                sb=sb,
                uri=ref_spec_dict_uri,
                kind="spec_dict",
                title=f"SpecDict {item_code}",
                content={
                    "spec_dict_key": _to_text(binding.get("spec_dict_key") or "").strip(),
                    "spec_item": _to_text(binding.get("spec_item") or "").strip(),
                    "spec_uri": ref_spec_uri,
                },
                metadata={"source": "boq.resolve_linked_gates"},
            )
        if ref_spec_item_uri:
            ensure_specir_object(
                sb=sb,
                uri=ref_spec_item_uri,
                kind="spec_item",
                title=f"SpecItem {item_code}",
                content={
                    "spec_dict_key": _to_text(binding.get("spec_dict_key") or "").strip(),
                    "spec_item": _to_text(binding.get("spec_item") or "").strip(),
                    "spec_uri": ref_spec_uri,
                },
                metadata={"source": "boq.resolve_linked_gates"},
            )
        for idx, ref_gate_uri in enumerate(ref_gate_uris):
            gate_id = linked_gate_ids[idx] if idx < len(linked_gate_ids) else ""
            ensure_specir_object(
                sb=sb,
                uri=ref_gate_uri,
                kind="gate",
                title=f"Gate {gate_id or item_code}",
                content={
                    "gate_id": gate_id,
                    "item_code": item_code,
                    "spec_uri": ref_spec_uri,
                    "spec_item_uri": ref_spec_item_uri,
                    "spec_dict_uri": ref_spec_dict_uri,
                    "gate_rule": linked_gate_rules[idx] if idx < len(linked_gate_rules) else {},
                },
                metadata={"source": "boq.resolve_linked_gates"},
            )
    except Exception:
        return


def _register_specir_spu_refs(
    *,
    sb: Any,
    item_code: str,
    item_name: str,
    ref_pack: dict[str, Any],
) -> None:
    if sb is None:
        return
    ref_spu_uri = _to_text(ref_pack.get("ref_spu_uri") or "").strip()
    ref_quota_uri = _to_text(ref_pack.get("ref_quota_uri") or "").strip()
    ref_meter_rule_uri = _to_text(ref_pack.get("ref_meter_rule_uri") or "").strip()
    try:
        if ref_spu_uri:
            ensure_specir_object(
                sb=sb,
                uri=ref_spu_uri,
                kind="spu",
                title=item_name or item_code,
                content=build_spu_ultimate_content(
                    spu_uri=ref_spu_uri,
                    title=item_name or item_code,
                    content={
                        "industry": "Highway",
                        "unit": _unit_from_meter_rule_uri(ref_meter_rule_uri),
                        "measure_statement": "Auto-registered from BOQ refs; refine in SpecIR editor.",
                        "measure_operator": "boq-auto-register",
                        "measure_expression": "approved_quantity",
                        "quota_ref": ref_quota_uri,
                        "meter_rule_ref": ref_meter_rule_uri,
                        "gate_refs": [],
                        "standard_codes": _default_norm_refs(item_no=item_code, item_name=item_name),
                    },
                ),
                metadata={"source": "boq.resolve_spu_ref_pack"},
            )
        if ref_quota_uri:
            ensure_specir_object(
                sb=sb,
                uri=ref_quota_uri,
                kind="quota",
                title=f"Quota {item_code}",
                content={"spu_ref": ref_spu_uri, "item_code": item_code},
                metadata={"source": "boq.resolve_spu_ref_pack"},
            )
        if ref_meter_rule_uri:
            ensure_specir_object(
                sb=sb,
                uri=ref_meter_rule_uri,
                kind="meter_rule",
                title=f"Meter Rule {item_code}",
                content={"spu_ref": ref_spu_uri, "item_code": item_code},
                metadata={"source": "boq.resolve_spu_ref_pack"},
            )
    except Exception:
        return


def resolve_linked_gates(
    *,
    item_code: str,
    fallback_spec_uri: str = "",
    sb: Any = None,
    persist_registry: bool = False,
) -> dict[str, Any]:
    code = _to_text(item_code).strip()
    fallback_spec = _to_text(fallback_spec_uri).strip()

    if sb is not None and code:
        try:
            db_binding = resolve_gate_binding(
                sb=sb,
                subitem_code=code,
                fallback_spec_uri=fallback_spec,
            )
            if bool(db_binding.get("from_registry")) and _to_text(db_binding.get("linked_gate_id") or "").strip():
                payload = {
                    "item_code": code,
                    "linked_gate_id": _to_text(db_binding.get("linked_gate_id") or "").strip(),
                    "linked_gate_ids": list(db_binding.get("linked_gate_ids") or []),
                    "linked_gate_rules": list(db_binding.get("linked_gate_rules") or []),
                    "linked_spec_uri": _to_text(db_binding.get("linked_spec_uri") or "").strip(),
                    "spec_dict_key": _to_text(db_binding.get("spec_dict_key") or "").strip(),
                    "spec_item": _to_text(db_binding.get("spec_item") or "").strip(),
                    "execution_strategy": _to_text(db_binding.get("execution_strategy") or "").strip(),
                    "fail_action": _to_text(db_binding.get("fail_action") or "").strip(),
                    "gate_template_lock": bool(db_binding.get("gate_template_lock")),
                    "gate_binding_hash": _to_text(db_binding.get("gate_binding_hash") or "").strip(),
                }
                payload.update(_build_binding_refs(payload))
                if persist_registry:
                    _register_specir_refs(sb=sb, item_code=code, binding=payload)
                return payload
        except Exception:
            pass

    linked_rules: list[dict[str, Any]] = []
    seen_gate_ids: set[str] = set()

    def _append_rules(match_kind: str, match_code: str, rule_keys: list[str]) -> None:
        for idx, key in enumerate(rule_keys):
            gate_meta = DEFAULT_QCGATE_RULE_LIBRARY.get(key) or {}
            gate_name = _to_text(gate_meta.get("gate_name") or key).strip()
            gate_id = f"QCGate::{code or match_code or 'UNKNOWN'}::{key}"
            if gate_id in seen_gate_ids:
                continue
            seen_gate_ids.add(gate_id)
            spec_uri = _to_text(gate_meta.get("spec_uri") or "").strip() or fallback_spec
            linked_rules.append(
                {
                    "gate_id": gate_id,
                    "gate_key": key,
                    "gate_name": gate_name,
                    "spec_uri": spec_uri,
                    "metric": _to_text(gate_meta.get("metric") or "").strip(),
                    "unit": _to_text(gate_meta.get("unit") or "").strip(),
                    "match_kind": match_kind,
                    "match_code": match_code,
                    "priority": idx,
                }
            )

    exact_rules = DEFAULT_QCGATE_BINDINGS.get("exact", {})
    prefix_rules = DEFAULT_QCGATE_BINDINGS.get("prefix", {})
    chapter_rules = DEFAULT_QCGATE_BINDINGS.get("chapter", {})

    if code in exact_rules:
        _append_rules("exact", code, list(exact_rules.get(code) or []))

    parts = _item_code_parts(code)
    for i in range(len(parts) - 1, 0, -1):
        prefix = "-".join(parts[:i])
        if prefix in prefix_rules:
            _append_rules("prefix", prefix, list(prefix_rules.get(prefix) or []))

    chapter = parts[0] if parts else ""
    if chapter and chapter in chapter_rules:
        _append_rules("chapter", chapter, list(chapter_rules.get(chapter) or []))

    if not linked_rules:
        fallback_gate_id = f"QCGate::{code or 'UNKNOWN'}::AUTO_DEFAULT"
        linked_rules.append(
            {
                "gate_id": fallback_gate_id,
                "gate_key": "auto_default",
                "gate_name": "自动门控规则",
                "spec_uri": fallback_spec,
                "metric": "",
                "unit": "",
                "match_kind": "fallback",
                "match_code": code,
                "priority": 0,
            }
        )

    linked_gate_ids = [_to_text(x.get("gate_id") or "").strip() for x in linked_rules if _to_text(x.get("gate_id") or "").strip()]
    linked_gate_id = linked_gate_ids[0] if linked_gate_ids else ""
    preferred_spec_uri = (
        _to_text(linked_rules[0].get("spec_uri") or "").strip()
        if linked_rules
        else fallback_spec
    )
    template_lock = bool(preferred_spec_uri or linked_gate_id)
    binding_hash = hashlib.sha256(
        json.dumps(
            {
                "item_code": code,
                "linked_gate_id": linked_gate_id,
                "linked_gate_ids": linked_gate_ids,
                "linked_gate_rules": linked_rules,
                "preferred_spec_uri": preferred_spec_uri,
                "template_lock": template_lock,
            },
            ensure_ascii=False,
            sort_keys=True,
            default=str,
        ).encode("utf-8")
    ).hexdigest()

    spec_dict_key = ""
    spec_item = ""
    if sb is not None and persist_registry and linked_rules:
        first_rule = linked_rules[0]
        first_spec_uri = _to_text(first_rule.get("spec_uri") or "").strip()
        parsed = re.match(
            r"^v://norm/(?P<code>[^/@#?]+)(?:@(?P<version>[^/#?]+))?(?:/(?P<path>[^#?]+))?(?:#(?P<fragment>[^?]+))?",
            first_spec_uri,
            flags=re.IGNORECASE,
        )
        if parsed:
            norm_code = _to_text(parsed.group("code") or "").strip().upper()
            norm_version = _to_text(parsed.group("version") or "").strip() or "v1.0"
            norm_path = _to_text(parsed.group("path") or "").strip().replace("/", ".")
            spec_item = _to_text(parsed.group("fragment") or "").strip() or _to_text(first_rule.get("metric") or "").strip()
            spec_dict_key = f"{norm_code}-{norm_version}-{norm_path}" if norm_path else f"{norm_code}-{norm_version}"
            try:
                save_spec_dict(
                    sb=sb,
                    spec_dict_key=spec_dict_key,
                    title=norm_code or "Norm SpecDict",
                    version=norm_version,
                    authority=norm_code,
                    spec_uri=first_spec_uri.split("#", 1)[0] if "#" in first_spec_uri else first_spec_uri,
                    items={
                        spec_item or "rule": {
                            "operator": _to_text(first_rule.get("operator") or "range").strip().lower() or "range",
                            "unit": _to_text(first_rule.get("unit") or "").strip(),
                            "mode": "absolute",
                            "default_threshold": None,
                            "context_rules": {},
                        }
                    },
                    metadata={"source": "resolve_linked_gates_fallback", "item_code": code},
                    is_active=True,
                )
                upsert_gate_binding(
                    sb=sb,
                    gate_id=linked_gate_id,
                    gate_id_base=linked_gate_id,
                    subitem_code=code,
                    spec_dict_key=spec_dict_key,
                    spec_item=spec_item or "rule",
                    match_kind="exact",
                    execution_strategy="all_pass",
                    fail_action="trigger_review_trip",
                    gate_rules=linked_rules,
                    metadata={"source": "resolve_linked_gates_fallback"},
                    is_active=True,
                )
            except Exception:
                pass

    payload = {
        "item_code": code,
        "linked_gate_id": linked_gate_id,
        "linked_gate_ids": linked_gate_ids,
        "linked_gate_rules": linked_rules,
        "linked_spec_uri": preferred_spec_uri,
        "spec_dict_key": spec_dict_key,
        "spec_item": spec_item,
        "execution_strategy": "",
        "fail_action": "",
        "gate_template_lock": template_lock,
        "gate_binding_hash": binding_hash,
    }
    payload.update(_build_binding_refs(payload))
    if persist_registry:
        _register_specir_refs(sb=sb, item_code=code, binding=payload)
    return payload


def auto_bind_gates(
    *,
    sb: Any,
    utxo_id: str,
) -> dict[str, Any]:
    proof_id = _to_text(utxo_id).strip()
    if not proof_id:
        raise ValueError("utxo_id is required")
    engine = ProofUTXOEngine(sb)
    row = engine.get_by_id(proof_id)
    if not row:
        raise ValueError("proof_utxo not found")
    state_data = row.get("state_data") if isinstance(row.get("state_data"), dict) else {}
    boq_item_uri = _to_text(
        state_data.get("boq_item_uri")
        or state_data.get("item_uri")
        or state_data.get("boq_uri")
        or row.get("segment_uri")
        or ""
    ).strip()
    item_code = _to_text(state_data.get("item_no") or _item_code_from_boq_uri(boq_item_uri)).strip()
    binding = resolve_linked_gates(
        item_code=item_code,
        fallback_spec_uri=_to_text(state_data.get("spec_uri") or row.get("norm_uri") or "").strip(),
        sb=sb,
        persist_registry=True,
    )
    refs = _build_binding_refs(binding)
    patched_state = dict(state_data)
    patched_state.update(
        {
            "linked_gate_id": _to_text(binding.get("linked_gate_id") or "").strip(),
            "linked_gate_ids": list(binding.get("linked_gate_ids") or []),
            "linked_spec_uri": _to_text(binding.get("linked_spec_uri") or "").strip(),
            "spec_dict_key": _to_text(binding.get("spec_dict_key") or "").strip(),
            "spec_item": _to_text(binding.get("spec_item") or "").strip(),
            "ref_gate_uri": _to_text(refs.get("ref_gate_uri") or "").strip(),
            "ref_gate_uris": list(refs.get("ref_gate_uris") or []),
            "ref_spec_uri": _to_text(refs.get("ref_spec_uri") or "").strip(),
            "ref_spec_dict_uri": _to_text(refs.get("ref_spec_dict_uri") or "").strip(),
            "ref_spec_item_uri": _to_text(refs.get("ref_spec_item_uri") or "").strip(),
            "gate_template_lock": bool(binding.get("gate_template_lock")),
            "gate_binding_hash": _to_text(binding.get("gate_binding_hash") or "").strip(),
        }
    )
    sb.table("proof_utxo").update({"state_data": patched_state}).eq("proof_id", proof_id).execute()
    return {
        "ok": True,
        "proof_id": proof_id,
        "boq_item_uri": boq_item_uri,
        "item_code": item_code,
        "linked_gate_id": patched_state.get("linked_gate_id"),
        "linked_gate_ids": patched_state.get("linked_gate_ids"),
        "ref_gate_uri": patched_state.get("ref_gate_uri"),
        "ref_spec_uri": patched_state.get("ref_spec_uri"),
        "spec_dict_key": patched_state.get("spec_dict_key"),
        "spec_item": patched_state.get("spec_item"),
        "gate_template_lock": patched_state.get("gate_template_lock"),
        "gate_binding_hash": patched_state.get("gate_binding_hash"),
    }


def _pick_node_name(
    *,
    code: str,
    node_type: str,
    explicit: BoqItem | None,
    leaf_ref: BoqItem | None,
) -> str:
    if explicit and _to_text(explicit.name).strip():
        return _to_text(explicit.name).strip()
    if explicit and _to_text(explicit.hierarchy_raw).strip():
        return _to_text(explicit.hierarchy_raw).strip()
    if node_type == "chapter" and explicit and _to_text(explicit.division).strip():
        return _to_text(explicit.division).strip()
    if node_type in {"section", "item"} and explicit and _to_text(explicit.subdivision).strip():
        return _to_text(explicit.subdivision).strip()
    if leaf_ref and _to_text(leaf_ref.division).strip() and node_type == "chapter":
        return _to_text(leaf_ref.division).strip()
    if leaf_ref and _to_text(leaf_ref.subdivision).strip() and node_type in {"section", "item"}:
        return _to_text(leaf_ref.subdivision).strip()
    return _default_node_name(code=code, node_type=node_type)


def _merkle_root(hashes: list[str]) -> str:
    layer = [_to_text(x).strip().lower() for x in hashes if _to_text(x).strip()]
    if not layer:
        return ""
    while len(layer) > 1:
        nxt: list[str] = []
        for i in range(0, len(layer), 2):
            left = layer[i]
            right = layer[i + 1] if i + 1 < len(layer) else left
            nxt.append(hashlib.sha256(f"{left}|{right}".encode("utf-8")).hexdigest())
        layer = nxt
    return layer[0]


def _build_hierarchy_nodes(
    *,
    boq_items: list[BoqItem],
    project_uri: str,
    boq_root_uri: str,
    norm_context_root_uri: str,
    source_file: str,
    bridge_mappings: dict[str, Any] | None = None,
    sb: Any = None,
    persist_registry: bool = False,
) -> list[BoqHierarchyNode]:
    if not boq_items:
        return []
    normalized_bridge_mappings = _normalize_bridge_mapping_keys(bridge_mappings)

    explicit_by_code: dict[str, BoqItem] = {}
    for item in boq_items:
        code = _to_text(item.item_no).strip()
        if not code:
            continue
        current = explicit_by_code.get(code)
        if current is None:
            explicit_by_code[code] = item
            continue
        # Prefer row with richer metadata.
        current_score = int(current.design_quantity is not None) + len(_to_text(current.name).strip())
        next_score = int(item.design_quantity is not None) + len(_to_text(item.name).strip())
        if next_score > current_score:
            explicit_by_code[code] = item

    all_codes: set[str] = set()
    for code in explicit_by_code:
        parts = _item_code_parts(code)
        for i in range(1, len(parts) + 1):
            all_codes.add("-".join(parts[:i]))

    children_map: dict[str, set[str]] = {code: set() for code in all_codes}
    for code in all_codes:
        parent = _direct_parent_code(code)
        if parent and parent in children_map:
            children_map[parent].add(code)

    representative_leaf: dict[str, BoqItem] = {}
    for code in sorted(all_codes, key=_item_code_sort_key):
        parts = _item_code_parts(code)
        for item in boq_items:
            item_code = _to_text(item.item_no).strip()
            item_parts = _item_code_parts(item_code)
            if len(item_parts) < len(parts):
                continue
            if item_parts[: len(parts)] == parts:
                representative_leaf[code] = item
                break

    ref_item_by_code: dict[str, BoqItem | None] = {
        code: (explicit_by_code.get(code) or representative_leaf.get(code))
        for code in all_codes
    }

    def _hierarchy_hint_for_code(code: str) -> dict[str, Any]:
        ref_item = ref_item_by_code.get(code)
        parts = _item_code_parts(code)
        chapter_hint = _extract_code_hint(
            (ref_item.division if ref_item else ""),
            (ref_item.hierarchy_raw if ref_item else ""),
            (ref_item.subdivision if ref_item else ""),
            (ref_item.sheet_name if ref_item else ""),
        )
        section_hint = _extract_code_hint(
            (ref_item.subdivision if ref_item else ""),
            (ref_item.hierarchy_raw if ref_item else ""),
        )
        chapter_code = chapter_hint or (parts[0] if parts else "")
        section_code = section_hint or (parts[0] if parts else chapter_code)
        subgroup_code = "-".join(parts[:3]) if len(parts) >= 3 else (section_code or chapter_code or code)
        return {
            "chapter_code": chapter_code,
            "section_code": section_code,
            "subgroup_code": subgroup_code,
            "code_parts": parts,
            "level": len(parts),
        }

    max_depth_by_code: dict[str, int] = {}
    for code in all_codes:
        max_depth = len(_item_code_parts(code))
        for item in boq_items:
            item_code = _to_text(item.item_no).strip()
            if item_code == code or not item_code.startswith(f"{code}-"):
                continue
            max_depth = max(max_depth, len(_item_code_parts(item_code)))
        max_depth_by_code[code] = max_depth

    node_payload: dict[str, dict[str, Any]] = {}
    for code in sorted(all_codes, key=_item_code_sort_key):
        explicit = explicit_by_code.get(code)
        leaf_ref = representative_leaf.get(code)
        depth = len(_item_code_parts(code))
        max_depth = int(max_depth_by_code.get(code) or depth)
        child_codes = sorted(children_map.get(code) or [], key=_item_code_sort_key)
        is_leaf = len(child_codes) == 0
        node_type = _node_type_for_depth(depth, max_depth)
        hierarchy_meta = _hierarchy_hint_for_code(code)
        uri = build_boq_item_uri(boq_root_uri=boq_root_uri, item_no=code, hierarchy=hierarchy_meta)
        norm_uri = build_norm_context_uri(norm_context_root_uri=norm_context_root_uri, item_no=code)
        parent_code = _direct_parent_code(code)
        if parent_code:
            parent_meta = _hierarchy_hint_for_code(parent_code)
            parent_uri = build_boq_item_uri(boq_root_uri=boq_root_uri, item_no=parent_code, hierarchy=parent_meta)
        else:
            parent_uri = ""
        name = _pick_node_name(code=code, node_type=node_type, explicit=explicit, leaf_ref=leaf_ref)
        ref_item = explicit or leaf_ref
        gate_binding = resolve_linked_gates(
            item_code=code,
            fallback_spec_uri="",
            sb=sb,
            persist_registry=persist_registry,
        )
        spu_ref_pack = (
            resolve_spu_ref_pack(
                item_code=code,
                item_name=name,
                quantity_unit=_to_text((explicit.unit if explicit else (leaf_ref.unit if leaf_ref else "")) or "").strip(),
            )
            if is_leaf
            else {"ref_spu_uri": "", "ref_quota_uri": "", "ref_meter_rule_uri": ""}
        )
        if persist_registry and is_leaf:
            _register_specir_spu_refs(
                sb=sb,
                item_code=code,
                item_name=name,
                ref_pack=spu_ref_pack,
            )

        hierarchy = {
            "chapter_code": hierarchy_meta.get("chapter_code") or "",
            "section_code": hierarchy_meta.get("section_code") or "",
            "subgroup_code": hierarchy_meta.get("subgroup_code") or "",
            "item_code": code,
            "code_parts": list(hierarchy_meta.get("code_parts") or []),
            "level": int(hierarchy_meta.get("level") or depth),
            "node_type": node_type,
            "is_leaf": is_leaf,
            "wbs_path": name,
        }
        bridge_uri = ""
        if is_leaf and normalized_bridge_mappings:
            candidates = [code, code.lower(), uri, uri.rstrip("/"), uri.lower(), uri.rstrip("/").lower()]
            for candidate in candidates:
                if candidate not in normalized_bridge_mappings:
                    continue
                resolved = _normalize_bridge_uri_value(project_uri=project_uri, raw_value=normalized_bridge_mappings[candidate])
                if resolved:
                    bridge_uri = resolved
                    break

        node_payload[code] = {
            "code": code,
            "uri": uri,
            "norm_context_uri": norm_uri,
            "parent_code": parent_code,
            "parent_uri": parent_uri,
            "depth": depth,
            "node_type": node_type,
            "is_leaf": is_leaf,
            "name": name,
            "unit": _to_text((ref_item.unit if ref_item else "") or "").strip(),
            "design_quantity": (explicit.design_quantity if (explicit and is_leaf) else None),
            "approved_quantity": (explicit.approved_quantity if (explicit and is_leaf) else None),
            "unit_price": (explicit.unit_price if (explicit and is_leaf) else None),
            "division": _to_text((ref_item.division if ref_item else "") or "").strip(),
            "subdivision": _to_text((ref_item.subdivision if ref_item else "") or "").strip(),
            "hierarchy_raw": _to_text((ref_item.hierarchy_raw if ref_item else "") or "").strip(),
            "source_file": source_file,
            "source_sheet": _to_text((ref_item.sheet_name if ref_item else "") or "").strip(),
            "source_row": int(ref_item.row_index) if ref_item else None,
            "children_codes": child_codes,
            "children_uris": [
                build_boq_item_uri(
                    boq_root_uri=boq_root_uri,
                    item_no=x,
                    hierarchy=_hierarchy_hint_for_code(x),
                )
                for x in child_codes
            ],
            "linked_gate_id": _to_text(gate_binding.get("linked_gate_id") or "").strip(),
            "linked_gate_ids": list(gate_binding.get("linked_gate_ids") or []),
            "linked_gate_rules": list(gate_binding.get("linked_gate_rules") or []),
            "linked_spec_uri": _to_text(gate_binding.get("linked_spec_uri") or "").strip(),
            "spec_dict_key": _to_text(gate_binding.get("spec_dict_key") or "").strip(),
            "spec_item": _to_text(gate_binding.get("spec_item") or "").strip(),
            "ref_gate_uri": _to_text(gate_binding.get("ref_gate_uri") or "").strip(),
            "ref_gate_uris": list(gate_binding.get("ref_gate_uris") or []),
            "ref_spec_uri": _to_text(gate_binding.get("ref_spec_uri") or "").strip(),
            "ref_spec_dict_uri": _to_text(gate_binding.get("ref_spec_dict_uri") or "").strip(),
            "ref_spec_item_uri": _to_text(gate_binding.get("ref_spec_item_uri") or "").strip(),
            "ref_spu_uri": _to_text(spu_ref_pack.get("ref_spu_uri") or "").strip(),
            "ref_quota_uri": _to_text(spu_ref_pack.get("ref_quota_uri") or "").strip(),
            "ref_meter_rule_uri": _to_text(spu_ref_pack.get("ref_meter_rule_uri") or "").strip(),
            "gate_template_lock": bool(gate_binding.get("gate_template_lock")),
            "gate_binding_hash": _to_text(gate_binding.get("gate_binding_hash") or "").strip(),
            "hierarchy": hierarchy,
            "bridge_uri": bridge_uri,
        }

    # Bottom-up subtree hashing so parent can seal descendants.
    subtree_hash_by_code: dict[str, str] = {}
    children_merkle_by_code: dict[str, str] = {}
    node_hash_by_code: dict[str, str] = {}

    for code in sorted(all_codes, key=lambda x: _item_code_sort_key(x), reverse=True):
        payload = node_payload[code]
        child_codes = payload["children_codes"]
        child_subtrees = [subtree_hash_by_code.get(c, "") for c in child_codes if subtree_hash_by_code.get(c, "")]
        children_merkle = _merkle_root(child_subtrees)
        children_merkle_by_code[code] = children_merkle
        node_canonical = {
            "code": payload["code"],
            "uri": payload["uri"],
            "parent_code": payload["parent_code"],
            "node_type": payload["node_type"],
            "is_leaf": payload["is_leaf"],
            "name": payload["name"],
            "unit": payload["unit"],
            "design_quantity": payload["design_quantity"],
            "approved_quantity": payload["approved_quantity"],
            "unit_price": payload["unit_price"],
            "children_codes": payload["children_codes"],
            "children_merkle_root": children_merkle,
            "linked_gate_id": payload.get("linked_gate_id"),
            "linked_gate_ids": payload.get("linked_gate_ids"),
            "linked_spec_uri": payload.get("linked_spec_uri"),
            "ref_gate_uri": payload.get("ref_gate_uri"),
            "ref_gate_uris": payload.get("ref_gate_uris"),
            "ref_spec_uri": payload.get("ref_spec_uri"),
            "ref_spec_dict_uri": payload.get("ref_spec_dict_uri"),
            "ref_spec_item_uri": payload.get("ref_spec_item_uri"),
            "ref_spu_uri": payload.get("ref_spu_uri"),
            "ref_quota_uri": payload.get("ref_quota_uri"),
            "ref_meter_rule_uri": payload.get("ref_meter_rule_uri"),
            "gate_template_lock": payload.get("gate_template_lock"),
            "hierarchy": payload["hierarchy"],
            "bridge_uri": payload.get("bridge_uri"),
            "source_file": payload["source_file"],
            "source_sheet": payload["source_sheet"],
            "source_row": payload["source_row"],
        }
        node_hash = hashlib.sha256(
            json.dumps(node_canonical, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
        ).hexdigest()
        node_hash_by_code[code] = node_hash
        subtree_hash_by_code[code] = hashlib.sha256(f"{node_hash}|{children_merkle}".encode("utf-8")).hexdigest()

    nodes: list[BoqHierarchyNode] = []
    for code in sorted(all_codes, key=_item_code_sort_key):
        payload = node_payload[code]
        node = BoqHierarchyNode(
            code=payload["code"],
            uri=payload["uri"],
            norm_context_uri=payload["norm_context_uri"],
            parent_code=payload["parent_code"],
            parent_uri=payload["parent_uri"],
            depth=int(payload["depth"]),
            node_type=_to_text(payload["node_type"]).strip(),
            is_leaf=bool(payload["is_leaf"]),
            name=_to_text(payload["name"]).strip(),
            unit=_to_text(payload["unit"]).strip(),
            design_quantity=payload["design_quantity"],
            approved_quantity=payload["approved_quantity"],
            unit_price=payload["unit_price"],
            division=_to_text(payload["division"]).strip(),
            subdivision=_to_text(payload["subdivision"]).strip(),
            hierarchy_raw=_to_text(payload["hierarchy_raw"]).strip(),
            source_file=_to_text(payload["source_file"]).strip(),
            source_sheet=_to_text(payload["source_sheet"]).strip(),
            source_row=payload["source_row"],
            children_codes=list(payload["children_codes"]),
            children_uris=list(payload["children_uris"]),
            children_merkle_root=children_merkle_by_code.get(code, ""),
            node_hash=node_hash_by_code.get(code, ""),
            subtree_hash=subtree_hash_by_code.get(code, ""),
            linked_gate_id=_to_text(payload.get("linked_gate_id") or "").strip(),
            linked_gate_ids=list(payload.get("linked_gate_ids") or []),
            linked_gate_rules=list(payload.get("linked_gate_rules") or []),
            linked_spec_uri=_to_text(payload.get("linked_spec_uri") or "").strip(),
            spec_dict_key=_to_text(payload.get("spec_dict_key") or "").strip(),
            spec_item=_to_text(payload.get("spec_item") or "").strip(),
            ref_gate_uri=_to_text(payload.get("ref_gate_uri") or "").strip(),
            ref_gate_uris=list(payload.get("ref_gate_uris") or []),
            ref_spec_uri=_to_text(payload.get("ref_spec_uri") or "").strip(),
            ref_spec_dict_uri=_to_text(payload.get("ref_spec_dict_uri") or "").strip(),
            ref_spec_item_uri=_to_text(payload.get("ref_spec_item_uri") or "").strip(),
            ref_spu_uri=_to_text(payload.get("ref_spu_uri") or "").strip(),
            ref_quota_uri=_to_text(payload.get("ref_quota_uri") or "").strip(),
            ref_meter_rule_uri=_to_text(payload.get("ref_meter_rule_uri") or "").strip(),
            gate_template_lock=bool(payload.get("gate_template_lock")),
            hierarchy=dict(payload["hierarchy"]),
            bridge_uri=_to_text(payload.get("bridge_uri") or "").strip(),
            genesis_hash=subtree_hash_by_code.get(code, ""),
        )
        nodes.append(node)
    return nodes


def parse_boq_hierarchy(
    excel_file: str | Path,
    *,
    sheet_name: str | None = None,
    boq_root_uri: str = "v://project/boq",
    norm_context_root_uri: str = "v://project/normContext",
) -> dict[str, Any]:
    source_path = Path(excel_file).expanduser().resolve()
    items = parse_boq_excel(source_path, sheet_name=sheet_name, leaf_only=False)
    nodes = _build_hierarchy_nodes(
        boq_items=items,
        project_uri="",
        boq_root_uri=boq_root_uri,
        norm_context_root_uri=norm_context_root_uri,
        source_file=str(source_path),
        persist_registry=False,
    )
    children_by_parent: dict[str, list[str]] = {}
    for node in nodes:
        children_by_parent[node.code] = list(node.children_codes)

    root_codes = [n.code for n in nodes if not n.parent_code]
    return {
        "source_file": str(source_path),
        "item_count": len(items),
        "node_count": len(nodes),
        "root_codes": root_codes,
        "nodes": [asdict(n) for n in nodes],
        "children": children_by_parent,
    }


def build_boq_item_uri(
    *,
    boq_root_uri: str,
    item_no: str,
    hierarchy: dict[str, Any] | None = None,
) -> str:
    root = _to_text(boq_root_uri).strip().rstrip("/")
    code = _to_text(item_no).strip()
    if not root:
        root = "v://project/boq"
    if not code:
        return root
    hierarchy = hierarchy or {}

    chapter_code = _to_text(hierarchy.get("chapter_code") or "").strip()
    section_code = _to_text(hierarchy.get("section_code") or "").strip()
    root_tail = root.split("/")[-1] if root else ""
    code_parts = list(hierarchy.get("code_parts") or _item_code_parts(code))

    segments: list[str] = []
    if chapter_code and root_tail != chapter_code:
        segments.append(chapter_code)
    if section_code and section_code not in {chapter_code, code} and root_tail != section_code:
        if not segments or segments[-1] != section_code:
            segments.append(section_code)

    chain: list[str] = []
    if code_parts:
        for i in range(1, len(code_parts) + 1):
            chain.append("-".join(code_parts[:i]))
    else:
        chain = [code]

    for seg in chain:
        if not seg or seg == root_tail:
            continue
        if seg == chapter_code or seg == section_code:
            continue
        if segments and segments[-1] == seg:
            continue
        segments.append(seg)

    if not segments:
        return root
    return f"{root}/{'/'.join(segments)}"


def build_norm_context_uri(*, norm_context_root_uri: str, item_no: str) -> str:
    root = _to_text(norm_context_root_uri).strip().rstrip("/")
    if not root:
        root = "v://project/normContext"
    return f"{root}/{_to_text(item_no).strip()}"


def compute_genesis_hash(item: BoqItem, *, boq_item_uri: str, source_file: str) -> str:
    contract_quantity = item.approved_quantity if item.approved_quantity is not None else item.design_quantity
    contract_quantity = _round4(contract_quantity)
    unit_price = _round4(item.unit_price)
    seed = f"{_to_text(item.item_no).strip()}|{contract_quantity or 0:.4f}|{unit_price or 0:.4f}"
    return hashlib.sha256(seed.encode("utf-8")).hexdigest()


def _proof_id_for_boq_item(project_uri: str, boq_item_uri: str, genesis_hash: str) -> str:
    seed = hashlib.sha256(f"{project_uri}|{boq_item_uri}|{genesis_hash}".encode("utf-8")).hexdigest()[:18].upper()
    return f"GP-BOQ-{seed}"


def _default_owner_uri(project_uri: str) -> str:
    base = _to_text(project_uri).strip().rstrip("/")
    if not base:
        return "v://project/executor/system/"
    return f"{base}/executor/system/"


def _slug_token(value: Any) -> str:
    text = re.sub(r"\s+", "-", _to_text(value).strip().lower())
    text = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff_-]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "bridge"


def _normalize_bridge_uri_value(*, project_uri: str, raw_value: Any) -> str:
    if isinstance(raw_value, dict):
        raw_uri = _to_text(raw_value.get("bridge_uri") or "").strip()
        if raw_uri.startswith("v://"):
            return raw_uri.rstrip("/")
        raw_name = _to_text(raw_value.get("bridge_name") or "").strip()
        if raw_name:
            return f"{project_uri.rstrip('/')}/bridge/{_slug_token(raw_name)}"
        return ""
    text = _to_text(raw_value).strip()
    if not text:
        return ""
    if text.startswith("v://"):
        return text.rstrip("/")
    return f"{project_uri.rstrip('/')}/bridge/{_slug_token(text)}"


def _normalize_bridge_mapping_keys(
    bridge_mappings: dict[str, Any] | None,
) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for raw_key, value in (bridge_mappings or {}).items():
        key = _to_text(raw_key).strip()
        if not key:
            continue
        normalized[key] = value
        normalized[key.rstrip("/")] = value
        normalized[key.lower()] = value
        if "/" in key:
            tail = key.rstrip("/").split("/")[-1]
            if tail:
                normalized[tail] = value
                normalized[tail.lower()] = value
    return normalized


def _write_zero_ledger_meta(
    *,
    sb: Any,
    project_uri: str,
    boq_item_uri: str,
    contract_price: float | None,
    approved_qty: float | None,
    genesis_hash: str,
    owner_uri: str,
    created_at: str,
) -> None:
    try:
        payload = {
            "project_uri": project_uri,
            "boq_item_uri": boq_item_uri,
            "contract_price": contract_price,
            "approved_qty": approved_qty,
            "genesis_hash": genesis_hash,
            "owner_uri": owner_uri,
            "created_at": created_at,
            "read_only_fields": ["contract_price", "approved_qty"],
        }
        sb.table("project_zero_ledger_meta").insert(payload).execute()
    except Exception:
        # Best-effort: ignore if table does not exist or permissions are missing.
        return


def _boq_full_line_uri(*, project_uri: str, item_no: str) -> str:
    base = _to_text(project_uri).strip().rstrip("/")
    code = _to_text(item_no).strip()
    if not base or not code:
        return ""
    return f"{base}/full-line/boq/{code}"


def _boq_bridge_scoped_uri(*, bridge_uri: str, item_no: str) -> str:
    bridge = _to_text(bridge_uri).strip().rstrip("/")
    code = _to_text(item_no).strip()
    if not bridge.startswith("v://") or not code:
        return ""
    return f"{bridge}/boq/{code}"


def _register_boq_item_sovereignty(
    *,
    sb: Any,
    project_uri: str,
    item_no: str,
    description: str,
    unit: str,
    boq_quantity: float | None,
    unit_price: float | None,
    contract_total: float | None,
    bridge_uri: str,
    source_file: str,
    source_sheet: str,
    source_row: int | None,
    commit: bool,
    fallback_boq_uri: str,
) -> dict[str, Any]:
    project = _to_text(project_uri).strip().rstrip("/")
    item_code = _to_text(item_no).strip()
    bridge = _to_text(bridge_uri).strip().rstrip("/")
    fallback_uri = _to_text(fallback_boq_uri).strip()
    metadata = {
        "description": _to_text(description).strip(),
        "unit": _normalize_unit(unit),
        "boq_quantity": boq_quantity,
        "unit_price": unit_price,
        "total_amount": contract_total,
        "chapter": _item_code_parts(item_code)[0] if _item_code_parts(item_code) else "",
        "bridge_uri": bridge,
        "bridge_name": bridge.split("/")[-1] if bridge.startswith("v://") else "",
        "source_file": _to_text(source_file).strip(),
        "source_sheet": _to_text(source_sheet).strip(),
        "source_row": source_row,
    }
    fallback_canonical = f"{project}/boq/{item_code}" if project and item_code else fallback_uri
    fallback_full_line = _boq_full_line_uri(project_uri=project, item_no=item_code)
    fallback_bridge_scoped = _boq_bridge_scoped_uri(bridge_uri=bridge, item_no=item_code)
    try:
        registration = register_boq_item(
            sb=sb,
            project_uri=project,
            identifier=item_code,
            metadata=metadata,
            bridge_uri=bridge,
            commit=bool(commit),
        )
        canonical_uri = _to_text(registration.get("canonical_uri") or fallback_canonical or fallback_uri).strip()
        full_line_uri = _to_text(registration.get("full_line_uri") or fallback_full_line).strip()
        bridge_scoped_uri = _to_text(registration.get("bridge_scoped_uri") or fallback_bridge_scoped).strip()
        aliases: list[str] = []
        for candidate in (full_line_uri, bridge_scoped_uri, fallback_uri):
            text = _to_text(candidate).strip()
            if text and text not in aliases:
                aliases.append(text)
        return {
            "ok": bool(registration.get("ok", True)),
            "canonical_uri": canonical_uri,
            "full_line_uri": full_line_uri,
            "bridge_scoped_uri": bridge_scoped_uri,
            "alias_uris": aliases,
            "metadata": metadata,
            "committed": bool(registration.get("committed")),
            "registration": registration.get("registration") if isinstance(registration.get("registration"), dict) else {},
        }
    except Exception as exc:
        aliases = []
        for candidate in (fallback_full_line, fallback_bridge_scoped, fallback_uri):
            text = _to_text(candidate).strip()
            if text and text not in aliases:
                aliases.append(text)
        return {
            "ok": False,
            "canonical_uri": fallback_canonical or fallback_uri,
            "full_line_uri": fallback_full_line,
            "bridge_scoped_uri": fallback_bridge_scoped,
            "alias_uris": aliases,
            "metadata": metadata,
            "committed": False,
            "registration_error": f"{exc.__class__.__name__}: {exc}",
        }


def initialize_boq_utxos(
    *,
    sb: Any,
    project_uri: str,
    project_id: str | None,
    boq_items: list[BoqItem],
    boq_root_uri: str = "v://project/boq",
    norm_context_root_uri: str = "v://project/normContext",
    owner_uri: str | None = None,
    source_file: str = "",
    bridge_mappings: dict[str, Any] | None = None,
    commit: bool = False,
) -> dict[str, Any]:
    if not _to_text(project_uri).strip():
        raise ValueError("project_uri is required")

    engine = ProofUTXOEngine(sb) if commit else None
    effective_owner = _to_text(owner_uri).strip() or _default_owner_uri(project_uri)

    hierarchy_nodes = _build_hierarchy_nodes(
        boq_items=boq_items,
        project_uri=project_uri,
        boq_root_uri=boq_root_uri,
        norm_context_root_uri=norm_context_root_uri,
        source_file=source_file,
        bridge_mappings=bridge_mappings,
        sb=sb,
        persist_registry=bool(commit),
    )
    if not hierarchy_nodes:
        return {
            "commit": bool(commit),
            "project_uri": project_uri,
            "project_id": project_id,
            "owner_uri": effective_owner,
            "source_file": source_file,
            "boq_root_uri": boq_root_uri,
            "norm_context_root_uri": norm_context_root_uri,
            "total_items": len(boq_items),
            "total_nodes": 0,
            "leaf_nodes": 0,
            "group_nodes": 0,
            "preview": [],
            "created": [],
            "errors": [],
            "success_count": 0,
        }

    proof_id_by_uri: dict[str, str] = {}
    for node in hierarchy_nodes:
        proof_id_by_uri[node.uri] = _proof_id_for_boq_item(project_uri, node.uri, node.genesis_hash)

    created: list[dict[str, Any]] = []
    preview: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for node in hierarchy_nodes:
        boq_item_uri = node.uri
        norm_context_uri = node.norm_context_uri
        hierarchy = dict(node.hierarchy)
        genesis_hash = node.genesis_hash
        proof_id = proof_id_by_uri.get(boq_item_uri) or _proof_id_for_boq_item(project_uri, boq_item_uri, genesis_hash)
        genesis_at = datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S")
        parent_utxo = proof_id_by_uri.get(node.parent_uri, "") if node.parent_uri else ""
        children_utxo = [proof_id_by_uri.get(uri, "") for uri in node.children_uris]
        children_utxo = [x for x in children_utxo if x]
        contract_quantity = _round4(node.approved_quantity if node.approved_quantity is not None else node.design_quantity)
        contract_unit_price = _round4(node.unit_price) if node.is_leaf else None
        design_total = None
        contract_total = None
        if node.is_leaf and node.design_quantity is not None and node.unit_price is not None:
            design_total = _round4(float(node.design_quantity) * float(node.unit_price))
        if node.is_leaf and contract_quantity is not None and node.unit_price is not None:
            contract_total = _round4(float(contract_quantity) * float(node.unit_price))
        boq_registration: dict[str, Any] = {}
        boq_item_canonical_uri = boq_item_uri
        boq_item_full_line_uri = ""
        boq_item_bridge_scoped_uri = ""
        boq_item_uri_aliases: list[str] = []
        if node.is_leaf:
            boq_registration = _register_boq_item_sovereignty(
                sb=sb,
                project_uri=project_uri,
                item_no=node.code,
                description=node.name,
                unit=node.unit,
                boq_quantity=contract_quantity,
                unit_price=contract_unit_price,
                contract_total=contract_total,
                bridge_uri=_to_text(node.bridge_uri).strip(),
                source_file=source_file,
                source_sheet=node.source_sheet,
                source_row=node.source_row,
                commit=bool(commit),
                fallback_boq_uri=boq_item_uri,
            )
            boq_item_canonical_uri = _to_text(boq_registration.get("canonical_uri") or boq_item_uri).strip()
            boq_item_full_line_uri = _to_text(boq_registration.get("full_line_uri")).strip()
            boq_item_bridge_scoped_uri = _to_text(boq_registration.get("bridge_scoped_uri")).strip()
            boq_item_uri_aliases = [
                _to_text(item).strip()
                for item in (boq_registration.get("alias_uris") or [])
                if _to_text(item).strip()
            ]
        state_data = {
            "asset_type": "boq_item" if node.is_leaf else "boq_group",
            "utxo_kind": "BOQ_INITIAL",
            "utxo_state": "UNSPENT",
            "status": "INITIAL",
            "status_code": 0,
            "status_label": "GENESIS",
            "lifecycle_stage": "INITIAL",
            "is_leaf": bool(node.is_leaf),
            "node_type": node.node_type,
            "node_code": node.code,
            "boq_item_uri": boq_item_uri,
            "boq_item_v_uri": boq_item_canonical_uri if node.is_leaf else boq_item_uri,
            "boq_item_canonical_uri": boq_item_canonical_uri if node.is_leaf else boq_item_uri,
            "boq_item_full_line_uri": boq_item_full_line_uri if node.is_leaf else "",
            "boq_item_bridge_scoped_uri": boq_item_bridge_scoped_uri if node.is_leaf else "",
            "boq_item_uri_aliases": boq_item_uri_aliases if node.is_leaf else [],
            "boq_item_registration": boq_registration if node.is_leaf else {},
            "bridge_uri": _to_text(node.bridge_uri).strip(),
            "norm_context_uri": norm_context_uri,
            "linked_gate_id": _to_text(node.linked_gate_id).strip(),
            "linked_gate_ids": list(node.linked_gate_ids),
            "linked_spec_uri": _to_text(node.linked_spec_uri).strip(),
            "spec_dict_key": _to_text(node.spec_dict_key).strip(),
            "spec_item": _to_text(node.spec_item).strip(),
            "ref_gate_uri": _to_text(node.ref_gate_uri).strip(),
            "ref_gate_uris": list(node.ref_gate_uris),
            "ref_spec_uri": _to_text(node.ref_spec_uri).strip(),
            "ref_spec_dict_uri": _to_text(node.ref_spec_dict_uri).strip(),
            "ref_spec_item_uri": _to_text(node.ref_spec_item_uri).strip(),
            "ref_spu_uri": _to_text(node.ref_spu_uri).strip(),
            "ref_quota_uri": _to_text(node.ref_quota_uri).strip(),
            "ref_meter_rule_uri": _to_text(node.ref_meter_rule_uri).strip(),
            "gate_template_lock": bool(node.gate_template_lock),
            "norm_refs": _default_norm_refs(item_no=node.code, item_name=node.name) if node.is_leaf else [],
            "item_no": node.code,
            "item_name": node.name,
            "unit": _normalize_unit(node.unit),
            "division": node.division,
            "subdivision": node.subdivision,
            "hierarchy_raw": node.hierarchy_raw,
            "hierarchy": hierarchy,
            "design_quantity": _round4(node.design_quantity) if node.is_leaf else None,
            "design_quantity_raw": "",
            "unit_price": _round4(node.unit_price) if node.is_leaf else None,
            "unit_price_raw": "",
            "approved_quantity": _round4(node.approved_quantity) if node.is_leaf else None,
            "approved_quantity_raw": "",
            "contract_quantity": contract_quantity if node.is_leaf else None,
            "contract_unit_price": contract_unit_price if node.is_leaf else None,
            "contract_price": contract_unit_price if node.is_leaf else None,
            "approved_qty": _round4(node.approved_quantity) if node.is_leaf else None,
            "design_total": design_total if node.is_leaf else None,
            "contract_total": contract_total if node.is_leaf else None,
            "read_only_fields": ["approved_quantity", "approved_qty", "contract_unit_price", "contract_price"],
            "remark": "",
            "hierarchy_tree": {
                "depth": int(node.depth),
                "node_type": node.node_type,
                "is_leaf": bool(node.is_leaf),
                "parent_uri": node.parent_uri,
                "parent_code": node.parent_code,
                "parent_utxo": parent_utxo,
                "children": list(node.children_uris),
                "children_codes": list(node.children_codes),
                "children_utxo": children_utxo,
                "children_merkle_root": node.children_merkle_root,
                "node_hash": node.node_hash,
                "subtree_hash": node.subtree_hash,
            },
            "genesis_hash": genesis_hash,
            "genesis_at": genesis_at,
            "genesis_proof": {
                "proof_id": proof_id,
                "algorithm": "sha256",
                "hash": genesis_hash,
                "created_at": genesis_at,
                "initial_quantity": node.design_quantity if node.is_leaf else None,
                "unit_price": node.unit_price if node.is_leaf else None,
                "contract_quantity": contract_quantity if node.is_leaf else None,
                "contract_unit_price": contract_unit_price if node.is_leaf else None,
                "design_total": design_total if node.is_leaf else None,
                "contract_total": contract_total if node.is_leaf else None,
                "wbs_path": hierarchy.get("wbs_path"),
                "children_merkle_root": node.children_merkle_root,
                "node_hash": node.node_hash,
                "subtree_hash": node.subtree_hash,
            },
            "source": {
                "file": source_file,
                "sheet": node.source_sheet,
                "row": node.source_row,
            },
            "gate_binding_hash": hashlib.sha256(
                json.dumps(
                    {
                        "item_code": node.code,
                        "linked_gate_id": _to_text(node.linked_gate_id).strip(),
                        "linked_gate_ids": list(node.linked_gate_ids),
                        "linked_spec_uri": _to_text(node.linked_spec_uri).strip(),
                        "spec_dict_key": _to_text(node.spec_dict_key).strip(),
                        "spec_item": _to_text(node.spec_item).strip(),
                        "ref_gate_uri": _to_text(node.ref_gate_uri).strip(),
                        "ref_spec_uri": _to_text(node.ref_spec_uri).strip(),
                        "ref_spec_dict_uri": _to_text(node.ref_spec_dict_uri).strip(),
                        "ref_spec_item_uri": _to_text(node.ref_spec_item_uri).strip(),
                        "ref_spu_uri": _to_text(node.ref_spu_uri).strip(),
                        "ref_quota_uri": _to_text(node.ref_quota_uri).strip(),
                        "ref_meter_rule_uri": _to_text(node.ref_meter_rule_uri).strip(),
                    },
                    ensure_ascii=False,
                    sort_keys=True,
                    default=str,
                ).encode("utf-8")
            ).hexdigest(),
        }
        if node.is_leaf:
            project_ref_payload = build_project_boq_item_ref(
                boq_v_uri=boq_item_canonical_uri,
                boq_item_id=node.code,
                description=node.name,
                quantity=contract_quantity,
                unit=_normalize_unit(node.unit),
                bridge_uri=_to_text(node.bridge_uri).strip(),
                ref_spu_uri=_to_text(node.ref_spu_uri).strip(),
                ref_quota_uri=_to_text(node.ref_quota_uri).strip(),
                ref_meter_rule_uri=_to_text(node.ref_meter_rule_uri).strip(),
            )
            standard_spu_snapshot = resolve_standard_spu_snapshot(
                sb=sb if bool(commit) else None,
                ref_spu_uri=_to_text(node.ref_spu_uri).strip(),
            )
            state_data["utxo_quantity"] = contract_quantity
            state_data["utxo_parent"] = parent_utxo or None
            state_data["project_boq_item_ref"] = project_ref_payload
            state_data["standard_spu_snapshot"] = standard_spu_snapshot
            state_data["repo_ir_layer"] = "project_ref_only"
            state_data["boq_item"] = {
                "boq_item_id": node.code,
                "description": node.name,
                "unit": _normalize_unit(node.unit),
                "boq_quantity": contract_quantity,
                "project_uri": project_uri,
                "bridge_uri": _to_text(node.bridge_uri).strip(),
                "spec_uri": _to_text(node.linked_spec_uri).strip(),
                "v_uri": boq_item_canonical_uri,
                "canonical_uri": boq_item_canonical_uri,
                "full_line_uri": boq_item_full_line_uri,
                "bridge_scoped_uri": boq_item_bridge_scoped_uri,
                "uri_aliases": boq_item_uri_aliases,
            }
            state_data["boq_utxo"] = {
                "utxo_id": proof_id,
                "boq_item_id": node.code,
                "kind": "BOQ_INITIAL",
                "state": "UNSPENT",
                "quantity": contract_quantity,
                "parent_utxo": parent_utxo or None,
                "bridge_uri": _to_text(node.bridge_uri).strip(),
                "boq_v_uri": boq_item_canonical_uri,
                "proof_hash": "",
                "timestamp": genesis_at,
            }
        if node.is_leaf and _to_text(node.linked_spec_uri).strip():
            state_data["spec_uri"] = _to_text(node.linked_spec_uri).strip()
        if node.is_leaf:
            state_data["ledger"] = {
                "initial_balance": contract_quantity if contract_quantity is not None else node.design_quantity,
                "balance_unit": node.unit,
                "balance_source": "contract_quantity" if contract_quantity is not None else "design_quantity",
            }
        else:
            state_data["aggregate"] = {
                "children_count": len(node.children_uris),
                "leaf_metric_mode": "recursive_summary_only",
            }

        payload_preview = {
            "proof_id": proof_id,
            "owner_uri": effective_owner,
            "project_id": project_id,
            "project_uri": project_uri,
            "proof_type": "zero_ledger",
            "result": "PENDING",
            "norm_uri": norm_context_uri,
            "state_data": state_data,
        }

        preview.append(payload_preview)

        if not commit:
            continue

        try:
            assert engine is not None
            row = engine.create(
                proof_id=proof_id,
                owner_uri=effective_owner,
                project_id=project_id,
                project_uri=project_uri,
                proof_type="zero_ledger",
                result="PENDING",
                state_data=state_data,
                conditions=[],
                parent_proof_id=parent_utxo or None,
                norm_uri=norm_context_uri,
                segment_uri=boq_item_canonical_uri if node.is_leaf else boq_item_uri,
                signer_uri=effective_owner,
                signer_role="SYSTEM",
                gitpeg_anchor=None,
                anchor_config=None,
            )
            created.append(row)
            if node.is_leaf:
                _write_zero_ledger_meta(
                    sb=sb,
                    project_uri=project_uri,
                    boq_item_uri=boq_item_canonical_uri,
                    contract_price=contract_unit_price,
                    approved_qty=contract_quantity,
                    genesis_hash=genesis_hash,
                    owner_uri=effective_owner,
                    created_at=genesis_at,
                )
        except Exception as exc:
            errors.append(
                {
                    "item_no": node.code,
                    "boq_item_uri": boq_item_uri,
                    "proof_id": proof_id,
                    "error": f"{exc.__class__.__name__}: {exc}",
                }
            )

    root_subtrees = [n.subtree_hash for n in hierarchy_nodes if not n.parent_code and n.subtree_hash]
    hierarchy_root_hash = _merkle_root(root_subtrees)
    leaf_nodes = sum(1 for n in hierarchy_nodes if n.is_leaf)
    group_nodes = len(hierarchy_nodes) - leaf_nodes

    return {
        "commit": bool(commit),
        "project_uri": project_uri,
        "project_id": project_id,
        "owner_uri": effective_owner,
        "source_file": source_file,
        "boq_root_uri": boq_root_uri,
        "norm_context_root_uri": norm_context_root_uri,
        "total_items": len(boq_items),
        "total_nodes": len(hierarchy_nodes),
        "leaf_nodes": leaf_nodes,
        "group_nodes": group_nodes,
        "hierarchy_root_hash": hierarchy_root_hash,
        "preview": preview,
        "created": created,
        "errors": errors,
        "success_count": len(created) if commit else 0,
    }


def boq_item_to_dict(item: BoqItem) -> dict[str, Any]:
    return asdict(item)

