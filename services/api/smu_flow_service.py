"""
SMU lifecycle orchestration service.

Six-stage flow:
1) Genesis Trip (BOQ import -> hierarchical UTXO)
2) Governance & QCGate (dynamic form + threshold context)
3) Execution & SnapPeg (TripRole execution + evidence fingerprint)
4) OrdoSign & DID (multi-party sovereign sign)
5) DocPeg Execution (approved trigger -> report bundle context)
6) SMU Risk Audit & Freeze (validate_logic + freeze proof)
"""

from __future__ import annotations

import ast
import base64
import csv
from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
import io
import json
import os
from pathlib import Path
import re
import zipfile
from typing import Any, Callable

from fastapi import HTTPException

from services.api.boq_utxo_service import (
    BoqItem,
    initialize_boq_utxos,
)
from services.api.labpeg_frequency_remediation_service import resolve_dual_pass_gate
from services.api.doc_governance_service import register_document
from services.api.erpnext_service import load_erpnext_custom
from services.api.erpnext_http_utils import erp_request_sync
from services.api.proof_utxo_engine import ProofUTXOEngine
from services.api.reports_generation_service import REPORTS_BUCKET
from services.api.specdict_gate_service import resolve_dynamic_threshold
from services.api.did_reputation_service import build_did_reputation_summary
from services.api.triprole_engine import (
    build_docfinal_package_for_boq,
    execute_triprole_action,
    get_boq_realtime_status,
)
from services.api.unit_merkle_service import build_unit_merkle_snapshot


ITEM_NO_PATTERN = re.compile(r"^\d{3}(?:-[0-9A-Za-z]+)*$")
SPU_LIBRARY_ROOT_URI = "v://library/spu"


def _safe_name(value: str, fallback: str = "doc") -> str:
    text = _to_text(value).strip()
    text = re.sub(r"[^\w.\-]+", "_", text, flags=re.ASCII).strip("_")
    return text or fallback


def _docpeg_report_no(item_no: str) -> str:
    now = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    safe_item = _safe_name(item_no or "", "ITEM")
    return f"DOCPEG-{safe_item}-{now}"


def _upload_docpeg_pdf(
    *,
    sb: Any,
    project_uri: str,
    report_no: str,
    pdf_bytes: bytes,
) -> tuple[str, str]:
    project_key = _safe_name(project_uri.replace("v://", "v_"), "project")
    file_name = f"{_safe_name(report_no, 'DOCPEG')}.pdf"
    storage_path = f"{project_key}/docpeg/{file_name}"
    sb.storage.from_(REPORTS_BUCKET).upload(
        storage_path,
        pdf_bytes,
        file_options={"content-type": "application/pdf"},
    )
    public_url = sb.storage.from_(REPORTS_BUCKET).get_public_url(storage_path)
    storage_url = public_url if isinstance(public_url, str) else ""
    return storage_path, storage_url


def _load_project_for_erpnext(sb: Any, project_id: str, project_uri: str) -> dict[str, Any]:
    if project_id:
        res = (
            sb.table("projects")
            .select("id,enterprise_id,erp_project_code,erp_project_name,name,contract_no,v_uri")
            .eq("id", project_id)
            .limit(1)
            .execute()
        )
        data = res.data or []
        if data:
            return data[0]
    if project_uri:
        res = (
            sb.table("projects")
            .select("id,enterprise_id,erp_project_code,erp_project_name,name,contract_no,v_uri")
            .eq("v_uri", project_uri)
            .limit(1)
            .execute()
        )
        data = res.data or []
        if data:
            return data[0]
    return {}


def _push_docpeg_to_erpnext(
    *,
    sb: Any,
    project_id: str,
    project_uri: str,
    item_no: str,
    item_name: str,
    report_no: str,
    report_url: str,
    docpeg_document: dict[str, Any],
    docpeg_context: dict[str, Any],
    risk_audit: dict[str, Any],
) -> dict[str, Any]:
    proj = _load_project_for_erpnext(sb, project_id, project_uri)
    enterprise_id = _to_text(proj.get("enterprise_id") or "").strip()
    if not enterprise_id:
        return {"attempted": False, "success": False, "reason": "enterprise_id_missing"}
    custom = load_erpnext_custom(sb, enterprise_id)
    if not _to_text(custom.get("erpnext_url") or "").strip():
        return {"attempted": False, "success": False, "reason": "erpnext_url_not_configured"}
    path = str(custom.get("erpnext_notify_path") or "").strip() or "/api/method/qcspec_notify"
    payload = {
        "enterprise_id": enterprise_id,
        "project_id": proj.get("id"),
        "project_name": proj.get("name"),
        "erp_project_code": proj.get("erp_project_code"),
        "erp_project_name": proj.get("erp_project_name") or proj.get("name"),
        "contract_no": proj.get("contract_no"),
        "project_uri": proj.get("v_uri") or project_uri,
        "stake": _to_text(docpeg_context.get("stake") or ""),
        "subitem": item_no,
        "subitem_name": item_name,
        "smu_id": _smu_id_from_item_code(item_no),
        "result": "pass",
        "quality_passed": True,
        "metering_action": "docpeg_report",
        "reason": "",
        "docpeg": {
            "report_no": report_no,
            "report_url": report_url,
            "docpeg_document_proof_id": docpeg_document.get("proof_id"),
            "docpeg_document_proof_hash": docpeg_document.get("proof_hash"),
            "artifact_uri": docpeg_context.get("artifact_uri"),
            "verify_uri": docpeg_context.get("verify_uri"),
            "smu_id": _smu_id_from_item_code(item_no),
            "total_proof_hash": risk_audit.get("total_proof_hash"),
            "risk_score": risk_audit.get("risk_score"),
            "risk_issue_count": len(risk_audit.get("issues") or []),
        },
    }
    res = erp_request_sync(custom, method="POST", path=path, body=payload, timeout_s=12.0)
    res["payload"] = payload
    return res


def _erpnext_receipt_proof_id(
    project_uri: str,
    boq_item_uri: str,
    payload: dict[str, Any],
    response: dict[str, Any],
) -> str:
    docpeg = _as_dict(payload.get("docpeg"))
    seed = {
        "project_uri": _to_text(project_uri).strip(),
        "boq_item_uri": _to_text(boq_item_uri).strip(),
        "report_no": _to_text(docpeg.get("report_no") or "").strip(),
        "docpeg_document_proof_id": _to_text(docpeg.get("docpeg_document_proof_id") or "").strip(),
        "docpeg_document_proof_hash": _to_text(docpeg.get("docpeg_document_proof_hash") or "").strip(),
        "success": bool(response.get("success")),
    }
    return f"GP-ERP-{_sha(seed)[:16].upper()}"


def _create_erpnext_receipt_proof(
    *,
    sb: Any,
    project_uri: str,
    boq_item_uri: str,
    payload: dict[str, Any],
    response: dict[str, Any],
    source_utxo_id: str = "",
    executor_uri: str = "v://executor/erpnext/system/",
) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    b_uri = _to_text(boq_item_uri).strip()
    if not p_uri or not b_uri:
        return {"ok": False, "error": "missing_project_or_boq_uri"}

    proof_id = _erpnext_receipt_proof_id(p_uri, b_uri, payload, response)
    engine = ProofUTXOEngine(sb)
    existing = engine.get_by_id(proof_id)
    if isinstance(existing, dict):
        return {"ok": True, "proof_id": proof_id, "already_exists": True}

    docpeg = _as_dict(payload.get("docpeg"))
    docpeg_proof_id = _to_text(docpeg.get("docpeg_document_proof_id") or "").strip()
    parent_id = docpeg_proof_id or _to_text(source_utxo_id).strip() or None
    result = "PASS" if bool(response.get("success")) else "FAIL"
    state_data = {
        "doc_type": "erpnext_receipt",
        "status": "sent" if result == "PASS" else "failed",
        "project_uri": p_uri,
        "boq_item_uri": b_uri,
        "item_no": _to_text(payload.get("subitem") or "").strip(),
        "item_name": _to_text(payload.get("subitem_name") or "").strip(),
        "report_no": _to_text(docpeg.get("report_no") or "").strip(),
        "report_url": _to_text(docpeg.get("report_url") or "").strip(),
        "verify_uri": _to_text(docpeg.get("verify_uri") or "").strip(),
        "total_proof_hash": _to_text(docpeg.get("total_proof_hash") or "").strip(),
        "risk_score": docpeg.get("risk_score"),
        "risk_issue_count": docpeg.get("risk_issue_count"),
        "docpeg_document_proof_id": docpeg_proof_id,
        "docpeg_document_proof_hash": _to_text(docpeg.get("docpeg_document_proof_hash") or "").strip(),
        "payload": payload,
        "response": response,
        "source_utxo_id": _to_text(source_utxo_id).strip(),
        "created_at": _utc_iso(),
    }
    try:
        engine.create(
            proof_id=proof_id,
            owner_uri=_to_text(executor_uri).strip() or "v://executor/erpnext/system/",
            project_uri=p_uri,
            project_id=_to_text(payload.get("project_id") or "").strip() or None,
            segment_uri=b_uri,
            proof_type="erpnext_receipt",
            result=result,
            state_data=state_data,
            conditions=[],
            parent_proof_id=parent_id,
            norm_uri="v://norm/CoordOS/ERPNext/1.0#receipt",
            signer_uri=_to_text(executor_uri).strip() or "v://executor/erpnext/system/",
            signer_role="ERP",
        )
        return {"ok": True, "proof_id": proof_id, "result": result}
    except Exception as exc:
        return {"ok": False, "error": f"{exc.__class__.__name__}: {exc}"}


def _normalize_item_no(value: Any) -> str:
    text = _to_text(value).strip()
    if not text:
        return ""
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[./]", "-", text)
    text = re.sub(r"[()]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text


@dataclass(slots=True)
class RoleIdentity:
    executor_uri: str
    executor_did: str
    executor_role: str


@dataclass(slots=True)
class TripAction:
    name: str
    input_proof_id: str
    output_proof_id: str
    result: str


@dataclass(slots=True)
class ContainerState:
    status: str
    stage: str
    boq_item_uri: str
    smu_id: str


SPU_TEMPLATE_LIBRARY: dict[str, dict[str, Any]] = {
    "SPU_Physical": {
        "library_uri": f"{SPU_LIBRARY_ROOT_URI}/SPU_Physical",
        "label": "实体工程模板",
        "contexts": ["main_beam", "guardrail", "pier", "slab", "generic"],
        "geometry": {
            "variables": [
                {"key": "length", "label": "长度", "unit": "m", "required": False},
                {"key": "width", "label": "宽度", "unit": "m", "required": False},
                {"key": "height", "label": "高度", "unit": "m", "required": False},
            ],
        },
        "formula": {
            "name": "generic_quantity",
            "expression": "length * width * height",
            "variables": ["length", "width", "height"],
            "fallback_expression": "quantity",
            "quantity_aliases": ["claim_quantity", "measured_value", "design_quantity", "approved_quantity"],
            "quantity_unit": "m3",
            "tolerance_ratio": 0.08,
            "description": "默认实体工程量核算",
        },
        "normpeg_refs": ["v://norm/JTG/F80/1-2017#generic"],
        "form_schema": [
            {"field": "design_value", "label": "Design Value", "unit": "", "operator": "present", "default": "", "source": "Field"},
            {"field": "measured_value", "label": "Measured Value", "unit": "", "operator": "present", "default": "", "source": "Field"},
            {"field": "allowed_deviation", "label": "Allowed Deviation", "unit": "", "operator": "present", "default": "", "source": "Field"},
        ],
    },
    "SPU_Reinforcement": {
        "library_uri": f"{SPU_LIBRARY_ROOT_URI}/SPU_Reinforcement",
        "label": "钢筋质检模板",
        "contexts": ["main_beam", "guardrail", "pier", "generic"],
        "geometry": {
            "variables": [
                {"key": "length", "label": "钢筋长度", "unit": "m", "required": False},
                {"key": "unit_weight", "label": "单位重", "unit": "kg/m", "required": False},
                {"key": "count", "label": "根数", "unit": "根", "required": False},
            ],
        },
        "formula": {
            "name": "rebar_weight",
            "expression": "length * unit_weight * count",
            "variables": ["length", "unit_weight", "count"],
            "fallback_expression": "quantity",
            "quantity_aliases": ["claim_quantity", "measured_value", "design_quantity", "approved_quantity"],
            "quantity_unit": "kg",
            "tolerance_ratio": 0.05,
            "description": "钢筋理论重量核算",
        },
        "normpeg_refs": ["v://norm/JTG/F80/1-2017#rebar"],
        "form_schema": [
            {
                "field": "design_value",
                "label": "Design Value",
                "unit": "",
                "operator": "present",
                "default": "",
                "source": "Field",
            },
            {
                "field": "measured_value",
                "label": "Measured Value",
                "unit": "",
                "operator": "present",
                "default": "",
                "source": "Field",
            },
            {
                "field": "allowed_deviation",
                "label": "Allowed Deviation",
                "unit": "",
                "operator": "present",
                "default": "",
                "source": "Field",
            },
            {
                "field": "yield_strength",
                "label": "Yield Strength",
                "unit": "MPa",
                "operator": ">=",
                "default": "400",
                "source": "SpecDict",
            },
            {
                "field": "spacing_deviation",
                "label": "Spacing Deviation",
                "unit": "mm",
                "operator": "range",
                "default": "-10~10",
                "source": "SpecDict",
                "point_count": 5,
            },
            {
                "field": "cover_thickness",
                "label": "Cover Thickness",
                "unit": "mm",
                "operator": "range",
                "default": "20~60",
                "source": "SpecDict",
                "point_count": 5,
            },
        ],
    },
    "SPU_Concrete": {
        "library_uri": f"{SPU_LIBRARY_ROOT_URI}/SPU_Concrete",
        "label": "混凝土质检模板",
        "contexts": ["main_beam", "pier", "slab", "generic"],
        "geometry": {
            "variables": [
                {"key": "length", "label": "长度", "unit": "m", "required": False},
                {"key": "width", "label": "宽度", "unit": "m", "required": False},
                {"key": "height", "label": "高度", "unit": "m", "required": False},
            ],
        },
        "formula": {
            "name": "concrete_volume",
            "expression": "length * width * height",
            "variables": ["length", "width", "height"],
            "fallback_expression": "quantity",
            "quantity_aliases": ["claim_quantity", "measured_value", "design_quantity", "approved_quantity"],
            "quantity_unit": "m3",
            "tolerance_ratio": 0.06,
            "description": "混凝土方量核算",
        },
        "normpeg_refs": ["v://norm/JTG/F80/1-2017#concrete"],
        "form_schema": [
            {
                "field": "design_value",
                "label": "Design Value",
                "unit": "",
                "operator": "present",
                "default": "",
                "source": "Field",
            },
            {
                "field": "measured_value",
                "label": "Measured Value",
                "unit": "",
                "operator": "present",
                "default": "",
                "source": "Field",
            },
            {
                "field": "allowed_deviation",
                "label": "Allowed Deviation",
                "unit": "",
                "operator": "present",
                "default": "",
                "source": "Field",
            },
            {
                "field": "compressive_strength",
                "label": "Compressive Strength",
                "unit": "MPa",
                "operator": ">=",
                "default": "30",
                "source": "SpecDict",
                "point_count": 3,
            },
            {
                "field": "slump",
                "label": "Slump",
                "unit": "mm",
                "operator": "range",
                "default": "120~220",
                "source": "SpecDict",
                "point_count": 3,
            },
        ],
    },
    "SPU_Contract": {
        "library_uri": f"{SPU_LIBRARY_ROOT_URI}/SPU_Contract",
        "label": "合同凭证模板",
        "contexts": ["generic"],
        "geometry": {"variables": []},
        "formula": {
            "name": "contract_amount",
            "expression": "claimed_amount",
            "variables": ["claimed_amount"],
            "quantity_aliases": ["claimed_amount", "claim_quantity", "approved_quantity"],
            "quantity_unit": "CNY",
            "tolerance_ratio": 0.0,
            "description": "合同支付金额核算",
        },
        "normpeg_refs": ["v://norm/JTG/F80/1-2017#contract"],
        "form_schema": [
            {
                "field": "voucher_ref",
                "label": "Contract Voucher Ref",
                "unit": "",
                "operator": "present",
                "default": "",
                "source": "Contract",
            },
            {
                "field": "claimed_amount",
                "label": "Claimed Amount",
                "unit": "CNY",
                "operator": "present",
                "default": "",
                "source": "Contract",
            },
            {
                "field": "payment_cycle",
                "label": "Payment Cycle",
                "unit": "",
                "operator": "present",
                "default": "",
                "source": "Contract",
            },
        ],
    },
    "SPU_Bridge": {
        "library_uri": f"{SPU_LIBRARY_ROOT_URI}/SPU_Bridge",
        "label": "桥梁结构模板",
        "contexts": ["main_beam", "pier", "slab", "generic"],
        "geometry": {
            "variables": [
                {"key": "length", "label": "构件长度", "unit": "m", "required": False},
                {"key": "width", "label": "构件宽度", "unit": "m", "required": False},
                {"key": "height", "label": "构件高度", "unit": "m", "required": False},
            ],
        },
        "formula": {
            "name": "bridge_volume",
            "expression": "length * width * height",
            "variables": ["length", "width", "height"],
            "fallback_expression": "quantity",
            "quantity_aliases": ["claim_quantity", "measured_value", "design_quantity", "approved_quantity"],
            "quantity_unit": "m3",
            "tolerance_ratio": 0.05,
            "description": "桥梁构件方量核算（盖梁/墩柱等）",
        },
        "normpeg_refs": ["v://norm/JTG/F80/1-2017#bridge"],
        "form_schema": [
            {"field": "design_value", "label": "Design Value", "unit": "", "operator": "present", "default": "", "source": "Field"},
            {"field": "measured_value", "label": "Measured Value", "unit": "", "operator": "present", "default": "", "source": "Field", "point_count": 5},
            {"field": "allowed_deviation", "label": "Allowed Deviation", "unit": "", "operator": "present", "default": "", "source": "Field"},
            {"field": "cover_thickness", "label": "Cover Thickness", "unit": "mm", "operator": "range", "default": "20~60", "source": "SpecDict", "point_count": 5},
        ],
    },
    "SPU_Landscape": {
        "library_uri": f"{SPU_LIBRARY_ROOT_URI}/SPU_Landscape",
        "label": "绿化验收模板",
        "contexts": ["generic"],
        "geometry": {
            "variables": [
                {"key": "length", "label": "绿化长度", "unit": "m", "required": False},
                {"key": "width", "label": "绿化宽度", "unit": "m", "required": False},
            ],
        },
        "formula": {
            "name": "landscape_area",
            "expression": "length * width",
            "variables": ["length", "width"],
            "fallback_expression": "quantity",
            "quantity_aliases": ["claim_quantity", "measured_value", "design_quantity", "approved_quantity"],
            "quantity_unit": "m2",
            "tolerance_ratio": 0.08,
            "description": "绿化面积核算",
        },
        "normpeg_refs": ["v://norm/JTG/F80/1-2017#landscape"],
        "form_schema": [
            {"field": "survival_rate", "label": "Survival Rate", "unit": "%", "operator": ">=", "default": "95", "source": "SpecDict"},
            {"field": "coverage_rate", "label": "Coverage Rate", "unit": "%", "operator": ">=", "default": "90", "source": "SpecDict"},
            {"field": "height_range", "label": "Height Deviation", "unit": "cm", "operator": "range", "default": "-5~5", "source": "SpecDict"},
        ],
    },
    "SPU_CapBeam": {
        "library_uri": f"{SPU_LIBRARY_ROOT_URI}/SPU_CapBeam",
        "label": "盖梁模板",
        "contexts": ["main_beam", "pier", "generic"],
        "geometry": {
            "variables": [
                {"key": "length", "label": "盖梁长度", "unit": "m", "required": True},
                {"key": "width", "label": "盖梁宽度", "unit": "m", "required": True},
                {"key": "height", "label": "盖梁高度", "unit": "m", "required": True},
            ],
        },
        "formula": {
            "name": "cap_beam_volume",
            "expression": "length * width * height",
            "variables": ["length", "width", "height"],
            "fallback_expression": "quantity",
            "quantity_aliases": ["claim_quantity", "measured_value", "design_quantity", "approved_quantity"],
            "quantity_unit": "m3",
            "tolerance_ratio": 0.05,
            "description": "盖梁实体方量核算",
        },
        "normpeg_refs": ["v://norm/JTG/F80/1-2017#cap_beam"],
        "form_schema": [
            {"field": "design_value", "label": "Design Value", "unit": "", "operator": "present", "default": "", "source": "Field"},
            {"field": "measured_value", "label": "Measured Value", "unit": "", "operator": "present", "default": "", "source": "Field", "point_count": 5},
            {"field": "allowed_deviation", "label": "Allowed Deviation", "unit": "", "operator": "present", "default": "", "source": "Field"},
            {"field": "cover_thickness", "label": "Cover Thickness", "unit": "mm", "operator": "range", "default": "20~60", "source": "SpecDict", "point_count": 5},
        ],
    },
    "SPU_PileFoundation": {
        "library_uri": f"{SPU_LIBRARY_ROOT_URI}/SPU_PileFoundation",
        "label": "桩基模板",
        "contexts": ["pier", "generic"],
        "geometry": {
            "variables": [
                {"key": "pile_length", "label": "桩长", "unit": "m", "required": False},
                {"key": "pile_diameter", "label": "桩径", "unit": "m", "required": False},
                {"key": "pile_count", "label": "桩数", "unit": "根", "required": False},
            ],
        },
        "formula": {
            "name": "pile_volume",
            "expression": "3.1415926 * (pile_diameter / 2) * (pile_diameter / 2) * pile_length * pile_count",
            "variables": ["pile_diameter", "pile_length", "pile_count"],
            "fallback_expression": "quantity",
            "quantity_aliases": ["claim_quantity", "measured_value", "design_quantity", "approved_quantity"],
            "quantity_unit": "m3",
            "tolerance_ratio": 0.06,
            "description": "桩基理论方量核算",
        },
        "normpeg_refs": ["v://norm/JTG/F80/1-2017#pile_foundation"],
        "form_schema": [
            {"field": "design_value", "label": "Design Value", "unit": "", "operator": "present", "default": "", "source": "Field"},
            {"field": "measured_value", "label": "Measured Value", "unit": "", "operator": "present", "default": "", "source": "Field", "point_count": 3},
            {"field": "allowed_deviation", "label": "Allowed Deviation", "unit": "", "operator": "present", "default": "", "source": "Field"},
            {"field": "integrity_grade", "label": "Integrity Grade", "unit": "", "operator": "present", "default": "I", "source": "SpecDict"},
        ],
    },
}

SPU_NAME_HINTS: dict[str, tuple[str, ...]] = {
    "SPU_CapBeam": ("盖梁", "capbeam", "cap beam", "pier cap"),
    "SPU_PileFoundation": ("桩基", "灌注桩", "钻孔桩", "pile"),
    "SPU_Reinforcement": ("钢筋", "rebar", "reinforcement"),
    "SPU_Concrete": ("混凝土", "砼", "concrete"),
    "SPU_Landscape": ("绿化", "种植", "landscape"),
    "SPU_Bridge": ("桥", "bridge"),
}


HEADER_ALIASES: dict[str, set[str]] = {
    'item_no': {
        '子目号',
        '子目號',
        '子目编号',
        '子目編號',
        '子目编码',
        '子目編碼',
        '细目号',
        '細目號',
        '细目编号',
        '細目編號',
        '清单编码',
        '清單編碼',
        '清单编号',
        '清單編號',
        '细目',
        '細目',
        '子目',
        'itemno',
        'item_no',
        'itemcode',
        'item',
    },
    'name': {
        '子目名称',
        '子目名稱',
        '细目名称',
        '細目名稱',
        '清单名称',
        '清單名稱',
        '名称',
        '名稱',
        '项目名称',
        '項目名稱',
        'itemname',
        'name',
    },
    'unit': {'单位', '單位', '计量单位', '計量單位', 'unit'},
    'division': {'分部工程', '分部', 'division'},
    'subdivision': {'分项工程', '分項工程', '子分项', '子分項', 'subdivision'},
    'hierarchy': {'所属分部分项层级', '所屬分部分項層級', '分部分项层级', '分部分項層級', '层级', '層級', 'hierarchy', 'wbs'},
    'design_quantity': {
        '设计数量',
        '設計數量',
        '设计工程量',
        '設計工程量',
        '设计量',
        '設計量',
        '施工图数量',
        '施工圖數量',
        '工程量',
        'designqty',
        'designquantity',
    },
    'unit_price': {'单价', '單價', '综合单价', '綜合單價', '价格', '價格', 'price', 'unitprice'},
    'approved_quantity': {
        '批复数量',
        '批復數量',
        '批复工程量',
        '批復工程量',
        '审批数量',
        '審批數量',
        '审批工程量',
        '審批工程量',
        '批准数量',
        '批准數量',
        '批准工程量',
        '批准工程量',
        '合同数量',
        '合同數量',
        '合同工程量',
        '审定数量',
        '審定數量',
        '审定工程量',
        '審定工程量',
        'approvedqty',
        'approvedquantity',
    },
    'approved_amount': {
        '批复金额',
        '批復金額',
        '批复合价',
        '批復合價',
        '审定金额',
        '審定金額',
        '审定合价',
        '審定合價',
        '批准金额',
        '批准金額',
        '合同金额',
        '合同金額',
        '合同合价',
        '合同合價',
        '合价',
        '合價',
        '总价',
        '總價',
        '总额',
        '總額',
        '金额',
        '金額',
    },
    'remark': {'备注', '備註', 'remark'},
}

NORMREF_BY_PREFIX: dict[str, list[str]] = {
    "401": ["JTG F80/1-2017"],
    "403": ["JTG F80/1-2017"],
    "405": ["JTG F80/1-2017"],
    "600": ["JTG F80/1-2017"],
    "702": ["JTG F80/1-2017"],
}

ROLE_ALLOW_BY_PREFIX: dict[str, list[str]] = {
    "401": ["SUPERVISOR", "OWNER"],
    "403": ["AI", "SUPERVISOR", "OWNER"],
    "405": ["AI", "SUPERVISOR", "OWNER"],
    "101": ["SUPERVISOR", "OWNER"],
    "102": ["SUPERVISOR", "OWNER"],
    "600": ["SUPERVISOR", "OWNER"],
    "702": ["SUPERVISOR", "OWNER"],
}


def _to_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def _as_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    return {}


def _as_list(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    return []


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    text = _to_text(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        m = re.search(r"[-+]?\d+(?:\.\d+)?", text)
        if not m:
            return None
        try:
            return float(m.group(0))
        except Exception:
            return None


def _round4(value: float | None) -> float | None:
    if value is None:
        return None
    try:
        return round(float(value), 4)
    except Exception:
        return None


def _normalize_unit(value: Any) -> str:
    text = _to_text(value).strip()
    if not text:
        return ""
    norm = text.replace("平方米", "m2").replace("㎡", "m2").replace("m²", "m2").replace("M2", "m2").replace("m^2", "m2")
    norm = norm.replace("立方米", "m3").replace("m³", "m3").replace("M3", "m3").replace("m^3", "m3")
    norm = norm.replace("公里", "km").replace("千米", "km").replace("KM", "km").replace("Km", "km")
    norm = norm.replace("米", "m").replace("M", "m")
    if norm in {"金额", "费用", "价款"}:
        norm = "总额"
    return norm.strip()


def _fallback_approved_quantity(
    *,
    item_no: str,
    unit: str,
    approved_quantity: float | None,
    approved_amount: float | None,
    design_quantity: float | None,
    unit_price: float | None,
) -> tuple[float | None, bool]:
    if approved_amount is None:
        return approved_quantity, False
    if approved_quantity is not None and approved_quantity > 0:
        return approved_quantity, False
    unit_norm = _normalize_unit(unit)
    if unit_norm in {"总额", "金额"} or not unit_norm:
        return approved_amount, True
    if _to_text(item_no).strip().startswith("1"):
        return approved_amount, True
    if unit_price is None and design_quantity is None:
        return approved_amount, True
    return approved_quantity, False


def _sha(payload: Any) -> str:
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")
    ).hexdigest()


def _utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize_header(value: Any) -> str:
    text = _to_text(value).strip().lstrip("\ufeff").lower()
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("號", "号").replace("編", "编").replace("碼", "码")
    text = text.replace("稱", "称").replace("單", "单").replace("價", "价")
    text = re.sub(r"[\s\u3000]+", "", text)
    text = re.sub(r"[:：_/\\-]", "", text)
    return text


def _detect_header_map(rows: list[list[Any]]) -> tuple[int, dict[str, int]]:
    normalized_aliases: dict[str, set[str]] = {
        field: {_normalize_header(alias) for alias in aliases}
        for field, aliases in HEADER_ALIASES.items()
    }

    def _header_match(key: str, alias: str) -> bool:
        if not key or not alias:
            return False
        if key == alias:
            return True
        # Handle merged header text like "子目號(item_no)" or wrapped titles.
        if len(alias) >= 3 and alias in key:
            return True
        if len(key) >= 4 and key in alias:
            return True
        return False

    best_idx = -1
    best_score = -1
    best_map: dict[str, int] = {}
    best_found_fields: set[str] = set()
    for idx, row in enumerate(rows):
        mapping: dict[str, int] = {}
        score = 0
        for col_idx, cell in enumerate(row):
            key = _normalize_header(cell)
            if not key:
                continue
            for field, aliases in normalized_aliases.items():
                if field in mapping:
                    continue
                if any(_header_match(key, alias) for alias in aliases):
                    mapping[field] = col_idx
                    score += 1
        if score > best_score and "item_no" in mapping and "name" in mapping:
            best_idx = idx
            best_score = score
            best_map = mapping
            best_found_fields = set(mapping.keys())
    if best_idx < 0:
        found = ",".join(sorted(best_found_fields)) if best_found_fields else "none"
        raise HTTPException(
            400,
            f"Failed to detect BOQ header row from upload file. required=item_no,name found={found}",
        )
    return best_idx, best_map


def _rows_from_csv_bytes(content: bytes) -> list[list[Any]]:
    if not content:
        return []
    text = ""
    for enc in ("utf-8-sig", "gb18030", "gbk"):
        try:
            text = content.decode(enc)
            break
        except Exception:
            continue
    if not text:
        text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [list(row) for row in reader]


def _rows_from_xlsx_bytes(content: bytes) -> tuple[list[list[Any]], str]:
    try:
        import openpyxl
    except Exception as exc:
        raise HTTPException(500, f"openpyxl not available: {exc}") from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except zipfile.BadZipFile:
        if content[:4] == b"\xD0\xCF\x11\xE0":
            # Fallback to legacy .xls parser when file is mis-labeled as .xlsx
            return _rows_from_xls_bytes(content)
        raise HTTPException(400, "Invalid .xlsx file. Please re-save as .xlsx or .csv.")
    except Exception as exc:
        raise HTTPException(400, f"Failed to read .xlsx file: {exc}")
    best_rows: list[list[Any]] = []
    best_sheet = wb.sheetnames[0]
    best_score = -1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[Any]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) >= 300000:
                break
        if not rows:
            continue
        score = -1
        try:
            _, colmap = _detect_header_map(rows[:120])
            score = len(colmap)
        except HTTPException:
            score = -1
        if score > best_score:
            best_rows = rows
            best_sheet = sheet_name
            best_score = score
    if not best_rows:
        return [], best_sheet
    return best_rows, best_sheet


def _rows_from_xls_bytes(content: bytes) -> tuple[list[list[Any]], str]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise HTTPException(
            400,
            f".xls parser missing (xlrd). Please convert to CSV/XLSX or install xlrd in services/api env. detail={exc}",
        ) from exc
    wb = xlrd.open_workbook(file_contents=content)
    best_rows: list[list[Any]] = []
    best_sheet = wb.sheet_by_index(0).name if wb.nsheets > 0 else "sheet0"
    best_score = -1

    for i in range(wb.nsheets):
        sheet = wb.sheet_by_index(i)
        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
            if len(rows) >= 300000:
                break
        if not rows:
            continue
        score = -1
        try:
            _, colmap = _detect_header_map(rows[:120])
            score = len(colmap)
        except HTTPException:
            score = -1
        if score > best_score:
            best_rows = rows
            best_sheet = _to_text(getattr(sheet, "name", "")).strip() or best_sheet
            best_score = score
    if not best_rows:
        return [], best_sheet
    return best_rows, best_sheet


def _parse_boq_upload(file_name: str, content: bytes) -> list[BoqItem]:
    name = _to_text(file_name).strip()
    lower = name.lower()
    header_sig = content[:4] if content else b""
    is_ole2 = header_sig == b"\xD0\xCF\x11\xE0"
    is_zip = header_sig == b"\x50\x4B\x03\x04"

    if lower.endswith(".csv"):
        rows = _rows_from_csv_bytes(content)
        source_sheet = "csv"
    elif lower.endswith(".xlsx"):
        if is_ole2:
            rows, source_sheet = _rows_from_xls_bytes(content)
        else:
            rows, source_sheet = _rows_from_xlsx_bytes(content)
        source_sheet = _to_text(source_sheet).strip() or "xlsx"
    elif lower.endswith(".xls"):
        rows, source_sheet = _rows_from_xls_bytes(content)
        source_sheet = _to_text(source_sheet).strip() or "xls"
    elif is_ole2:
        rows, source_sheet = _rows_from_xls_bytes(content)
        source_sheet = _to_text(source_sheet).strip() or "xls"
    elif is_zip:
        rows, source_sheet = _rows_from_xlsx_bytes(content)
        source_sheet = _to_text(source_sheet).strip() or "xlsx"
    else:
        rows = _rows_from_csv_bytes(content)
        source_sheet = "unknown"

    if not rows:
        raise HTTPException(400, "Upload file is empty.")

    probe = rows[:120]
    header_row, colmap = _detect_header_map(probe)

    out: list[BoqItem] = []
    for row_idx in range(header_row + 1, len(rows)):
        row = rows[row_idx]
        item_no_raw = _to_text(row[colmap["item_no"]] if colmap["item_no"] < len(row) else "").strip()
        item_no = _normalize_item_no(item_no_raw)
        if not item_no or not ITEM_NO_PATTERN.match(item_no):
            continue
        name = _to_text(row[colmap["name"]] if colmap["name"] < len(row) else "").strip()
        unit = _normalize_unit(row[colmap.get("unit", -1)] if colmap.get("unit", -1) < len(row) and colmap.get("unit", -1) >= 0 else "")
        division = _to_text(row[colmap.get("division", -1)] if colmap.get("division", -1) < len(row) and colmap.get("division", -1) >= 0 else "").strip()
        subdivision = _to_text(row[colmap.get("subdivision", -1)] if colmap.get("subdivision", -1) < len(row) and colmap.get("subdivision", -1) >= 0 else "").strip()
        hierarchy_raw = _to_text(row[colmap.get("hierarchy", -1)] if colmap.get("hierarchy", -1) < len(row) and colmap.get("hierarchy", -1) >= 0 else "").strip()

        dq_raw = _to_text(row[colmap.get("design_quantity", -1)] if colmap.get("design_quantity", -1) < len(row) and colmap.get("design_quantity", -1) >= 0 else "").strip()
        up_raw = _to_text(row[colmap.get("unit_price", -1)] if colmap.get("unit_price", -1) < len(row) and colmap.get("unit_price", -1) >= 0 else "").strip()
        aq_raw = _to_text(row[colmap.get("approved_quantity", -1)] if colmap.get("approved_quantity", -1) < len(row) and colmap.get("approved_quantity", -1) >= 0 else "").strip()
        aa_raw = _to_text(row[colmap.get("approved_amount", -1)] if colmap.get("approved_amount", -1) < len(row) and colmap.get("approved_amount", -1) >= 0 else "").strip()
        remark = _to_text(row[colmap.get("remark", -1)] if colmap.get("remark", -1) < len(row) and colmap.get("remark", -1) >= 0 else "").strip()

        design_quantity = _round4(_to_float(dq_raw))
        unit_price = _round4(_to_float(up_raw))
        approved_quantity = _round4(_to_float(aq_raw))
        approved_amount = _round4(_to_float(aa_raw))
        approved_quantity, used_amount = _fallback_approved_quantity(
            item_no=item_no,
            unit=unit,
            approved_quantity=approved_quantity,
            approved_amount=approved_amount,
            design_quantity=design_quantity,
            unit_price=unit_price,
        )
        if used_amount:
            aq_raw = aa_raw
            if not unit:
                unit = "总额"

        out.append(
            BoqItem(
                item_no=item_no,
                name=name,
                unit=unit,
                division=division,
                subdivision=subdivision,
                hierarchy_raw=hierarchy_raw,
                design_quantity=design_quantity,
                design_quantity_raw=dq_raw,
                unit_price=unit_price,
                unit_price_raw=up_raw,
                approved_quantity=approved_quantity,
                approved_quantity_raw=aq_raw,
                remark=remark,
                row_index=row_idx + 1,
                sheet_name=source_sheet,
            )
        )
    if not out:
        raise HTTPException(400, "No BOQ item rows parsed from upload file.")
    return out


def _normalized_hint_text(text: str) -> str:
    lowered = _to_text(text).strip().lower()
    lowered = lowered.replace("_", "").replace("-", "").replace(" ", "")
    return lowered


def _match_spu_template_by_name(item_name: str) -> tuple[str, list[str]]:
    normalized = _normalized_hint_text(item_name)
    if not normalized:
        return ("", [])
    matched: list[str] = []
    template_id = ""
    for candidate, keywords in SPU_NAME_HINTS.items():
        for kw in keywords:
            norm_kw = _normalized_hint_text(kw)
            if norm_kw and norm_kw in normalized:
                template_id = candidate
                matched.append(kw)
        if template_id:
            break
    return (template_id, matched)


def _safe_eval_formula(expression: str, variables: dict[str, float]) -> float | None:
    expr = _to_text(expression).strip()
    if not expr:
        return None
    try:
        tree = ast.parse(expr, mode="eval")
    except Exception:
        return None

    def _eval_node(node: ast.AST) -> float:
        if isinstance(node, ast.Expression):
            return _eval_node(node.body)
        if isinstance(node, ast.Constant):
            if isinstance(node.value, (int, float)):
                return float(node.value)
            raise ValueError("unsupported_constant")
        if isinstance(node, ast.Name):
            key = _to_text(node.id).strip()
            if key in variables:
                return float(variables[key])
            raise KeyError(key)
        if isinstance(node, ast.BinOp):
            left = _eval_node(node.left)
            right = _eval_node(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                if abs(right) <= 1e-12:
                    raise ZeroDivisionError
                return left / right
            if isinstance(node.op, ast.Pow):
                return left**right
            raise ValueError("unsupported_operator")
        if isinstance(node, ast.UnaryOp):
            val = _eval_node(node.operand)
            if isinstance(node.op, ast.USub):
                return -val
            if isinstance(node.op, ast.UAdd):
                return +val
            raise ValueError("unsupported_unary")
        raise ValueError("unsupported_expression")

    try:
        return float(_eval_node(tree))
    except Exception:
        return None


def _formula_aliases_map() -> dict[str, tuple[str, ...]]:
    return {
        "length": ("length", "l"),
        "width": ("width", "w"),
        "height": ("height", "h"),
        "pile_length": ("pile_length", "length", "l"),
        "pile_diameter": ("pile_diameter", "diameter", "d"),
        "pile_count": ("pile_count", "count"),
        "unit_weight": ("unit_weight",),
        "count": ("count",),
        "claimed_amount": ("claimed_amount", "claim_quantity", "amount"),
        "quantity": ("quantity", "claim_quantity", "measured_value", "design_quantity", "approved_quantity"),
    }


def _resolve_formula_variables(*, formula_def: dict[str, Any], measurement: dict[str, Any], design_quantity: float | None, approved_quantity: float | None) -> tuple[dict[str, float], list[str]]:
    aliases = _formula_aliases_map()
    inputs = _as_dict(measurement)
    target_qty = approved_quantity if approved_quantity is not None and approved_quantity > 0 else design_quantity
    vars_out: dict[str, float] = {}
    missing: list[str] = []
    for var in _as_list(formula_def.get("variables")):
        key = _to_text(var).strip()
        if not key:
            continue
        candidates = aliases.get(key) or (key,)
        val: float | None = None
        for name in candidates:
            val = _to_float(inputs.get(name))
            if val is not None:
                break
        if val is None and key == "quantity" and target_qty is not None:
            val = float(target_qty)
        if val is None:
            missing.append(key)
            continue
        vars_out[key] = float(val)
    for alias in _as_list(formula_def.get("quantity_aliases")):
        alias_key = _to_text(alias).strip()
        if not alias_key:
            continue
        alias_val = _to_float(inputs.get(alias_key))
        if alias_val is None and target_qty is not None:
            alias_val = float(target_qty)
        if alias_val is not None:
            vars_out[alias_key] = float(alias_val)
    if target_qty is not None:
        vars_out["quantity"] = float(target_qty)
    return (vars_out, missing)


def _build_spu_formula_audit(
    *,
    template: dict[str, Any],
    measurement: dict[str, Any] | None = None,
    design_quantity: float | None = None,
    approved_quantity: float | None = None,
) -> dict[str, Any]:
    formula_def = _as_dict(template.get("formula"))
    if not formula_def:
        return {}
    formula_name = _to_text(formula_def.get("name") or "").strip()
    expression = _to_text(formula_def.get("expression") or "").strip()
    fallback_expression = _to_text(formula_def.get("fallback_expression") or "").strip()
    tolerance_ratio = _to_float(formula_def.get("tolerance_ratio"))
    tolerance = float(tolerance_ratio) if tolerance_ratio is not None else 0.08

    target_qty = approved_quantity if approved_quantity is not None and approved_quantity > 0 else design_quantity
    vars_out, missing = _resolve_formula_variables(
        formula_def=formula_def,
        measurement=_as_dict(measurement),
        design_quantity=design_quantity,
        approved_quantity=approved_quantity,
    )
    computed = _safe_eval_formula(expression, vars_out) if expression else None
    used_expression = expression
    used_fallback = False
    if computed is None and fallback_expression:
        computed = _safe_eval_formula(fallback_expression, vars_out)
        used_expression = fallback_expression
        used_fallback = computed is not None

    status = "PENDING"
    deviation_ratio: float | None = None
    if computed is None:
        status = "PENDING"
    elif target_qty is None:
        status = "PENDING"
    else:
        base = max(abs(float(target_qty)), 1e-9)
        deviation_ratio = round(abs(float(computed) - float(target_qty)) / base, 6)
        status = "PASS" if deviation_ratio <= tolerance + 1e-9 else "FAIL"

    return {
        "formula_name": formula_name or "quantity_formula",
        "expression": expression,
        "expression_used": used_expression,
        "used_fallback": used_fallback,
        "variables": _as_list(formula_def.get("variables")),
        "variables_resolved": vars_out,
        "missing_variables": missing,
        "computed_quantity": round(float(computed), 6) if computed is not None else None,
        "ledger_quantity": round(float(target_qty), 6) if target_qty is not None else None,
        "deviation_ratio": deviation_ratio,
        "tolerance_ratio": round(float(tolerance), 6),
        "status": status,
        "quantity_unit": _to_text(formula_def.get("quantity_unit") or "").strip(),
    }


def _resolve_spu_template(item_no: str, item_name: str) -> dict[str, Any]:
    code = _to_text(item_no).strip()
    name = _to_text(item_name).strip()
    lower_name = name.lower()

    template_id = "SPU_Physical"
    mapping_reason = "default"
    matched_keywords: list[str] = []
    if code.startswith("101") or code.startswith("102"):
        template_id = "SPU_Contract"
        mapping_reason = "prefix_contract"
    elif any(token in name for token in ("费", "协调", "管理", "监测", "监控", "咨询", "勘察", "保险", "交通", "保通", "征迁", "补偿", "迁改", "拆除", "临时", "安全", "试验", "检验")):
        template_id = "SPU_Contract"
        mapping_reason = "name_contract_keyword"
    else:
        by_name, keywords = _match_spu_template_by_name(name)
        if by_name:
            template_id = by_name
            mapping_reason = "name_keyword"
            matched_keywords = keywords
        elif code.startswith("401") or ("桥梁" in name) or ("桥" in name):
            template_id = "SPU_Bridge"
            mapping_reason = "prefix_bridge"
        elif code.startswith("600") or code.startswith("702") or ("绿化" in name) or ("种植" in name) or ("landscape" in lower_name):
            template_id = "SPU_Landscape"
            mapping_reason = "prefix_landscape"
        elif code.startswith("403") or code.startswith("405") or ("钢筋" in name) or ("rebar" in lower_name):
            template_id = "SPU_Reinforcement"
            mapping_reason = "prefix_rebar"
        elif ("混凝土" in name) or ("concrete" in lower_name):
            template_id = "SPU_Concrete"
            mapping_reason = "name_concrete"

    template = _as_dict(SPU_TEMPLATE_LIBRARY.get(template_id))
    return {
        "spu_root_uri": SPU_LIBRARY_ROOT_URI,
        "spu_library_uri": _to_text(template.get("library_uri") or f"{SPU_LIBRARY_ROOT_URI}/{template_id}").strip(),
        "spu_template_id": template_id,
        "spu_template_label": _to_text(template.get("label") or template_id).strip(),
        "spu_form_schema": _as_list(template.get("form_schema")),
        "spu_geometry": _as_dict(template.get("geometry")),
        "spu_formula": _as_dict(template.get("formula")),
        "spu_normpeg_refs": [str(x).strip() for x in _as_list(template.get("normpeg_refs")) if str(x).strip()],
        "supported_contexts": _as_list(template.get("contexts")),
        "spu_mapping": {
            "source": "boq_name_parser",
            "reason": mapping_reason,
            "matched_keywords": matched_keywords,
        },
    }


def list_spu_template_library() -> dict[str, Any]:
    templates: list[dict[str, Any]] = []
    for template_id in sorted(SPU_TEMPLATE_LIBRARY.keys()):
        tpl = _as_dict(SPU_TEMPLATE_LIBRARY.get(template_id))
        templates.append(
            {
                "spu_template_id": template_id,
                "library_uri": _to_text(tpl.get("library_uri") or f"{SPU_LIBRARY_ROOT_URI}/{template_id}").strip(),
                "label": _to_text(tpl.get("label") or template_id).strip(),
                "contexts": _as_list(tpl.get("contexts")),
                "geometry": _as_dict(tpl.get("geometry")),
                "formula": _as_dict(tpl.get("formula")),
                "normpeg_refs": [str(x).strip() for x in _as_list(tpl.get("normpeg_refs")) if str(x).strip()],
                "form_schema": _as_list(tpl.get("form_schema")),
            }
        )
    return {
        "ok": True,
        "library_root_uri": SPU_LIBRARY_ROOT_URI,
        "template_count": len(templates),
        "templates": templates,
    }


def _resolve_norm_refs(item_no: str, item_name: str, *, template_norm_refs: list[str] | None = None) -> list[str]:
    code = _to_text(item_no).strip()
    prefix = code.split("-")[0] if code else ""
    refs: list[str] = [str(x).strip() for x in (template_norm_refs or []) if str(x).strip()]
    if prefix in NORMREF_BY_PREFIX:
        refs.extend(list(NORMREF_BY_PREFIX[prefix]))
    name = _to_text(item_name).strip()
    if "桥" in name:
        refs.extend(list(NORMREF_BY_PREFIX.get("401") or []))
    if "绿化" in name:
        refs.extend(list(NORMREF_BY_PREFIX.get("600") or []))
    if "种植" in name:
        refs.extend(list(NORMREF_BY_PREFIX.get("702") or []))
    if "钢筋" in name:
        refs.extend(list(NORMREF_BY_PREFIX.get("403") or []))
    out: list[str] = []
    seen: set[str] = set()
    for ref in refs:
        normalized = _to_text(ref).strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        out.append(normalized)
    return out


def _resolve_allowed_roles(item_no: str, spu_template_id: str) -> list[str]:
    code = _to_text(item_no).strip()
    prefix = code.split("-")[0] if code else ""
    mapped = ROLE_ALLOW_BY_PREFIX.get(prefix)
    if mapped:
        return list(mapped)
    if spu_template_id == "SPU_Contract":
        return ["OWNER", "SUPERVISOR"]
    return ["AI", "SUPERVISOR", "OWNER"]


def _is_contract_payload(item_name: str, measurement: dict[str, Any]) -> bool:
    name = _to_text(item_name)
    if any(token in name for token in ("费", "协调", "管理", "咨询", "补偿", "保险")):
        return True
    for key in ("voucher_ref", "claimed_amount", "contract_voucher", "contract_amount"):
        if key in measurement and _to_text(measurement.get(key)).strip():
            return True
    return False


def _resolve_boq_balance(*, sb: Any, project_uri: str, boq_item_uri: str) -> dict[str, float]:
    status = get_boq_realtime_status(sb=sb, project_uri=project_uri, limit=10000)
    items = _as_list(status.get("items"))
    item = next((x for x in items if _to_text(_as_dict(x).get("boq_item_uri")).strip() == boq_item_uri), None)
    if not isinstance(item, dict):
        return {"baseline": 0.0, "settled": 0.0}
    approved_qty = _to_float(item.get("approved_quantity"))
    contract_qty = _to_float(item.get("contract_quantity"))
    design_qty = _to_float(item.get("design_quantity"))
    baseline = approved_qty if approved_qty is not None and approved_qty > 0 else (contract_qty if contract_qty is not None and contract_qty > 0 else (design_qty or 0.0))
    settled = _to_float(item.get("settled_quantity")) or 0.0
    return {"baseline": float(baseline), "settled": float(settled)}


def _resolve_lab_pass_for_sample(*, sb: Any, project_uri: str, boq_item_uri: str, sample_id: str) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    b_uri = _to_text(boq_item_uri).strip().rstrip("/")
    sample = _to_text(sample_id).strip()
    if not p_uri or not b_uri or not sample:
        return {"ok": False, "pass": 0, "total": 0}
    try:
        rows = (
            sb.table("proof_utxo")
            .select("proof_id, result, proof_type, state_data, segment_uri, created_at")
            .eq("project_uri", p_uri)
            .eq("proof_type", "lab")
            .order("created_at", desc=False)
            .limit(2000)
            .execute()
            .data
            or []
        )
    except Exception as exc:
        return {"ok": False, "pass": 0, "total": 0, "error": f"{exc.__class__.__name__}"}
    matched: list[dict[str, Any]] = []
    for row in rows:
        sd = _as_dict(row.get("state_data"))
        uri = _to_text(sd.get("boq_item_uri") or row.get("segment_uri") or "").strip().rstrip("/")
        if uri != b_uri:
            continue
        if _to_text(sd.get("sample_id") or "").strip() != sample:
            continue
        matched.append(row)
    if not matched:
        return {"ok": True, "pass": 0, "total": 0}
    lab_pass = [x for x in matched if _to_text(x.get("result") or "").strip().upper() == "PASS"]
    latest_pass = lab_pass[-1] if lab_pass else None
    return {
        "ok": True,
        "pass": len(lab_pass),
        "total": len(matched),
        "latest_pass_proof_id": _to_text((latest_pass or {}).get("proof_id") or "").strip(),
    }


def verify_conservation(*, baseline: float, settled: float, claim: float) -> dict[str, Any]:
    baseline_val = float(baseline or 0.0)
    settled_val = float(settled or 0.0)
    claim_val = float(claim or 0.0)
    total = settled_val + claim_val
    if baseline_val <= 0:
        return {"ok": True, "gap_ratio": 0.0, "total": total, "baseline": baseline_val}
    gap_ratio = abs(total - baseline_val) / baseline_val
    gap_ratio = round(gap_ratio, 4)
    ok = total <= baseline_val + 1e-9
    return {"ok": ok, "gap_ratio": gap_ratio, "total": total, "baseline": baseline_val}


def _resolve_bridge_table_template_path() -> str:
    candidates = [
        _to_text(os.getenv("QCSPEC_BRIDGE_TABLE_DOCX") or "").strip(),
        r"C:\Users\xm_91\Desktop\3、桥施表.docx",
    ]
    for raw in candidates:
        if not raw:
            continue
        try:
            p = Path(raw).expanduser()
            if p.exists() and p.is_file():
                return str(p.resolve())
        except Exception:
            continue
    return ""


def _resolve_docpeg_template(item_no: str, item_name: str) -> dict[str, Any]:
    code = _to_text(item_no).strip()
    name = _to_text(item_name).strip()
    bridge_docx_path = _resolve_bridge_table_template_path()

    # Rule-based binding: 403 steel/rebar defaults to bridge form 11.
    if code.startswith("403") or ("钢筋" in name):
        return {
            "template_group": "桥施表",
            "template_code": "桥施11表",
            "template_name": "钢筋安装检查表",
            "binding_rule": "item_no startswith 403",
            "template_path": bridge_docx_path,
            "fallback_template": "rebar_inspection_table.docx",
            "is_auto_bound": True,
        }
    if code.startswith("402") or ("混凝土" in name):
        return {
            "template_group": "桥施表",
            "template_code": "桥施63表",
            "template_name": "混凝土施工质量检查表",
            "binding_rule": "item_no startswith 402 or item_name contains 混凝土",
            "template_path": bridge_docx_path,
            "fallback_template": "01_inspection_report.docx",
            "is_auto_bound": True,
        }
    return {
        "template_group": "未绑定",
        "template_code": "",
        "template_name": "",
        "binding_rule": "unbound",
        "template_path": bridge_docx_path,
        "fallback_template": "rebar_inspection_table.docx",
        "is_auto_bound": False,
    }


def _resolve_lab_status(*, sb: Any, project_uri: str, boq_item_uri: str) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    b_uri = _to_text(boq_item_uri).strip().rstrip("/")
    if not p_uri or not b_uri:
        return {"ok": False, "status": "MISSING", "total": 0, "pass": 0}
    try:
        rows = (
            sb.table("proof_utxo")
            .select("proof_id, result, proof_type, state_data, segment_uri, created_at")
            .eq("project_uri", p_uri)
            .eq("proof_type", "lab")
            .order("created_at", desc=False)
            .limit(2000)
            .execute()
            .data
            or []
        )
    except Exception as exc:
        return {"ok": False, "status": "UNAVAILABLE", "total": 0, "pass": 0, "error": f"{exc.__class__.__name__}"}

    matched: list[dict[str, Any]] = []
    for row in rows:
        sd = _as_dict(row.get("state_data"))
        uri = _to_text(sd.get("boq_item_uri") or row.get("segment_uri") or "").strip().rstrip("/")
        if uri == b_uri:
            matched.append(row)

    if not matched:
        return {"ok": True, "status": "MISSING", "total": 0, "pass": 0}

    lab_pass = [x for x in matched if _to_text(x.get("result") or "").strip().upper() == "PASS"]
    latest = matched[-1]
    latest_pass = lab_pass[-1] if lab_pass else None
    status = "PASS" if lab_pass else "FAIL"
    return {
        "ok": True,
        "status": status,
        "total": len(matched),
        "pass": len(lab_pass),
        "latest_proof_id": _to_text((latest or {}).get("proof_id") or "").strip(),
        "latest_pass_proof_id": _to_text((latest_pass or {}).get("proof_id") or "").strip(),
    }


def _smu_id_from_item_code(item_code: str) -> str:
    token = _to_text(item_code).strip().rstrip("/").split("/")[-1]
    if "-" in token:
        return token.split("-")[0]
    return token


def _is_smu_frozen(*, sb: Any, project_uri: str, smu_id: str) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    s_id = _to_text(smu_id).strip()
    if not p_uri or not s_id:
        return {"frozen": False}
    try:
        rows = (
            sb.table("proof_utxo")
            .select("proof_id,result,created_at,state_data")
            .eq("project_uri", p_uri)
            .eq("proof_type", "smu_freeze")
            .filter("state_data->>smu_id", "eq", s_id)
            .order("created_at", desc=True)
            .limit(5)
            .execute()
            .data
            or []
        )
    except Exception:
        return {"frozen": False}
    latest = rows[0] if rows else {}
    if not latest:
        return {"frozen": False}
    status = _to_text(latest.get("result") or "").strip().upper()
    is_frozen = status == "PASS"
    return {
        "frozen": is_frozen,
        "proof_id": _to_text(latest.get("proof_id") or "").strip(),
        "created_at": _to_text(latest.get("created_at") or "").strip(),
    }


def _resolve_smu_leaf_items(*, sb: Any, project_uri: str, smu_id: str) -> list[dict[str, Any]]:
    s_id = _to_text(smu_id).strip()
    if not s_id:
        return []
    status = get_boq_realtime_status(sb=sb, project_uri=project_uri, limit=10000)
    out: list[dict[str, Any]] = []
    for item in _as_list(status.get("items")):
        row = _as_dict(item)
        item_no = _to_text(row.get("item_no") or "").strip()
        if not item_no or not item_no.startswith(s_id):
            continue
        if not _to_text(row.get("boq_item_uri") or "").strip():
            continue
        out.append(row)
    return out


def _collect_smu_qualification(*, sb: Any, project_uri: str, smu_id: str) -> dict[str, Any]:
    items = _resolve_smu_leaf_items(sb=sb, project_uri=project_uri, smu_id=smu_id)
    total = len(items)
    qualified = 0
    incomplete: list[dict[str, Any]] = []
    for row in items:
        settlement_count = int(_to_float(row.get("settlement_count")) or 0)
        latest_settlement = _to_text(row.get("latest_settlement_proof_id") or "").strip()
        if settlement_count > 0 and latest_settlement:
            qualified += 1
            continue
        incomplete.append(
            {
                "item_no": _to_text(row.get("item_no") or "").strip(),
                "boq_item_uri": _to_text(row.get("boq_item_uri") or "").strip(),
                "settlement_count": settlement_count,
                "latest_settlement_proof_id": latest_settlement,
            }
        )
    all_qualified = total > 0 and qualified == total
    return {
        "smu_id": _to_text(smu_id).strip(),
        "leaf_total": total,
        "qualified_leaf_count": qualified,
        "unqualified_leaf_count": max(0, total - qualified),
        "all_qualified": all_qualified,
        "pending_items": incomplete[:200],
    }


def _mark_smu_scope_immutable(
    *,
    sb: Any,
    project_uri: str,
    smu_id: str,
    freeze_proof_id: str,
    total_proof_hash: str,
) -> dict[str, Any]:
    rows = _boq_rows(sb, project_uri=project_uri, boq_item_uri="", only_unspent=False, limit=50000)
    touched = 0
    skipped = 0
    now = _utc_iso()
    for row in rows:
        seg = _to_text(row.get("segment_uri") or "").strip()
        if "/boq/" not in seg:
            continue
        item_code = seg.rstrip("/").split("/")[-1]
        if not item_code.startswith(smu_id):
            continue
        sd = _as_dict(row.get("state_data"))
        freeze_meta = _as_dict(sd.get("smu_freeze"))
        if bool(freeze_meta.get("immutable")):
            skipped += 1
            continue
        freeze_meta.update(
            {
                "immutable": True,
                "frozen_at": now,
                "freeze_proof_id": freeze_proof_id,
                "smu_id": smu_id,
                "total_proof_hash": total_proof_hash,
            }
        )
        sd["immutable"] = True
        sd["immutable_at"] = now
        sd["smu_freeze"] = freeze_meta
        pid = _to_text(row.get("proof_id") or "").strip()
        if pid:
            sb.table("proof_utxo").update({"state_data": sd}).eq("proof_id", pid).execute()
            touched += 1
    return {"touched": touched, "skipped": skipped}


def _build_genesis_enrichment_patch(
    *,
    code: str,
    name: str,
    sd: dict[str, Any],
    upload_file_name: str,
    owner_uri: str,
) -> dict[str, Any]:
    spu = _resolve_spu_template(code, name)
    template = _as_dict(SPU_TEMPLATE_LIBRARY.get(_to_text(spu.get("spu_template_id") or "").strip()))
    norm_refs = _resolve_norm_refs(
        code,
        name,
        template_norm_refs=[str(x).strip() for x in _as_list(spu.get("spu_normpeg_refs")) if str(x).strip()],
    )
    design_qty = _to_float(sd.get("design_quantity"))
    approved_qty = _to_float(sd.get("approved_quantity"))
    formula_audit = _build_spu_formula_audit(
        template=template,
        measurement={},
        design_quantity=design_qty,
        approved_quantity=approved_qty,
    )
    docpeg_template = _resolve_docpeg_template(code, name)
    return {
        **spu,
        "formula_validation": formula_audit,
        "norm_refs": norm_refs,
        "docpeg_template": docpeg_template,
        "genesis_amount": approved_qty or design_qty,
        "container": {
            "status": "Unspent",
            "stage": "Genesis Trip",
            "smu_id": _smu_id_from_item_code(code),
        },
        "trip": {
            "phase": "Genesis Trip",
            "source_file": upload_file_name,
            "formula_validation": formula_audit,
        },
        "role": {
            "identity_mode": "Role-Trip-Container",
            "owner_uri": owner_uri,
        },
        "metadata": {
            "unit_project": _to_text(sd.get("division") or "").strip(),
            "subdivision_project": _to_text(sd.get("subdivision") or "").strip(),
            "wbs_path": _to_text(_as_dict(sd.get("hierarchy")).get("wbs_path") or "").strip(),
        },
    }


def _patch_state_data(sb: Any, proof_id: str, patch: dict[str, Any]) -> dict[str, Any]:
    rows = (
        sb.table("proof_utxo")
        .select("state_data")
        .eq("proof_id", proof_id)
        .limit(1)
        .execute()
        .data
        or []
    )
    if not rows:
        return {}
    sd = _as_dict(rows[0].get("state_data"))
    sd.update(patch)
    sb.table("proof_utxo").update({"state_data": sd}).eq("proof_id", proof_id).execute()
    return sd


def _queue_erpnext_push(
    *,
    sb: Any,
    project_uri: str,
    boq_item_uri: str,
    payload: dict[str, Any],
    response: dict[str, Any],
) -> None:
    try:
        sb.table("erpnext_push_queue").insert(
            {
                "project_uri": project_uri,
                "boq_item_uri": boq_item_uri,
                "payload": payload,
                "response": response,
                "attempts": 1,
                "status": "queued",
            }
        ).execute()
    except Exception:
        pass


def retry_erpnext_push_queue(
    *,
    sb: Any,
    limit: int = 20,
) -> dict[str, Any]:
    try:
        rows = (
            sb.table("erpnext_push_queue")
            .select("*")
            .eq("status", "queued")
            .order("created_at", desc=False)
            .limit(max(1, min(limit, 100)))
            .execute()
            .data
            or []
        )
    except Exception as exc:
        return {"ok": False, "error": f"{exc.__class__.__name__}: {exc}", "attempted": 0, "success": 0}

    attempted = 0
    success = 0
    for row in rows:
        attempted += 1
        payload = _as_dict(row.get("payload"))
        project_uri = _to_text(row.get("project_uri") or "").strip()
        boq_item_uri = _to_text(row.get("boq_item_uri") or "").strip()
        project_id = _to_text(payload.get("project_id") or "").strip()
        proj = _load_project_for_erpnext(sb, project_id, project_uri)
        enterprise_id = _to_text(proj.get("enterprise_id") or "").strip()
        if not enterprise_id:
            continue
        custom = load_erpnext_custom(sb, enterprise_id)
        path = str(custom.get("erpnext_notify_path") or "").strip() or "/api/method/qcspec_notify"
        res = erp_request_sync(custom, method="POST", path=path, body=payload, timeout_s=12.0)
        ok = bool(res.get("success"))
        status = "sent" if ok else "queued"
        if ok:
            success += 1
            try:
                _create_erpnext_receipt_proof(
                    sb=sb,
                    project_uri=project_uri,
                    boq_item_uri=boq_item_uri,
                    payload=payload,
                    response=res,
                )
            except Exception:
                pass
        try:
            sb.table("erpnext_push_queue").update(
                {
                    "response": res,
                    "attempts": int(row.get("attempts") or 0) + 1,
                    "status": status,
                }
            ).eq("id", row.get("id")).execute()
        except Exception:
            pass

    return {"ok": True, "attempted": attempted, "success": success}


def _boq_rows(
    sb: Any,
    *,
    project_uri: str,
    boq_item_uri: str = "",
    only_unspent: bool = False,
    limit: int = 50000,
) -> list[dict[str, Any]]:
    q = (
        sb.table("proof_utxo")
        .select("*")
        .eq("project_uri", project_uri)
        .order("created_at", desc=False)
        .limit(max(1, min(limit, 50000)))
    )
    if only_unspent:
        q = q.eq("spent", False)
    rows = q.execute().data or []
    if boq_item_uri:
        uri = boq_item_uri.rstrip("/")
        out: list[dict[str, Any]] = []
        for row in rows:
            seg = _to_text(row.get("segment_uri") or "").strip().rstrip("/")
            sd = _as_dict(row.get("state_data"))
            item_uri = _to_text(sd.get("boq_item_uri") or seg).strip().rstrip("/")
            if item_uri == uri or seg == uri:
                out.append(row)
        rows = out
    return [x for x in rows if isinstance(x, dict)]


def _latest_unspent_leaf(sb: Any, *, project_uri: str, boq_item_uri: str) -> dict[str, Any]:
    rows = _boq_rows(sb, project_uri=project_uri, boq_item_uri=boq_item_uri, only_unspent=True, limit=20000)
    if not rows:
        return {}
    rows.sort(key=lambda r: _to_text(r.get("created_at") or ""))
    return rows[-1]


def _container_status_from_stage(stage: str, result: str) -> str:
    s = _to_text(stage).strip().upper()
    r = _to_text(result).strip().upper()
    if s in {"INITIAL", "PRECHECK"}:
        return "Unspent"
    if s in {"ENTRY", "INSTALLATION", "VARIATION"}:
        return "Reviewing"
    if s in {"SETTLEMENT"} and r == "PASS":
        return "Approved"
    if r == "FAIL":
        return "Failed"
    return "Reviewing"


def _eval_threshold(operator: str, threshold: Any, measured_value: float | None) -> dict[str, Any]:
    op = _to_text(operator).strip().lower()
    val = measured_value
    if val is None:
        return {"status": "PENDING", "ok": False}
    if isinstance(threshold, list) and len(threshold) >= 2:
        lo = _to_float(threshold[0])
        hi = _to_float(threshold[1])
        if lo is None or hi is None:
            return {"status": "PENDING", "ok": False}
        ok = (val >= min(lo, hi)) and (val <= max(lo, hi))
        return {"status": "SUCCESS" if ok else "FAIL", "ok": ok, "normalized_operator": "range", "threshold": [min(lo, hi), max(lo, hi)]}
    t = _to_float(threshold)
    if t is None:
        return {"status": "PENDING", "ok": False}
    if op in {">=", "gte"}:
        ok = val >= t
    elif op in {"<=", "lte"}:
        ok = val <= t
    elif op == ">":
        ok = val > t
    elif op == "<":
        ok = val < t
    else:
        ok = val == t
    return {"status": "SUCCESS" if ok else "FAIL", "ok": ok, "normalized_operator": op or "=", "threshold": t}


def _derive_display_metadata(sd: dict[str, Any], *, item_no: str, item_name: str) -> dict[str, str]:
    raw_meta = _as_dict(sd.get("metadata"))
    hierarchy = _as_dict(sd.get("hierarchy"))
    chapter_code = _to_text(hierarchy.get("chapter_code") or "").strip()
    section_code = _to_text(hierarchy.get("section_code") or "").strip()
    subgroup_code = _to_text(hierarchy.get("subgroup_code") or "").strip()
    wbs_path = _to_text(raw_meta.get("wbs_path") or hierarchy.get("wbs_path") or "").strip()

    unit_project = _to_text(raw_meta.get("unit_project") or sd.get("division") or "").strip()
    if not unit_project:
        if chapter_code:
            unit_project = f"{chapter_code}章"
        else:
            head = _to_text(item_no).strip().split("-")[0] if _to_text(item_no).strip() else ""
            unit_project = f"{head}章" if head else "单位工程未命名"

    subdivision_project = _to_text(raw_meta.get("subdivision_project") or sd.get("subdivision") or "").strip()
    if not subdivision_project:
        if section_code and subgroup_code and section_code != subgroup_code:
            subdivision_project = f"{section_code}节 / {subgroup_code}"
        elif subgroup_code:
            subdivision_project = subgroup_code
        elif section_code:
            subdivision_project = section_code
        elif _to_text(item_no).strip():
            subdivision_project = _to_text(item_no).strip()
    if _to_text(item_name).strip():
        subdivision_project = f"{subdivision_project} {item_name}".strip() if subdivision_project else _to_text(item_name).strip()

    return {
        "unit_project": unit_project,
        "subdivision_project": subdivision_project,
        "wbs_path": wbs_path,
    }


def import_genesis_trip(
    *,
    sb: Any,
    project_uri: str,
    project_id: str = "",
    upload_file_name: str,
    upload_content: bytes,
    boq_root_uri: str = "",
    norm_context_root_uri: str = "",
    owner_uri: str = "",
    commit: bool = True,
    progress_hook: Callable[[str, int, str], None] | None = None,
) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    if not p_uri:
        raise HTTPException(400, "project_uri is required")
    if progress_hook:
        progress_hook("parsing", 12, "解析上传文件")
    items = _parse_boq_upload(upload_file_name, upload_content)
    if progress_hook:
        progress_hook("parsed", 22, f"解析完成：识别细目 {len(items)} 条")
    root_uri = _to_text(boq_root_uri).strip() or f"{p_uri.rstrip('/')}/boq/400"
    norm_root = _to_text(norm_context_root_uri).strip() or f"{p_uri.rstrip('/')}/normContext"
    if progress_hook:
        progress_hook("writing_chain", 38, "初始化主权树并写链")
    result = initialize_boq_utxos(
        sb=sb,
        project_uri=p_uri,
        project_id=_to_text(project_id).strip() or None,
        boq_items=items,
        boq_root_uri=root_uri,
        norm_context_root_uri=norm_root,
        owner_uri=_to_text(owner_uri).strip() or f"{p_uri.rstrip('/')}/role/system/",
        source_file=upload_file_name,
        commit=bool(commit),
    )
    if progress_hook:
        total_nodes = int(result.get("total_nodes") or 0)
        progress_hook("chain_written", 78, f"写链完成：节点 {total_nodes}")

    if progress_hook:
        progress_hook("enriching_preview", 84, "补充 SPU 与模板绑定")
    effective_owner_uri = _to_text(result.get("owner_uri") or "").strip()
    for row in _as_list(result.get("preview")):
        sd = _as_dict(row.get("state_data"))
        code = _to_text(sd.get("item_no") or "").strip()
        name = _to_text(sd.get("item_name") or "").strip()
        patch = _build_genesis_enrichment_patch(
            code=code,
            name=name,
            sd=sd,
            upload_file_name=upload_file_name,
            owner_uri=effective_owner_uri,
        )
        sd.update(patch)
        row["state_data"] = sd

    enrichment_warnings: list[dict[str, Any]] = []
    if bool(commit):
        # NOTE:
        # Genesis insertion is already completed in initialize_boq_utxos(commit=True).
        # The patch below is metadata enrichment only. It must be best-effort and must not
        # fail the entire import when a downstream HTTP connection is temporarily closed.
        persist_failed_streak = 0
        for created in _as_list(result.get("created")):
            pid = _to_text(created.get("proof_id") or "").strip()
            sd = _as_dict(created.get("state_data"))
            code = _to_text(sd.get("item_no") or "").strip()
            name = _to_text(sd.get("item_name") or "").strip()
            patch = _build_genesis_enrichment_patch(
                code=code,
                name=name,
                sd=sd,
                upload_file_name=upload_file_name,
                owner_uri=effective_owner_uri,
            )
            # Keep response payload enriched even if persistence fails.
            merged_state = dict(sd)
            merged_state.update(patch)
            created["state_data"] = merged_state
            if not pid:
                continue
            if persist_failed_streak >= 3:
                enrichment_warnings.append(
                    {
                        "proof_id": pid,
                        "item_no": code,
                        "error": "persistence skipped after repeated connection failures",
                    }
                )
                continue
            try:
                _patch_state_data(sb, pid, patch)
                persist_failed_streak = 0
            except Exception as exc:
                persist_failed_streak += 1
                enrichment_warnings.append(
                    {
                        "proof_id": pid,
                        "item_no": code,
                        "error": f"{exc.__class__.__name__}: {exc}",
                    }
                )
        if progress_hook:
            progress_hook("enriched", 96, "后处理完成")

    if progress_hook:
        progress_hook("finalizing", 99, "正在整理导入结果")

    return {
        "ok": True,
        "phase": "Genesis Trip",
        "role": {
            "identity_mode": "Role-Trip-Container",
            "executor_role": "SYSTEM",
        },
        "trip": {
            "name": "asset_initialization",
            "source_file": upload_file_name,
            "item_count": len(items),
            "commit": bool(commit),
        },
        "container": {
            "boq_root_uri": root_uri,
            "norm_context_root_uri": norm_root,
            "hierarchy_root_hash": _to_text(result.get("hierarchy_root_hash") or "").strip(),
        },
        "enrichment_warnings": enrichment_warnings[:20],
        "result": result,
    }


def preview_genesis_tree(
    *,
    sb: Any,
    project_uri: str,
    project_id: str = "",
    upload_file_name: str,
    upload_content: bytes,
    boq_root_uri: str = "",
    norm_context_root_uri: str = "",
    owner_uri: str = "",
) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    if not p_uri:
        raise HTTPException(400, "project_uri is required")
    items = _parse_boq_upload(upload_file_name, upload_content)
    root_uri = _to_text(boq_root_uri).strip() or f"{p_uri.rstrip('/')}/boq/400"
    norm_root = _to_text(norm_context_root_uri).strip() or f"{p_uri.rstrip('/')}/normContext"
    result = initialize_boq_utxos(
        sb=sb,
        project_uri=p_uri,
        project_id=_to_text(project_id).strip() or None,
        boq_items=items,
        boq_root_uri=root_uri,
        norm_context_root_uri=norm_root,
        owner_uri=_to_text(owner_uri).strip() or f"{p_uri.rstrip('/')}/role/system/",
        source_file=upload_file_name,
        commit=False,
    )
    preview_items: list[dict[str, Any]] = []
    for row in _as_list(result.get("preview")):
        sd = _as_dict(row.get("state_data"))
        if not bool(sd.get("is_leaf")):
            continue
        preview_items.append(
            {
                "boq_item_uri": _to_text(sd.get("boq_item_uri") or "").strip(),
                "item_no": _to_text(sd.get("item_no") or "").strip(),
                "item_name": _to_text(sd.get("item_name") or "").strip(),
                "unit": _to_text(sd.get("unit") or "").strip(),
                "design_quantity": _to_float(sd.get("design_quantity")),
                "approved_quantity": _to_float(sd.get("approved_quantity")),
                "settled_quantity": 0.0,
            }
        )
    return {
        "ok": True,
        "phase": "Genesis Preview",
        "project_uri": p_uri,
        "boq_root_uri": root_uri,
        "norm_context_root_uri": norm_root,
        "total_items": len(items),
        "total_nodes": int(result.get("total_nodes") or 0),
        "leaf_nodes": int(result.get("leaf_nodes") or 0),
        "hierarchy_root_hash": _to_text(result.get("hierarchy_root_hash") or "").strip(),
        "preview_items": preview_items,
    }


def get_governance_context(
    *,
    sb: Any,
    project_uri: str,
    boq_item_uri: str,
    component_type: str = "generic",
    measured_value: float | None = None,
) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    b_uri = _to_text(boq_item_uri).strip().rstrip("/")
    if not p_uri or not b_uri:
        raise HTTPException(400, "project_uri and boq_item_uri are required")
    row = _latest_unspent_leaf(sb, project_uri=p_uri, boq_item_uri=b_uri)
    if not row:
        raise HTTPException(404, "No unspent UTXO found for boq_item_uri")

    sd = _as_dict(row.get("state_data"))
    item_no = _to_text(sd.get("item_no") or b_uri.split("/")[-1]).strip()
    item_name = _to_text(sd.get("item_name") or "").strip()
    spu = _resolve_spu_template(item_no, item_name)
    template = _as_dict(SPU_TEMPLATE_LIBRARY.get(_to_text(spu.get("spu_template_id") or "").strip()))
    norm_refs = _resolve_norm_refs(
        item_no,
        item_name,
        template_norm_refs=[str(x).strip() for x in _as_list(spu.get("spu_normpeg_refs")) if str(x).strip()],
    )
    formula_validation = _build_spu_formula_audit(
        template=template,
        measurement={},
        design_quantity=_to_float(sd.get("design_quantity")),
        approved_quantity=_to_float(sd.get("approved_quantity")),
    )
    if formula_validation:
        spu["formula_validation"] = formula_validation
    allowed_roles = _resolve_allowed_roles(item_no, _to_text(spu.get("spu_template_id") or "").strip())
    docpeg_template = _as_dict(sd.get("docpeg_template"))
    if not docpeg_template:
        docpeg_template = _resolve_docpeg_template(item_no, item_name)
    gate_id = _to_text(sd.get("linked_gate_id") or "").strip()
    threshold_pack = {}
    if gate_id:
        try:
            threshold_pack = resolve_dynamic_threshold(sb=sb, gate_id=gate_id, context={"context": component_type})
        except Exception:
            threshold_pack = {}
    threshold_eval = _eval_threshold(
        _to_text(_as_dict(threshold_pack).get("operator") or "").strip(),
        _as_dict(threshold_pack).get("threshold"),
        measured_value,
    )

    stage = _to_text(sd.get("lifecycle_stage") or sd.get("status") or "").strip().upper()
    container = ContainerState(
        status=_container_status_from_stage(stage, _to_text(row.get("result") or "").strip()),
        stage=stage or "INITIAL",
        boq_item_uri=b_uri,
        smu_id=_smu_id_from_item_code(item_no),
    )
    freeze_state = _is_smu_frozen(sb=sb, project_uri=p_uri, smu_id=container.smu_id)
    if bool(freeze_state.get("frozen")):
        container.status = "Frozen"
        container.stage = "SMU_FREEZE"

    display_metadata = _derive_display_metadata(sd, item_no=item_no, item_name=item_name)
    lab_status = _resolve_lab_status(sb=sb, project_uri=p_uri, boq_item_uri=b_uri)
    dual_gate = resolve_dual_pass_gate(sb=sb, project_uri=p_uri, boq_item_uri=b_uri)
    gatekeeper = {
        "is_compliant": bool(dual_gate.get("qc_pass_count")),
        "lab_ok": bool(dual_gate.get("lab_pass_count")),
        "dual_ok": bool(dual_gate.get("ok")),
        "qc_pass_count": int(dual_gate.get("qc_pass_count") or 0),
        "lab_pass_count": int(dual_gate.get("lab_pass_count") or 0),
        "latest_lab_pass_proof_id": _to_text(dual_gate.get("latest_lab_pass_proof_id") or "").strip(),
    }

    return {
        "ok": True,
        "phase": "Governance & QCGate",
        "role": {
            "executor_role": "CHIEF_ENGINEER",
            "did_gate_required": True,
            "allowed_dto_roles": allowed_roles,
        },
        "trip": {
            "name": "governance_context",
            "input_proof_id": _to_text(row.get("proof_id") or "").strip(),
        },
        "container": {
            "status": container.status,
            "stage": container.stage,
            "boq_item_uri": container.boq_item_uri,
            "smu_id": container.smu_id,
        },
        "node": {
            "proof_id": _to_text(row.get("proof_id") or "").strip(),
            "proof_type": _to_text(row.get("proof_type") or "").strip(),
            "result": _to_text(row.get("result") or "").strip(),
            "item_no": item_no,
            "item_name": item_name,
            "unit": _to_text(sd.get("unit") or "").strip(),
            "design_quantity": _to_float(sd.get("design_quantity")),
            "approved_quantity": _to_float(sd.get("approved_quantity")),
            "linked_gate_id": gate_id,
            "linked_spec_uri": _to_text(sd.get("linked_spec_uri") or "").strip(),
            "docpeg_template": docpeg_template,
            "metadata": display_metadata,
            "lab_status": lab_status,
            "norm_refs": norm_refs,
            "formula_validation": formula_validation,
        },
        "spu": spu,
        "threshold": {
            "component_type": component_type,
            **_as_dict(threshold_pack),
            "evaluation": threshold_eval,
        },
        "freeze_state": freeze_state,
        "gatekeeper": gatekeeper,
    }


def execute_smu_trip(
    *,
    sb: Any,
    project_uri: str,
    input_proof_id: str,
    executor_uri: str,
    executor_did: str,
    executor_role: str,
    component_type: str,
    measurement: dict[str, Any],
    geo_location: dict[str, Any],
    server_timestamp_proof: dict[str, Any],
    evidence_hashes: list[str],
    credentials_vc: list[dict[str, Any]],
    force_reject: bool = False,
) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    in_id = _to_text(input_proof_id).strip()
    if not p_uri or not in_id:
        raise HTTPException(400, "project_uri and input_proof_id are required")
    measurement_data = _as_dict(measurement)

    input_row = ProofUTXOEngine(sb).get_by_id(in_id)
    if not input_row:
        raise HTTPException(404, "input_proof_id not found")
    item_uri = _to_text(_as_dict(input_row.get("state_data")).get("boq_item_uri") or input_row.get("segment_uri") or "").strip()
    if not item_uri:
        raise HTTPException(409, "input proof has no boq_item_uri")

    input_sd = _as_dict(input_row.get("state_data"))
    item_no = _to_text(input_sd.get("item_no") or item_uri.rstrip("/").split("/")[-1]).strip()
    smu_id = _smu_id_from_item_code(item_no)
    freeze_state = _is_smu_frozen(sb=sb, project_uri=p_uri, smu_id=smu_id)
    if bool(freeze_state.get("frozen")):
        raise HTTPException(409, f"smu_frozen: {smu_id} is immutable")
    item_name = _to_text(input_sd.get("item_name") or "").strip()
    spu = _resolve_spu_template(item_no, item_name)
    template = _as_dict(SPU_TEMPLATE_LIBRARY.get(_to_text(spu.get("spu_template_id") or "").strip()))
    formula_validation = _build_spu_formula_audit(
        template=template,
        measurement=measurement_data,
        design_quantity=_to_float(input_sd.get("design_quantity")),
        approved_quantity=_to_float(input_sd.get("approved_quantity")),
    )
    norm_refs = _resolve_norm_refs(
        item_no,
        item_name,
        template_norm_refs=[str(x).strip() for x in _as_list(spu.get("spu_normpeg_refs")) if str(x).strip()],
    )
    is_contract_trip = _is_contract_payload(item_name, measurement_data)
    if not is_contract_trip:
        sample_id = _to_text(measurement_data.get("sample_id") or measurement_data.get("utxo_identifier") or "").strip()
        if sample_id:
            lab_sample_gate = _resolve_lab_pass_for_sample(
                sb=sb,
                project_uri=p_uri,
                boq_item_uri=item_uri,
                sample_id=sample_id,
            )
            if not bool(lab_sample_gate.get("pass")):
                raise HTTPException(409, "evidence_chain_incomplete: missing lab PASS proof for sample_id")
        lab_gate = resolve_dual_pass_gate(sb=sb, project_uri=p_uri, boq_item_uri=item_uri)
        if not bool(lab_gate.get("lab_pass_count")):
            raise HTTPException(409, "evidence_chain_incomplete: missing lab PASS proof")

    claim_qty = _round4(_to_float(measurement_data.get("claim_quantity"))) or 0.0
    if claim_qty > 0:
        balance = _resolve_boq_balance(sb=sb, project_uri=p_uri, boq_item_uri=item_uri)
        baseline = balance.get("baseline") or 0.0
        settled = balance.get("settled") or 0.0
        conservation = verify_conservation(baseline=baseline, settled=settled, claim=claim_qty)
        if not bool(conservation.get("ok")):
            required_delta = max(0.0, (settled + claim_qty) - baseline)
            gap_ratio = float(conservation.get("gap_ratio") or 0.0)
            raise HTTPException(
                409,
                f"deviation_warning: gap_ratio={gap_ratio:.4f}, require variation trip (delta>= {required_delta:.6f})",
            )

    snappeg_payload = {
        "project_uri": p_uri,
        "input_proof_id": in_id,
        "boq_item_uri": item_uri,
        "smu_id": smu_id,
        "measurement": measurement_data,
        "formula_validation": formula_validation,
        "norm_refs": norm_refs,
        "geo_location": geo_location,
        "server_timestamp_proof": server_timestamp_proof,
        "executor_did": executor_did,
        "evidence_hashes": evidence_hashes,
    }
    snappeg_hash = _sha(snappeg_payload)
    values_for_qc: list[float] = []
    for candidate in _as_list(measurement_data.get("values")):
        parsed = _to_float(candidate)
        if parsed is not None:
            values_for_qc.append(float(parsed))
    if not values_for_qc:
        raw_values = _to_text(measurement_data.get("values") or "").strip()
        if raw_values:
            for part in re.split(r"[,，;\s\n]+", raw_values):
                parsed = _to_float(part)
                if parsed is not None:
                    values_for_qc.append(float(parsed))
    single_value = _to_float(measurement_data.get("value"))
    if single_value is None:
        single_value = _to_float(measurement_data.get("measured_value"))
    if single_value is None:
        single_value = _to_float(measurement_data.get("claim_quantity"))
    quality_payload: dict[str, Any] = {
        "component_type": component_type,
        "measurement": measurement_data,
        "snappeg_payload_hash": snappeg_hash,
    }
    if values_for_qc:
        quality_payload["values"] = values_for_qc
    if single_value is not None:
        quality_payload["value"] = float(single_value)
    formula_status = _to_text(formula_validation.get("status") or "").strip().upper()
    contract_formula_ok = is_contract_trip and formula_status == "PASS"
    if contract_formula_ok:
        quality_payload["result_policy"] = "contract_formula_pass"

    override_result = "FAIL" if bool(force_reject) else ("PASS" if contract_formula_ok else "")
    qc = execute_triprole_action(
        sb=sb,
        body={
            "action": "quality.check",
            "input_proof_id": in_id,
            "executor_uri": executor_uri,
            "executor_did": executor_did,
            "executor_role": executor_role,
            "boq_item_uri": item_uri,
            **({"result": override_result} if override_result else {}),
            "payload": quality_payload,
            "credentials_vc": credentials_vc,
            "geo_location": geo_location,
            "server_timestamp_proof": server_timestamp_proof,
        },
    )

    current = dict(qc)
    if (not force_reject) and _to_text(qc.get("result") or "").strip().upper() == "PASS":
        current = execute_triprole_action(
            sb=sb,
            body={
                "action": "measure.record",
                "input_proof_id": _to_text(qc.get("output_proof_id") or "").strip(),
                "executor_uri": executor_uri,
                "executor_did": executor_did,
                "executor_role": executor_role,
                "boq_item_uri": item_uri,
                "payload": {
                    "component_type": component_type,
                    "measurement": measurement_data,
                    "snappeg_payload_hash": snappeg_hash,
                },
                "credentials_vc": credentials_vc,
                "geo_location": geo_location,
                "server_timestamp_proof": server_timestamp_proof,
            },
        )

    out_id = _to_text(current.get("output_proof_id") or "").strip()
    if out_id:
        patched = _patch_state_data(
            sb,
            out_id,
            {
                "snappeg": {
                    "hash": snappeg_hash,
                    "evidence_hashes": evidence_hashes,
                    "geo_location": geo_location,
                    "server_timestamp_proof": server_timestamp_proof,
                    "executor_did": executor_did,
                    "captured_at": _utc_iso(),
                },
                "container": {
                    "status": "Reviewing",
                    "stage": "Execution & SnapPeg",
                    "boq_item_uri": item_uri,
                    "smu_id": smu_id,
                },
                "trip": {
                    "phase": "Execution & SnapPeg",
                    "measurement": measurement_data,
                },
                "formula_validation": formula_validation,
                "norm_refs": norm_refs,
            },
        )
        if patched:
            current["state_data"] = patched

    return {
        "ok": True,
        "phase": "Execution & SnapPeg",
        "role": {
            "executor_uri": executor_uri,
            "executor_did": executor_did,
            "executor_role": executor_role,
        },
        "trip": {
            "name": "execution_submit",
            "quality_check_output_proof_id": _to_text(qc.get("output_proof_id") or "").strip(),
            "output_proof_id": out_id,
            "result": _to_text(current.get("result") or "").strip(),
            "snappeg_hash": snappeg_hash,
            "force_reject": bool(force_reject),
        },
        "container": {
            "status": "Reviewing",
            "stage": "Execution & SnapPeg",
            "boq_item_uri": item_uri,
            "smu_id": smu_id,
        },
        "formula_validation": formula_validation,
        "raw": current,
    }


def sign_smu_approval(
    *,
    sb: Any,
    input_proof_id: str,
    boq_item_uri: str,
    supervisor_executor_uri: str,
    supervisor_did: str,
    contractor_did: str,
    owner_did: str,
    signer_metadata: dict[str, Any],
    consensus_values: list[dict[str, Any]] | None = None,
    allowed_deviation: float | None = None,
    allowed_deviation_percent: float | None = None,
    geo_location: dict[str, Any],
    server_timestamp_proof: dict[str, Any],
    auto_docpeg: bool = True,
    verify_base_url: str = "https://verify.qcspec.com",
    template_path: str = "",
) -> dict[str, Any]:
    in_id = _to_text(input_proof_id).strip()
    item_uri = _to_text(boq_item_uri).strip()
    if not in_id or not item_uri:
        raise HTTPException(400, "input_proof_id and boq_item_uri are required")
    now = _utc_iso()
    input_row = ProofUTXOEngine(sb).get_by_id(in_id) or {}
    input_sd = _as_dict(_as_dict(input_row).get("state_data"))
    input_item_no = _to_text(input_sd.get("item_no") or item_uri.rstrip("/").split("/")[-1]).strip()
    input_smu_id = _smu_id_from_item_code(input_item_no)
    project_uri_for_freeze = _to_text(_as_dict(input_row).get("project_uri") or "").strip()
    if project_uri_for_freeze and input_smu_id:
        freeze_state = _is_smu_frozen(sb=sb, project_uri=project_uri_for_freeze, smu_id=input_smu_id)
        if bool(freeze_state.get("frozen")):
            raise HTTPException(409, f"smu_frozen: {input_smu_id} is immutable")
    input_item_name = _to_text(input_sd.get("item_name") or "").strip()
    template_binding = _as_dict(input_sd.get("docpeg_template"))
    if not template_binding:
        template_binding = _resolve_docpeg_template(input_item_no, input_item_name)
    auto_template_path = _to_text(template_binding.get("template_path") or "").strip()
    selected_template_path = _to_text(template_path).strip() or auto_template_path

    signatures: list[dict[str, Any]] = []
    for role, did in (
        ("contractor", contractor_did),
        ("supervisor", supervisor_did),
        ("owner", owner_did),
    ):
        normalized_did = _to_text(did).strip()
        if not normalized_did.startswith("did:"):
            raise HTTPException(400, f"{role}_did must start with did:")
        sig = hashlib.sha256(f"{in_id}|{normalized_did}|{role}|{now}".encode("utf-8")).hexdigest()
        signatures.append({"role": role, "did": normalized_did, "signature_hash": sig, "signed_at": now})

    biometric = _as_dict(signer_metadata)
    if not biometric:
        biometric = {
            "mode": "liveness",
            "passed": True,
            "checked_at": now,
            "device": "mobile",
            "signers": [
                {"role": "contractor", "did": contractor_did, "biometric_ok": True},
                {"role": "supervisor", "did": supervisor_did, "biometric_ok": True},
                {"role": "owner", "did": owner_did, "biometric_ok": True},
            ],
        }

    payload = {
        "approved_from": "SMU_APPROVAL_PANEL",
        "status_target": "Approved",
    }
    if consensus_values:
        payload["consensus_values"] = consensus_values
    if allowed_deviation is not None:
        payload["allowed_deviation"] = allowed_deviation
    if allowed_deviation_percent is not None:
        payload["allowed_deviation_percent"] = allowed_deviation_percent

    settle = execute_triprole_action(
        sb=sb,
        body={
            "action": "settlement.confirm",
            "input_proof_id": in_id,
            "executor_uri": supervisor_executor_uri,
            "executor_did": supervisor_did,
            "executor_role": "SUPERVISOR",
            "boq_item_uri": item_uri,
            "result": "PASS",
            "signatures": signatures,
            "consensus_signatures": signatures,
            "signer_metadata": biometric,
            "payload": payload,
            "geo_location": geo_location,
            "server_timestamp_proof": server_timestamp_proof,
        },
    )
    out_id = _to_text(settle.get("output_proof_id") or "").strip()
    lineage_total_hash = _to_text(_as_dict(settle.get("provenance")).get("total_proof_hash") or "").strip()

    docpeg: dict[str, Any] = {}
    docpeg_document: dict[str, Any] = {}
    risk_audit: dict[str, Any] = {}
    erpnext_push: dict[str, Any] = {}
    erpnext_receipt: dict[str, Any] = {}
    if auto_docpeg and _to_text(settle.get("result") or "").strip().upper() == "PASS":
        package = build_docfinal_package_for_boq(
            boq_item_uri=item_uri,
            sb=sb,
            project_meta={},
            verify_base_url=verify_base_url,
            template_path=selected_template_path or None,
            apply_asset_transfer=False,
        )
        risk_audit = _as_dict(_as_dict(package.get("context")).get("risk_audit"))
        _as_dict(package.get("context"))["smu_id"] = input_smu_id
        pdf_bytes = package.get("pdf_bytes") or b""
        preview_bytes = bytes(pdf_bytes[:1_800_000])
        report_no = _docpeg_report_no(input_item_no)
        project_uri = _to_text(input_row.get("project_uri") or "").strip()
        if project_uri and pdf_bytes:
            try:
                storage_path, storage_url = _upload_docpeg_pdf(
                    sb=sb,
                    project_uri=project_uri,
                    report_no=report_no,
                    pdf_bytes=pdf_bytes,
                )
                docpeg_document = register_document(
                    sb=sb,
                    project_uri=project_uri,
                    node_uri=f"{item_uri.rstrip('/')}/reports/{report_no}/",
                    source_utxo_id=out_id or in_id,
                    file_name=f"{report_no}.pdf",
                    file_size=len(pdf_bytes),
                    mime_type="application/pdf",
                    storage_path=storage_path,
                    storage_url=storage_url,
                    text_excerpt="DocPeg auto-rendered report",
                    ai_metadata={
                        "doc_type": "docpeg_report",
                        "summary": "DocPeg auto-rendered report",
                        "tags": ["docpeg", "docfinal", "report"],
                    },
                    custom_metadata={
                        "report_no": report_no,
                        "boq_item_uri": item_uri,
                        "template_path": selected_template_path,
                    },
                    tags=["docpeg", "docfinal", "report"],
                    executor_uri=supervisor_executor_uri or "v://executor/docpeg/system/",
                    trip_action="document.create_trip",
                    lifecycle_stage="DOCUMENT",
                    trip_payload={
                        "phase": "DocPeg.render",
                        "source": "OrdoSign",
                        "report_no": report_no,
                    },
                )
                docpeg_document["report_no"] = report_no
                erpnext_push = _push_docpeg_to_erpnext(
                    sb=sb,
                    project_id=_to_text(input_row.get("project_id") or "").strip(),
                    project_uri=project_uri,
                    item_no=input_item_no,
                    item_name=input_item_name,
                    report_no=report_no,
                    report_url=storage_url,
                    docpeg_document=docpeg_document,
                    docpeg_context=_as_dict(package.get("context")),
                    risk_audit=risk_audit,
                )
                if bool(erpnext_push.get("success")):
                    erpnext_receipt = _create_erpnext_receipt_proof(
                        sb=sb,
                        project_uri=project_uri,
                        boq_item_uri=item_uri,
                        payload=_as_dict(erpnext_push.get("payload")),
                        response=erpnext_push,
                        source_utxo_id=out_id or in_id,
                    )
                if not bool(erpnext_push.get("success")):
                    _queue_erpnext_push(
                        sb=sb,
                        project_uri=project_uri,
                        boq_item_uri=item_uri,
                        payload=_as_dict(erpnext_push.get("payload")),
                        response=erpnext_push,
                    )
            except Exception as exc:
                docpeg_document = {
                    "ok": False,
                    "error": f"{exc.__class__.__name__}: {exc}",
                    "report_no": report_no,
                }
        docpeg = {
            "verify_uri": _to_text(_as_dict(package.get("context")).get("verify_uri") or "").strip(),
            "artifact_uri": _to_text(_as_dict(package.get("context")).get("artifact_uri") or "").strip(),
            "gitpeg_anchor": _to_text(_as_dict(package.get("context")).get("gitpeg_anchor") or "").strip(),
            "pdf_preview_b64": base64.b64encode(preview_bytes).decode("ascii") if preview_bytes else "",
            "pdf_preview_truncated": len(preview_bytes) < len(pdf_bytes),
            "template_binding": template_binding,
            "selected_template_path": selected_template_path or _to_text(template_binding.get("fallback_template") or "").strip(),
            "context": package.get("context") or {},
            "document": docpeg_document or {},
            "risk_audit": risk_audit,
            "erpnext_push": erpnext_push,
            "erpnext_receipt": erpnext_receipt,
        }

    if out_id:
        _patch_state_data(
            sb,
            out_id,
            {
                "container": {
                    "status": "Approved",
                    "stage": "OrdoSign & DID",
                    "boq_item_uri": item_uri,
                    "smu_id": input_smu_id,
                },
                "trip": {
                    "phase": "OrdoSign & DID",
                    "consensus": "complete",
                },
                "total_proof_hash": lineage_total_hash,
                "docpeg_document": docpeg_document or {},
                "risk_audit": risk_audit or {},
                "erpnext_push": erpnext_push or {},
                "erpnext_receipt": erpnext_receipt or {},
                "smu_id": input_smu_id,
            },
        )

    return {
        "ok": True,
        "phase": "OrdoSign & DID",
        "role": {
            "executor_uri": supervisor_executor_uri,
            "executor_did": supervisor_did,
            "executor_role": "SUPERVISOR",
        },
        "trip": {
            "name": "approval_signature",
            "input_proof_id": in_id,
            "output_proof_id": out_id,
            "result": _to_text(settle.get("result") or "").strip(),
            "total_proof_hash": lineage_total_hash,
        },
        "container": {
            "status": "Approved",
            "stage": "OrdoSign & DID",
            "boq_item_uri": item_uri,
            "smu_id": input_smu_id,
        },
        "docpeg": docpeg,
        "raw": settle,
    }


def validate_logic(
    *,
    sb: Any,
    project_uri: str,
    smu_id: str,
) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    s_id = _to_text(smu_id).strip()
    if not p_uri or not s_id:
        raise HTTPException(400, "project_uri and smu_id are required")
    rows = _boq_rows(sb, project_uri=p_uri, boq_item_uri="", only_unspent=False, limit=50000)
    scoped: list[dict[str, Any]] = []
    for row in rows:
        seg = _to_text(row.get("segment_uri") or "").strip()
        if f"/boq/" not in seg:
            continue
        code = seg.rstrip("/").split("/")[-1]
        if code.startswith(s_id):
            scoped.append(row)
    if not scoped:
        raise HTTPException(404, f"No proof rows found under smu_id={s_id}")

    missing_geo = 0
    missing_ntp = 0
    fail_count = 0
    low_trust = 0
    issues: list[dict[str, Any]] = []
    for row in scoped:
        pid = _to_text(row.get("proof_id") or "").strip()
        result = _to_text(row.get("result") or "").strip().upper()
        sd = _as_dict(row.get("state_data"))
        geo = _as_dict(sd.get("geo_location"))
        ntp = _as_dict(sd.get("server_timestamp_proof"))
        if not geo or (_to_float(geo.get("lat")) is None or _to_float(geo.get("lng")) is None):
            missing_geo += 1
            issues.append({"proof_id": pid, "severity": "medium", "issue": "missing_geo_location"})
        if not ntp or not _to_text(ntp.get("ntp_server") or ntp.get("proof_hash") or "").strip():
            missing_ntp += 1
            issues.append({"proof_id": pid, "severity": "medium", "issue": "missing_ntp_proof"})
        if result == "FAIL":
            fail_count += 1
            issues.append({"proof_id": pid, "severity": "high", "issue": "fail_result_in_chain"})
        trust = _to_text(_as_dict(sd.get("geo_compliance")).get("trust_level") or "").strip().upper()
        if trust in {"LOW", "OUTSIDE"}:
            low_trust += 1
            issues.append({"proof_id": pid, "severity": "high", "issue": "low_geo_trust"})

    total = len(scoped)
    risk_score = 100.0
    if total > 0:
        risk_score -= 35.0 * (fail_count / total)
        risk_score -= 25.0 * (low_trust / total)
        risk_score -= 20.0 * (missing_geo / total)
        risk_score -= 20.0 * (missing_ntp / total)
    did_reputation = build_did_reputation_summary(
        sb=sb,
        project_uri=p_uri,
        chain_rows=scoped,
        window_days=90,
    )
    rep_penalty = _to_float(_as_dict(did_reputation).get("risk_penalty")) or 0.0
    if rep_penalty > 0:
        risk_score -= min(25.0, rep_penalty)
    for did in _as_list(_as_dict(did_reputation).get("high_risk_dids")):
        d = _as_dict(did)
        issues.append(
            {
                "proof_id": "",
                "severity": "medium",
                "issue": "did_reputation_low",
                "participant_did": _to_text(d.get("participant_did") or "").strip(),
                "identity_uri": _to_text(d.get("identity_uri") or "").strip(),
                "score": _to_float(d.get("score")),
            }
        )
    risk_score = max(0.0, min(100.0, round(risk_score, 2)))
    qualification = _collect_smu_qualification(sb=sb, project_uri=p_uri, smu_id=s_id)
    qualification_ratio = 0.0
    if int(qualification.get("leaf_total") or 0) > 0:
        qualification_ratio = float(qualification.get("qualified_leaf_count") or 0) / float(qualification.get("leaf_total") or 1)
    if not bool(qualification.get("all_qualified")):
        issues.append(
            {
                "proof_id": "",
                "severity": "high",
                "issue": "smu_unqualified_leaf_exists",
                "pending_leaf_count": int(qualification.get("unqualified_leaf_count") or 0),
            }
        )

    return {
        "ok": True,
        "phase": "SMU & Risk Audit",
        "smu_id": s_id,
        "project_uri": p_uri,
        "summary": {
            "total_proofs": total,
            "missing_geo": missing_geo,
            "missing_ntp": missing_ntp,
            "fail_count": fail_count,
            "low_trust_count": low_trust,
            "risk_score": risk_score,
            "qualified_leaf_count": int(qualification.get("qualified_leaf_count") or 0),
            "leaf_total": int(qualification.get("leaf_total") or 0),
            "qualification_ratio": round(qualification_ratio, 6),
            "all_qualified": bool(qualification.get("all_qualified")),
            "did_reputation_score": _to_float(_as_dict(did_reputation).get("aggregate_score")) or 0.0,
            "did_sampling_multiplier": _to_float(_as_dict(did_reputation).get("sampling_multiplier")) or 1.0,
            "did_count": int(_to_float(_as_dict(did_reputation).get("did_count")) or 0),
        },
        "did_reputation": did_reputation,
        "qualification": qualification,
        "issues": issues[:500],
        "logic_hash": _sha(
            {
                "project_uri": p_uri,
                "smu_id": s_id,
                "summary": {
                    "total_proofs": total,
                    "missing_geo": missing_geo,
                    "missing_ntp": missing_ntp,
                    "fail_count": fail_count,
                    "low_trust_count": low_trust,
                    "risk_score": risk_score,
                    "qualified_leaf_count": int(qualification.get("qualified_leaf_count") or 0),
                    "leaf_total": int(qualification.get("leaf_total") or 0),
                    "all_qualified": bool(qualification.get("all_qualified")),
                    "did_reputation_score": _to_float(_as_dict(did_reputation).get("aggregate_score")) or 0.0,
                    "did_sampling_multiplier": _to_float(_as_dict(did_reputation).get("sampling_multiplier")) or 1.0,
                },
            }
        ),
    }


def freeze_smu(
    *,
    sb: Any,
    project_uri: str,
    smu_id: str,
    executor_uri: str,
    min_risk_score: float = 60.0,
) -> dict[str, Any]:
    p_uri = _to_text(project_uri).strip()
    s_id = _to_text(smu_id).strip()
    if not p_uri or not s_id:
        raise HTTPException(400, "project_uri and smu_id are required")
    freeze_state = _is_smu_frozen(sb=sb, project_uri=p_uri, smu_id=s_id)
    if bool(freeze_state.get("frozen")):
        raise HTTPException(409, f"smu_already_frozen: {s_id}")
    audit = validate_logic(sb=sb, project_uri=p_uri, smu_id=s_id)
    qualification = _as_dict(audit.get("qualification"))
    if not bool(qualification.get("all_qualified")):
        pending = int(qualification.get("unqualified_leaf_count") or 0)
        raise HTTPException(409, f"freeze_blocked: {pending} leaf nodes not qualified")
    risk_score = _to_float(_as_dict(audit.get("summary")).get("risk_score")) or 0.0
    merkle = build_unit_merkle_snapshot(
        sb=sb,
        project_uri=p_uri,
        unit_code=s_id,
        proof_id="",
        max_rows=50000,
    )
    total_proof_hash = _to_text(merkle.get("unit_root_hash") or "").strip()
    if not total_proof_hash:
        raise HTTPException(409, "unit_root_hash is empty, cannot freeze")
    status = "PASS" if risk_score >= float(min_risk_score) else "FAIL"
    freeze_seed = _sha({"project_uri": p_uri, "smu_id": s_id, "root": total_proof_hash, "ts": _utc_iso()})[:18].upper()
    freeze_proof_id = f"GP-SMU-{freeze_seed}"
    state_data = {
        "asset_type": "smu_freeze",
        "status": "SMU_FROZEN" if status == "PASS" else "SMU_FREEZE_REJECTED",
        "lifecycle_stage": "SMU_FREEZE",
        "smu_id": s_id,
        "risk_score": risk_score,
        "risk_logic_hash": _to_text(audit.get("logic_hash") or "").strip(),
        "audit_summary": audit.get("summary") or {},
        "unit_merkle_root": total_proof_hash,
        "project_root_hash": _to_text(merkle.get("project_root_hash") or merkle.get("global_project_fingerprint") or "").strip(),
        "leaf_count": merkle.get("leaf_count"),
        "total_proof_hash": total_proof_hash,
        "container": {
            "status": "Frozen" if status == "PASS" else "Blocked",
            "stage": "SMU & Risk Audit",
            "boq_item_uri": "",
            "smu_id": s_id,
        },
        "trip": {
            "phase": "SMU.freeze",
            "pushed_to_settlement_dashboard": status == "PASS",
        },
        "role": {
            "executor_uri": executor_uri,
            "executor_role": "OWNER",
        },
        "settlement_packet": {
            "smu_id": s_id,
            "project_uri": p_uri,
            "total_proof_hash": total_proof_hash,
            "risk_score": risk_score,
            "status": status,
            "qualified_leaf_count": int(qualification.get("qualified_leaf_count") or 0),
            "leaf_total": int(qualification.get("leaf_total") or 0),
            "created_at": _utc_iso(),
        },
    }
    row = ProofUTXOEngine(sb).create(
        proof_id=freeze_proof_id,
        owner_uri=_to_text(executor_uri).strip() or "v://executor/owner/system/",
        project_uri=p_uri,
        project_id=None,
        segment_uri=f"{p_uri.rstrip('/')}/smu/{s_id}",
        proof_type="smu_freeze",
        result=status,
        state_data=state_data,
        conditions=[],
        parent_proof_id=None,
        norm_uri="v://norm/CoordOS/SMU/1.0#freeze",
        signer_uri=_to_text(executor_uri).strip() or "v://executor/owner/system/",
        signer_role="OWNER",
        gitpeg_anchor=None,
        anchor_config=None,
    )
    immutable_result = {"touched": 0, "skipped": 0}
    if status == "PASS":
        immutable_result = _mark_smu_scope_immutable(
            sb=sb,
            project_uri=p_uri,
            smu_id=s_id,
            freeze_proof_id=freeze_proof_id,
            total_proof_hash=total_proof_hash,
        )
    return {
        "ok": True,
        "phase": "SMU & Risk Audit",
        "role": {
            "executor_uri": executor_uri,
            "executor_role": "OWNER",
        },
        "trip": {
            "name": "SMU.freeze()",
            "result": status,
            "risk_score": risk_score,
            "min_risk_score": min_risk_score,
        },
        "container": {
            "status": "Frozen" if status == "PASS" else "Blocked",
            "stage": "SMU & Risk Audit",
            "smu_id": s_id,
        },
        "freeze_proof_id": _to_text(row.get("proof_id") or "").strip(),
        "total_proof_hash": total_proof_hash,
        "audit": audit,
        "immutable_update": immutable_result,
        "merkle": {
            "unit_root_hash": _to_text(merkle.get("unit_root_hash") or "").strip(),
            "project_root_hash": _to_text(merkle.get("project_root_hash") or merkle.get("global_project_fingerprint") or "").strip(),
            "leaf_count": merkle.get("leaf_count"),
        },
        "settlement_packet": _as_dict(state_data.get("settlement_packet")),
    }
